"""
Service de gestion du plan comptable SYSCOHADA
Conforme aux normes OHADA révisées 2017
"""
from typing import List, Dict, Optional, Tuple
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.utils import timezone
from decimal import Decimal
import uuid
from datetime import date
from ..models import Company, ChartOfAccounts, JournalEntry, JournalEntryLine, FiscalYear


class PlanComptableService:
    """
    Service de gestion avancée du plan comptable SYSCOHADA
    Performance optimisée et conformité totale aux normes OHADA
    """
    
    # Plan comptable SYSCOHADA standard
    SYSCOHADA_STANDARD = {
        # Classe 1 - Comptes de capitaux
        '10': 'Capital',
        '101': 'Capital social',
        '1011': 'Capital souscrit, appelé, versé',
        '1012': 'Capital souscrit, appelé, non versé',
        '1013': 'Capital souscrit, non appelé',
        '105': 'Primes liées au capital social',
        '106': 'Réserves',
        '1061': 'Réserve légale',
        '1063': 'Réserves statutaires',
        '1068': 'Autres réserves',
        '11': 'Report à nouveau',
        '110': 'Report à nouveau créditeur',
        '119': 'Report à nouveau débiteur',
        '12': 'Résultat de l\'exercice',
        '120': 'Résultat de l\'exercice (bénéfice)',
        '129': 'Résultat de l\'exercice (perte)',
        '13': 'Résultat en instance d\'affectation',
        '14': 'Subventions d\'investissement',
        '15': 'Provisions pour risques et charges',
        '151': 'Provisions pour litiges',
        '155': 'Provisions pour impôts',
        '158': 'Autres provisions pour risques et charges',
        '16': 'Emprunts et dettes assimilées',
        '161': 'Emprunts obligataires',
        '162': 'Emprunts et dettes auprès des établissements de crédit',
        '163': 'Avances reçues de l\'Etat',
        '164': 'Avances reçues et comptes courants bloqués',
        '165': 'Dépôts et cautionnements reçus',
        '166': 'Participation des travailleurs aux bénéfices',
        '167': 'Emprunts et dettes assimilées',
        '168': 'Autres emprunts et dettes assimilées',
        
        # Classe 2 - Comptes d'immobilisations
        '20': 'Charges immobilisées',
        '201': 'Frais d\'établissement',
        '205': 'Frais de recherche et de développement',
        '21': 'Immobilisations incorporelles',
        '211': 'Frais de développement',
        '212': 'Brevets, licences, logiciels',
        '213': 'Fonds commercial',
        '22': 'Terrains',
        '221': 'Terrains agricoles et forestiers',
        '222': 'Terrains nus',
        '223': 'Terrains bâtis',
        '228': 'Autres terrains',
        '23': 'Bâtiments, installations techniques et agencements',
        '231': 'Bâtiments industriels, agricoles et commerciaux',
        '232': 'Bâtiments administratifs et sociaux',
        '233': 'Ouvrages d\'infrastructure',
        '238': 'Autres installations et agencements',
        '24': 'Matériel',
        '241': 'Matériel et outillage industriel et commercial',
        '242': 'Matériel et outillage agricole',
        '243': 'Matériel d\'emballage récupérable',
        '244': 'Matériel et mobilier',
        '245': 'Matériel de transport',
        '246': 'Matériel informatique',
        '247': 'Agencements et aménagements',
        '248': 'Autres matériels',
        
        # Classe 3 - Comptes de stocks
        '31': 'Marchandises',
        '311': 'Marchandises A',
        '312': 'Marchandises B',
        '32': 'Matières premières et fournitures liées',
        '321': 'Matières A',
        '322': 'Matières B',
        '33': 'Autres approvisionnements',
        '331': 'Matières consommables',
        '332': 'Fournitures consommables',
        '333': 'Produits d\'entretien',
        '334': 'Fournitures de bureau',
        '335': 'Emballages',
        '34': 'Produits en cours',
        '341': 'Etudes en cours',
        '342': 'Travaux en cours',
        '35': 'Services en cours',
        '351': 'Etudes en cours',
        '352': 'Prestations de services en cours',
        '36': 'Produits finis',
        '361': 'Produits A',
        '362': 'Produits B',
        '37': 'Produits intermédiaires',
        '371': 'Produits intermédiaires A',
        '372': 'Produits intermédiaires B',
        '38': 'Stocks en cours de route, en dépôt ou en consignation',
        
        # Classe 4 - Comptes de tiers
        '40': 'Fournisseurs et comptes rattachés',
        '401': 'Fournisseurs',
        '4011': 'Fournisseurs - factures non parvenues',
        '4017': 'Fournisseurs - retenues de garantie',
        '402': 'Fournisseurs, effets à payer',
        '408': 'Fournisseurs, factures non parvenues',
        '409': 'Fournisseurs débiteurs',
        '41': 'Clients et comptes rattachés',
        '411': 'Clients',
        '4111': 'Clients - ventes de biens ou de prestations de services',
        '4117': 'Clients - retenues de garantie',
        '412': 'Clients, effets à recevoir',
        '413': 'Clients douteux',
        '414': 'Clients, débiteurs divers et créditeurs divers',
        '418': 'Clients, produits non encore facturés',
        '419': 'Clients créditeurs',
        '42': 'Personnel',
        '421': 'Personnel, avances et acomptes',
        '422': 'Comités d\'entreprise, comités d\'établissement',
        '423': 'Participation des salariés aux résultats',
        '424': 'Participation des salariés aux résultats',
        '425': 'Personnel, avances sur salaires',
        '426': 'Personnel, dépôts',
        '427': 'Personnel, oppositions',
        '428': 'Personnel, charges à payer et produits à recevoir',
        '43': 'Organismes sociaux',
        '431': 'Sécurité sociale',
        '432': 'Caisse de retraite',
        '433': 'Autres organismes sociaux',
        '438': 'Organismes sociaux, charges à payer et produits à recevoir',
        '44': 'Etat et collectivités publiques',
        '441': 'Etat, subventions à recevoir',
        '442': 'Etat, impôts et taxes recouvrables sur des tiers',
        '443': 'Etat, TVA facturée sur ventes',
        '444': 'Etat, TVA due ou crédit de TVA',
        '445': 'Etat, TVA récupérable sur autres biens et services',
        '446': 'Etat, TVA récupérable sur immobilisations',
        '447': 'Etat, autres impôts, taxes et versements assimilés',
        '448': 'Etat, charges à payer et produits à recevoir',
        '45': 'Groupe et associés',
        '451': 'Groupe',
        '455': 'Associés, comptes courants',
        '456': 'Associés, opérations sur le capital',
        '457': 'Associés, dividendes à payer',
        '458': 'Groupe, charges à payer et produits à recevoir',
        '46': 'Débiteurs divers et créditeurs divers',
        '462': 'Créances sur cessions d\'immobilisations',
        '464': 'Dettes sur acquisitions de valeurs mobilières de placement',
        '465': 'Créances sur cessions de valeurs mobilières de placement',
        '467': 'Autres comptes débiteurs ou créditeurs',
        '47': 'Comptes transitoires ou d\'attente',
        '471': 'Comptes d\'attente',
        '472': 'Comptes d\'attente',
        '48': 'Charges et produits constatés d\'avance',
        '481': 'Charges à répartir sur plusieurs exercices',
        '486': 'Charges constatées d\'avance',
        '487': 'Produits constatés d\'avance',
        
        # Classe 5 - Comptes de trésorerie
        '50': 'Titres de placement',
        '501': 'Parts dans des entreprises liées',
        '502': 'Actions',
        '503': 'Obligations',
        '504': 'Bons du Trésor et bons de caisse à court terme',
        '505': 'Bons de caisse à moyen terme',
        '51': 'Valeurs à encaisser',
        '511': 'Effets à recevoir',
        '512': 'Effets à l\'encaissement',
        '513': 'Effets à l\'escompte',
        '52': 'Banques',
        '521': 'Banques locales',
        '522': 'Banques étrangères',
        '524': 'Banques, intérêts courus',
        '53': 'Etablissements financiers et assimilés',
        '531': 'Chèques postaux',
        '532': 'Trésor',
        '533': 'Etablissements financiers',
        '54': 'Instruments de trésorerie',
        '56': 'Banques, crédits de trésorerie et d\'escompte',
        '561': 'Crédits de trésorerie',
        '564': 'Escompte de crédit ordinaire',
        '565': 'Escompte de crédit de campagne',
        '566': 'Intérêts courus sur crédits de trésorerie',
        '57': 'Caisse',
        '571': 'Caisse siège social',
        '572': 'Caisse succursale A',
        '573': 'Caisse succursale B',
        '58': 'Virements internes',
        '580': 'Virements de fonds',
        '581': 'Chèques à encaisser',
        '582': 'Chèques à encaisser sur places',
        '59': 'Dépréciations et provisions pour dépréciation',
        
        # Classe 6 - Comptes de charges
        '60': 'Achats et variations de stocks',
        '601': 'Achats de marchandises',
        '602': 'Achats de matières premières et fournitures liées',
        '603': 'Variations de stocks de biens achetés',
        '604': 'Achats d\'études et prestations de services',
        '605': 'Achats de matériel, équipements et travaux',
        '608': 'Achats d\'emballages',
        '61': 'Transports',
        '611': 'Transports sur achats',
        '612': 'Transports sur ventes',
        '613': 'Transports pour le compte de tiers',
        '614': 'Transports du personnel',
        '618': 'Autres frais de transport',
        '62': 'Services extérieurs A',
        '621': 'Sous-traitance générale',
        '622': 'Locations et charges locatives',
        '623': 'Réparations et entretien',
        '624': 'Etudes, recherches et documentation',
        '625': 'Personnel extérieur à l\'entreprise',
        '626': 'Redevances pour brevets, licences',
        '627': 'Services bancaires',
        '628': 'Cotisations et divers',
        '63': 'Services extérieurs B',
        '631': 'Assurances',
        '632': 'Rémunérations d\'intermédiaires et honoraires',
        '633': 'Publicité, publications, relations publiques',
        '634': 'Transports de biens et transports collectifs du personnel',
        '635': 'Déplacements, missions et réceptions',
        '636': 'Services de télécommunication',
        '637': 'Services financiers',
        '638': 'Autres services extérieurs',
        '64': 'Impôts, taxes et versements assimilés',
        '641': 'Impôts, taxes et versements assimilés sur rémunérations',
        '645': 'Autres impôts, taxes et versements assimilés',
        '646': 'Droits d\'enregistrement',
        '647': 'Pénalités et amendes fiscales',
        '648': 'Autres impôts, taxes et versements assimilés',
        '65': 'Autres charges',
        '651': 'Redevances pour concessions, brevets, licences',
        '652': 'Moins-values des cessions courantes',
        '653': 'Jetons de présence',
        '654': 'Pertes sur créances irrécouvrables',
        '655': 'Quote-part de résultat sur opérations faites en commun',
        '658': 'Charges diverses',
        '66': 'Charges de personnel',
        '661': 'Rémunérations du personnel',
        '662': 'Charges sociales',
        '663': 'Autres charges sociales',
        '664': 'Rémunérations de l\'exploitant individuel',
        '665': 'Charges sociales de l\'exploitant',
        '67': 'Charges financières',
        '671': 'Intérêts des emprunts et dettes',
        '672': 'Intérêts dans loyers de crédit-bail',
        '673': 'Escomptes accordés',
        '674': 'Autres charges financières',
        '675': 'Escomptes accordés',
        '676': 'Différences de change',
        '677': 'Pertes sur cessions de valeurs mobilières de placement',
        '68': 'Dotations aux amortissements',
        '681': 'Dotations aux amortissements des charges immobilisées',
        '682': 'Dotations aux amortissements des immobilisations incorporelles',
        '683': 'Dotations aux amortissements des immobilisations corporelles',
        '69': 'Dotations aux provisions',
        '691': 'Dotations aux provisions d\'exploitation',
        '695': 'Dotations aux provisions financières',
        '697': 'Dotations aux provisions exceptionnelles',
        
        # Classe 7 - Comptes de produits
        '70': 'Ventes',
        '701': 'Ventes de marchandises',
        '702': 'Ventes de produits finis',
        '703': 'Ventes de produits intermédiaires',
        '704': 'Travaux',
        '705': 'Etudes',
        '706': 'Autres prestations de services',
        '707': 'Ventes de marchandises dans l\'Union Européenne',
        '708': 'Produits des activités annexes',
        '71': 'Subventions d\'exploitation',
        '711': 'Subventions d\'exploitation reçues dans l\'exercice',
        '719': 'Reprises sur subventions d\'exploitation',
        '72': 'Production immobilisée',
        '721': 'Immobilisations incorporelles',
        '722': 'Immobilisations corporelles',
        '73': 'Variations de stocks de produits',
        '734': 'Variations de stocks de produits en cours',
        '735': 'Variations de stocks de services en cours',
        '736': 'Variations de stocks de produits finis',
        '737': 'Variations de stocks de produits intermédiaires',
        '74': 'Autres produits',
        '741': 'Produits des cessions courantes d\'immobilisations',
        '742': 'Produits des cessions de créances',
        '743': 'Réductions obtenues des fournisseurs',
        '744': 'Plus-values des cessions courantes',
        '745': 'Quote-part de résultat sur opérations faites en commun',
        '748': 'Autres produits divers',
        '75': 'Produits accessoires',
        '751': 'Redevances pour concessions, brevets, licences',
        '752': 'Revenus des immeubles non affectés à l\'exploitation',
        '753': 'Jetons de présence et rémunérations d\'administrateurs',
        '754': 'Ristournes perçues des coopératives',
        '758': 'Produits divers',
        '76': 'Produits financiers',
        '761': 'Produits de participation',
        '762': 'Autres produits de valeurs mobilières',
        '763': 'Revenus de créances rattachées à des participations',
        '764': 'Revenus de prêts',
        '765': 'Escomptes obtenus',
        '766': 'Gains de change',
        '767': 'Produits nets sur cessions de valeurs mobilières',
        '77': 'Revenus financiers et assimilés',
        '771': 'Intérêts des prêts',
        '772': 'Intérêts des comptes bancaires',
        '773': 'Escomptes obtenus',
        '774': 'Revenus de valeurs mobilières de placement',
        '78': 'Transferts de charges',
        '781': 'Transferts de charges d\'exploitation',
        '786': 'Transferts de charges financières',
        '787': 'Transferts de charges exceptionnelles',
        '79': 'Reprises de provisions et de dépréciations',
        '791': 'Reprises de provisions d\'exploitation',
        '795': 'Reprises de provisions financières',
        '797': 'Reprises de provisions exceptionnelles',
        '798': 'Reprises d\'amortissements',
    }
    
    @staticmethod
    def create_standard_chart(company: Company, system_type: str = 'NORMAL') -> Dict[str, any]:
        """
        Crée le plan comptable SYSCOHADA standard pour une société
        Performance optimisée : traitement par lots
        """
        with transaction.atomic():
            created_accounts = []
            
            # Sélection des comptes selon le système
            accounts_to_create = PlanComptableService._filter_accounts_by_system(system_type)
            
            # Création par lots pour performance
            for code, name in accounts_to_create.items():
                account_data = PlanComptableService._prepare_account_data(code, name, company)
                
                account = ChartOfAccounts(
                    company=company,
                    code=code,
                    name=name,
                    account_class=account_data['class'],
                    account_type=account_data['type'],
                    normal_balance=account_data['balance'],
                    is_reconcilable=account_data['reconcilable'],
                    allow_direct_entry=account_data['direct_entry'],
                )
                created_accounts.append(account)
            
            # Insertion en lot - Performance optimisée
            ChartOfAccounts.objects.bulk_create(created_accounts, batch_size=100)
            
            # Mise à jour des relations hiérarchiques
            PlanComptableService._setup_account_hierarchy(company)
            
        return {
            'success': True,
            'created_count': len(created_accounts),
            'system_type': system_type,
            'message': f'Plan comptable SYSCOHADA {system_type} créé avec succès'
        }
    
    @staticmethod
    def _filter_accounts_by_system(system_type: str) -> Dict[str, str]:
        """Filtre les comptes selon le système SYSCOHADA"""
        all_accounts = PlanComptableService.SYSCOHADA_STANDARD
        
        if system_type == 'MINIMAL':
            # Système minimal de trésorerie - comptes essentiels uniquement
            essential_prefixes = ['10', '11', '12', '40', '41', '52', '57', '60', '61', '66', '70']
            return {code: name for code, name in all_accounts.items() 
                   if any(code.startswith(prefix) for prefix in essential_prefixes)}
        
        elif system_type == 'ALLEGE':
            # Système allégé - comptes principaux
            return {code: name for code, name in all_accounts.items() if len(code) <= 3}
        
        else:  # NORMAL
            return all_accounts
    
    @staticmethod
    def _prepare_account_data(code: str, name: str, company: Company) -> Dict[str, any]:
        """Prépare les données métadonnées pour un compte"""
        class_code = code[0]
        
        # Détermination du sens normal selon SYSCOHADA
        if class_code in ['1', '2', '3', '6']:
            normal_balance = 'DEBIT'
        else:  # Classes 4, 5, 7
            normal_balance = 'CREDIT'
        
        # Comptes lettrables (tiers principalement)
        is_reconcilable = class_code == '4' or code.startswith(('401', '411', '421'))
        
        # Type de compte
        if len(code) == 1:
            account_type = 'TOTAL'
        elif len(code) <= 2:
            account_type = 'TOTAL'
        else:
            account_type = 'DETAIL'
        
        # Saisie directe autorisée (pas sur les comptes de regroupement)
        allow_direct_entry = account_type == 'DETAIL'
        
        return {
            'class': class_code,
            'type': account_type,
            'balance': normal_balance,
            'reconcilable': is_reconcilable,
            'direct_entry': allow_direct_entry,
        }
    
    @staticmethod
    def _setup_account_hierarchy(company: Company):
        """Etablit la hiérarchie des comptes - optimisé"""
        accounts = ChartOfAccounts.objects.filter(company=company).order_by('code')
        
        for account in accounts:
            parent_code = account.code[:-1]  # Parent = code sans dernier chiffre
            
            if parent_code and len(parent_code) >= 1:
                try:
                    parent = ChartOfAccounts.objects.get(
                        company=company, 
                        code=parent_code
                    )
                    account.parent_account = parent
                    account.level = len(account.code)
                    account.save(update_fields=['parent_account', 'level'])
                except ChartOfAccounts.DoesNotExist:
                    pass  # Pas de parent trouvé
    
    @staticmethod
    def get_account_balance(account: ChartOfAccounts, 
                           fiscal_year: Optional[FiscalYear] = None,
                           as_of_date: Optional[date] = None) -> Decimal:
        """
        Calcule le solde d'un compte avec performance optimisée
        Cache automatique et requêtes optimisées
        """
        from django.db.models import Sum, Q
        
        # Clé de cache unique
        cache_key = f"balance_{account.id}_{fiscal_year.id if fiscal_year else 'all'}_{as_of_date or 'current'}"
        
        # Vérification cache (5 minutes)
        cached_balance = cache.get(cache_key)
        if cached_balance is not None:
            return cached_balance
        
        # Construction de la requête optimisée
        query = Q(account=account)
        
        if fiscal_year:
            query &= Q(entry__fiscal_year=fiscal_year)
        
        if as_of_date:
            query &= Q(entry__entry_date__lte=as_of_date)
        
        # Requête agrégée optimisée - une seule requête DB
        aggregates = JournalEntryLine.objects.filter(query).aggregate(
            total_debit=Sum('debit_amount') or Decimal('0'),
            total_credit=Sum('credit_amount') or Decimal('0')
        )
        
        balance = aggregates['total_debit'] - aggregates['total_credit']
        
        # Ajustement selon le sens normal du compte
        if account.normal_balance == 'CREDIT':
            balance = -balance
        
        # Mise en cache (5 minutes)
        cache.set(cache_key, balance, 300)
        
        return balance
    
    @staticmethod
    def search_accounts(company: Company, 
                       query: str, 
                       limit: int = 20) -> List[ChartOfAccounts]:
        """
        Recherche rapide et intelligente dans le plan comptable
        Full-text search optimisé
        """
        from django.db.models import Q
        
        if not query or len(query) < 2:
            return []
        
        # Recherche multi-critères optimisée
        search_query = Q(company=company, is_active=True)
        
        # Recherche par code (priorité 1)
        if query.isdigit():
            search_query &= Q(code__startswith=query)
        else:
            # Recherche par nom (full-text)
            search_query &= Q(name__icontains=query)
        
        # Requête avec optimisations
        results = ChartOfAccounts.objects.filter(search_query).select_related(
            'parent_account'
        ).order_by(
            'code'
        )[:limit]
        
        return list(results)
    
    @staticmethod
    def get_balance_sheet_structure(company: Company, 
                                   fiscal_year: Optional[FiscalYear] = None) -> Dict[str, any]:
        """
        Génère la structure du bilan SYSCOHADA avec soldes
        Performance ultra-optimisée avec requêtes batch
        """
        # Structure bilan SYSCOHADA
        balance_sheet = {
            'actif': {
                'actif_immobilise': {
                    'immobilisations_incorporelles': [],
                    'immobilisations_corporelles': [],
                    'immobilisations_financieres': []
                },
                'actif_circulant': {
                    'stocks': [],
                    'creances': [],
                    'tresorerie_actif': []
                }
            },
            'passif': {
                'capitaux_propres': {
                    'capital': [],
                    'reserves': [],
                    'resultat': []
                },
                'passif_circulant': {
                    'dettes_fournisseurs': [],
                    'autres_dettes': [],
                    'tresorerie_passif': []
                }
            }
        }
        
        # Récupération optimisée des comptes avec soldes
        accounts = ChartOfAccounts.objects.filter(
            company=company,
            is_active=True,
            account_type='DETAIL'
        ).prefetch_related('entry_lines')
        
        for account in accounts:
            balance = PlanComptableService.get_account_balance(account, fiscal_year)
            
            if balance != 0:  # N'afficher que les comptes avec solde
                account_data = {
                    'code': account.code,
                    'name': account.name,
                    'balance': float(balance),
                    'balance_formatted': f"{balance:,.0f}"
                }
                
                # Classification selon SYSCOHADA
                PlanComptableService._classify_account_for_balance_sheet(
                    account, account_data, balance_sheet
                )
        
        return balance_sheet
    
    @staticmethod
    def _classify_account_for_balance_sheet(account, account_data, balance_sheet):
        """Classe un compte dans la structure du bilan"""
        code = account.code
        
        # Classification ACTIF
        if code.startswith('2'):  # Immobilisations
            if code.startswith('21'):
                balance_sheet['actif']['actif_immobilise']['immobilisations_incorporelles'].append(account_data)
            elif code.startswith(('22', '23', '24')):
                balance_sheet['actif']['actif_immobilise']['immobilisations_corporelles'].append(account_data)
            else:
                balance_sheet['actif']['actif_immobilise']['immobilisations_financieres'].append(account_data)
                
        elif code.startswith('3'):  # Stocks
            balance_sheet['actif']['actif_circulant']['stocks'].append(account_data)
            
        elif code.startswith('4'):  # Tiers - Créances
            if code.startswith('41'):  # Clients
                balance_sheet['actif']['actif_circulant']['creances'].append(account_data)
                
        elif code.startswith('5'):  # Trésorerie
            balance_sheet['actif']['actif_circulant']['tresorerie_actif'].append(account_data)
        
        # Classification PASSIF
        elif code.startswith('1'):  # Capitaux
            if code.startswith('10'):
                balance_sheet['passif']['capitaux_propres']['capital'].append(account_data)
            elif code.startswith('11'):
                balance_sheet['passif']['capitaux_propres']['reserves'].append(account_data)
            elif code.startswith('12'):
                balance_sheet['passif']['capitaux_propres']['resultat'].append(account_data)
                
        elif code.startswith('4'):  # Tiers - Dettes
            if code.startswith('40'):
                balance_sheet['passif']['passif_circulant']['dettes_fournisseurs'].append(account_data)
            else:
                balance_sheet['passif']['passif_circulant']['autres_dettes'].append(account_data)
    
    @staticmethod
    def validate_account_code(code: str, company: Company) -> Tuple[bool, str]:
        """
        Validation avancée d'un code compte SYSCOHADA
        Contrôles de conformité complets
        """
        if not code:
            return False, "Le code compte est obligatoire"
        
        # Longueur selon paramétrage société
        if len(code) > company.account_length:
            return False, f"Code trop long (max {company.account_length} caractères)"
        
        # Format numérique uniquement
        if not code.isdigit():
            return False, "Le code doit contenir uniquement des chiffres"
        
        # Classe SYSCOHADA valide (1-9)
        if code[0] not in '123456789':
            return False, "La classe doit être entre 1 et 9"
        
        # Unicité
        if ChartOfAccounts.objects.filter(company=company, code=code).exists():
            return False, "Ce code compte existe déjà"
        
        return True, "Code valide"
    
    @staticmethod
    def export_chart_to_excel(company: Company) -> bytes:
        """
        Export du plan comptable vers Excel
        Format professionnel avec mise en forme
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan Comptable SYSCOHADA"
        
        # En-têtes
        headers = ['Code', 'Libellé', 'Classe', 'Type', 'Sens Normal', 'Lettrable', 'Actif']
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        accounts = ChartOfAccounts.objects.filter(
            company=company
        ).order_by('code')
        
        for row, account in enumerate(accounts, 2):
            ws.cell(row=row, column=1, value=account.code)
            ws.cell(row=row, column=2, value=account.name)
            ws.cell(row=row, column=3, value=account.account_class)
            ws.cell(row=row, column=4, value=account.account_type)
            ws.cell(row=row, column=5, value=account.normal_balance)
            ws.cell(row=row, column=6, value='Oui' if account.is_reconcilable else 'Non')
            ws.cell(row=row, column=7, value='Oui' if account.is_active else 'Non')
        
        # Ajustement automatique des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarde en bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    @staticmethod
    def get_performance_stats(company: Company) -> Dict[str, any]:
        """
        Statistiques de performance du plan comptable
        Métriques utiles pour le monitoring
        """
        from django.db.models import Count, Sum
        
        stats = {
            'total_accounts': ChartOfAccounts.objects.filter(company=company).count(),
            'active_accounts': ChartOfAccounts.objects.filter(company=company, is_active=True).count(),
            'accounts_by_class': {},
            'reconcilable_accounts': ChartOfAccounts.objects.filter(
                company=company, is_reconcilable=True
            ).count(),
            'total_entries': JournalEntry.objects.filter(company=company).count(),
            'total_lines': JournalEntryLine.objects.filter(
                entry__company=company
            ).count(),
        }
        
        # Répartition par classe
        for i in range(1, 10):
            count = ChartOfAccounts.objects.filter(
                company=company, 
                account_class=str(i)
            ).count()
            if count > 0:
                stats['accounts_by_class'][f'classe_{i}'] = count
        
        return stats