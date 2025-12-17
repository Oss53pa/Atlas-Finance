"""
Service d'export des données du dashboard
Génère des exports en PDF, Excel et CSV
"""
import os
import csv
from io import StringIO
from decimal import Decimal
from django.core.files.base import ContentFile
from django.utils import timezone
from apps.dashboard.models import DashboardExport
from .financial_calculator import FinancialCalculator


class DashboardExportService:
    """
    Service pour générer des exports de données du dashboard
    """

    def __init__(self, company=None, user=None):
        self.company = company
        self.user = user
        self.calculator = FinancialCalculator(company=company)

    def create_export(self, export_type, file_name, filters=None):
        """
        Crée un nouvel export

        Args:
            export_type (str): 'pdf', 'excel' ou 'csv'
            file_name (str): Nom du fichier
            filters (dict): Filtres appliqués

        Returns:
            DashboardExport: Instance créée
        """
        export = DashboardExport.objects.create(
            user=self.user,
            company=self.company,
            export_type=export_type,
            file_name=file_name,
            filters=filters or {},
            status='pending'
        )
        return export

    def generate_csv_export(self, export_id):
        """
        Génère un export CSV

        Args:
            export_id: ID de l'export

        Returns:
            bool: True si succès
        """
        try:
            export = DashboardExport.objects.get(id=export_id)
            export.status = 'processing'
            export.save()

            # Récupérer les données
            financial_metrics = self.calculator.calculate_all_metrics()
            operational_metrics = self.calculator.calculate_operational_metrics()

            # Créer le CSV en mémoire
            output = StringIO()
            writer = csv.writer(output)

            # En-têtes
            writer.writerow(['Dashboard KPIs - WiseBook'])
            writer.writerow([f'Date: {timezone.now().strftime("%Y-%m-%d %H:%M")}'])
            writer.writerow([f'Société: {self.company.name if self.company else "Toutes"}'])
            writer.writerow([])

            # Métriques financières
            writer.writerow(['MÉTRIQUES FINANCIÈRES'])
            writer.writerow(['Indicateur', 'Valeur'])
            writer.writerow(['Total Actif', f'{financial_metrics["total_assets"]:.2f}'])
            writer.writerow(['Total Passif', f'{financial_metrics["total_liabilities"]:.2f}'])
            writer.writerow(['Capitaux Propres', f'{financial_metrics["equity"]:.2f}'])
            writer.writerow(['Chiffre d\'Affaires', f'{financial_metrics["revenue"]:.2f}'])
            writer.writerow(['Résultat Net', f'{financial_metrics["net_income"]:.2f}'])
            writer.writerow(['EBITDA', f'{financial_metrics["ebitda"]:.2f}'])
            writer.writerow(['Position Trésorerie', f'{financial_metrics["cash_position"]:.2f}'])
            writer.writerow(['Fonds de Roulement', f'{financial_metrics["working_capital"]:.2f}'])
            writer.writerow([])

            # Ratios financiers
            writer.writerow(['RATIOS FINANCIERS'])
            writer.writerow(['Ratio', 'Valeur'])
            writer.writerow(['Ratio de Liquidité Immédiate', f'{financial_metrics["quick_ratio"]:.2f}'])
            writer.writerow(['Ratio de Liquidité Générale', f'{financial_metrics["current_ratio"]:.2f}'])
            writer.writerow(['Dette / Capitaux Propres', f'{financial_metrics["debt_to_equity"]:.2f}'])
            writer.writerow(['ROE (Return on Equity)', f'{financial_metrics["roe"]:.2f}%'])
            writer.writerow(['ROA (Return on Assets)', f'{financial_metrics["roa"]:.2f}%'])
            writer.writerow([])

            # Métriques opérationnelles
            writer.writerow(['MÉTRIQUES OPÉRATIONNELLES'])
            writer.writerow(['Indicateur', 'Valeur'])
            writer.writerow(['Taux de Réalisation Commandes', f'{operational_metrics["order_fulfillment_rate"]:.1f}%'])
            writer.writerow(['Rotation des Stocks', f'{operational_metrics["inventory_turnover"]:.1f}'])
            writer.writerow(['Taux de Productivité', f'{operational_metrics["productivity_rate"]:.1f}%'])
            writer.writerow(['Satisfaction Client', f'{operational_metrics["customer_satisfaction"]:.1f}/5'])
            writer.writerow(['Disponibilité Système', f'{operational_metrics["system_uptime"]:.1f}%'])

            # Sauvegarder le fichier
            csv_content = output.getvalue()
            file_name = f'dashboard_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
            file_content = ContentFile(csv_content.encode('utf-8-sig'))  # BOM pour Excel

            export.file.save(file_name, file_content)
            export.file_size = len(csv_content.encode('utf-8-sig'))
            export.status = 'completed'
            export.save()

            return True

        except Exception as e:
            export.mark_as_failed(str(e))
            return False

    def generate_excel_export(self, export_id):
        """
        Génère un export Excel

        Args:
            export_id: ID de l'export

        Returns:
            bool: True si succès
        """
        try:
            # Vérifier si openpyxl est disponible
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                raise Exception("openpyxl n'est pas installé. Utilisez: pip install openpyxl")

            export = DashboardExport.objects.get(id=export_id)
            export.status = 'processing'
            export.save()

            # Récupérer les données
            financial_metrics = self.calculator.calculate_all_metrics()
            operational_metrics = self.calculator.calculate_operational_metrics()

            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Dashboard KPIs"

            # Styles
            header_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # En-tête
            ws['A1'] = 'Dashboard KPIs - WiseBook'
            ws['A1'].font = header_font
            ws['A2'] = f'Date: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A3'] = f'Société: {self.company.name if self.company else "Toutes"}'

            row = 5

            # Métriques financières
            ws[f'A{row}'] = 'MÉTRIQUES FINANCIÈRES'
            ws[f'A{row}'].font = section_font
            row += 1

            ws[f'A{row}'] = 'Indicateur'
            ws[f'B{row}'] = 'Valeur'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            metrics_data = [
                ('Total Actif', financial_metrics['total_assets']),
                ('Total Passif', financial_metrics['total_liabilities']),
                ('Capitaux Propres', financial_metrics['equity']),
                ('Chiffre d\'Affaires', financial_metrics['revenue']),
                ('Résultat Net', financial_metrics['net_income']),
                ('EBITDA', financial_metrics['ebitda']),
                ('Position Trésorerie', financial_metrics['cash_position']),
                ('Fonds de Roulement', financial_metrics['working_capital']),
            ]

            for metric_name, metric_value in metrics_data:
                ws[f'A{row}'] = metric_name
                ws[f'B{row}'] = float(metric_value)
                ws[f'B{row}'].number_format = '#,##0.00'
                row += 1

            row += 1

            # Ratios financiers
            ws[f'A{row}'] = 'RATIOS FINANCIERS'
            ws[f'A{row}'].font = section_font
            row += 1

            ws[f'A{row}'] = 'Ratio'
            ws[f'B{row}'] = 'Valeur'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            ratios_data = [
                ('Ratio de Liquidité Immédiate', financial_metrics['quick_ratio']),
                ('Ratio de Liquidité Générale', financial_metrics['current_ratio']),
                ('Dette / Capitaux Propres', financial_metrics['debt_to_equity']),
                ('ROE (Return on Equity)', financial_metrics['roe']),
                ('ROA (Return on Assets)', financial_metrics['roa']),
            ]

            for ratio_name, ratio_value in ratios_data:
                ws[f'A{row}'] = ratio_name
                ws[f'B{row}'] = ratio_value
                ws[f'B{row}'].number_format = '0.00'
                row += 1

            # Ajuster les colonnes
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20

            # Sauvegarder
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            file_name = f'dashboard_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            export.file.save(file_name, ContentFile(output.read()))
            export.file_size = output.tell()
            export.status = 'completed'
            export.save()

            return True

        except Exception as e:
            export.mark_as_failed(str(e))
            return False

    def generate_pdf_export(self, export_id):
        """
        Génère un export PDF

        Args:
            export_id: ID de l'export

        Returns:
            bool: True si succès
        """
        try:
            # Vérifier si reportlab est disponible
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib import colors
            except ImportError:
                raise Exception("reportlab n'est pas installé. Utilisez: pip install reportlab")

            export = DashboardExport.objects.get(id=export_id)
            export.status = 'processing'
            export.save()

            # Récupérer les données
            financial_metrics = self.calculator.calculate_all_metrics()
            operational_metrics = self.calculator.calculate_operational_metrics()

            # Créer le PDF
            from io import BytesIO
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []

            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30
            )

            # Titre
            title = Paragraph('Dashboard KPIs - WiseBook', title_style)
            elements.append(title)

            # Informations
            info_text = f"""
            <b>Date:</b> {timezone.now().strftime("%Y-%m-%d %H:%M")}<br/>
            <b>Société:</b> {self.company.name if self.company else "Toutes"}
            """
            elements.append(Paragraph(info_text, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))

            # Section: Métriques financières
            elements.append(Paragraph('<b>MÉTRIQUES FINANCIÈRES</b>', styles['Heading2']))
            elements.append(Spacer(1, 0.2*inch))

            financial_data = [
                ['Indicateur', 'Valeur'],
                ['Total Actif', f'{financial_metrics["total_assets"]:,.2f}'],
                ['Total Passif', f'{financial_metrics["total_liabilities"]:,.2f}'],
                ['Capitaux Propres', f'{financial_metrics["equity"]:,.2f}'],
                ['Chiffre d\'Affaires', f'{financial_metrics["revenue"]:,.2f}'],
                ['Résultat Net', f'{financial_metrics["net_income"]:,.2f}'],
                ['EBITDA', f'{financial_metrics["ebitda"]:,.2f}'],
                ['Position Trésorerie', f'{financial_metrics["cash_position"]:,.2f}'],
                ['Fonds de Roulement', f'{financial_metrics["working_capital"]:,.2f}'],
            ]

            financial_table = Table(financial_data, colWidths=[4*inch, 2*inch])
            financial_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(financial_table)
            elements.append(Spacer(1, 0.3*inch))

            # Section: Ratios financiers
            elements.append(Paragraph('<b>RATIOS FINANCIERS</b>', styles['Heading2']))
            elements.append(Spacer(1, 0.2*inch))

            ratios_data = [
                ['Ratio', 'Valeur'],
                ['Ratio de Liquidité Immédiate', f'{financial_metrics["quick_ratio"]:.2f}'],
                ['Ratio de Liquidité Générale', f'{financial_metrics["current_ratio"]:.2f}'],
                ['Dette / Capitaux Propres', f'{financial_metrics["debt_to_equity"]:.2f}'],
                ['ROE (Return on Equity)', f'{financial_metrics["roe"]:.2f}%'],
                ['ROA (Return on Assets)', f'{financial_metrics["roa"]:.2f}%'],
            ]

            ratios_table = Table(ratios_data, colWidths=[4*inch, 2*inch])
            ratios_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(ratios_table)

            # Construire le PDF
            doc.build(elements)
            pdf_content = buffer.getvalue()
            buffer.close()

            # Sauvegarder
            file_name = f'dashboard_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            export.file.save(file_name, ContentFile(pdf_content))
            export.file_size = len(pdf_content)
            export.status = 'completed'
            export.save()

            return True

        except Exception as e:
            export.mark_as_failed(str(e))
            return False

    def generate_export(self, export_id):
        """
        Génère un export selon son type

        Args:
            export_id: ID de l'export

        Returns:
            bool: True si succès
        """
        export = DashboardExport.objects.get(id=export_id)

        if export.export_type == 'csv':
            return self.generate_csv_export(export_id)
        elif export.export_type == 'excel':
            return self.generate_excel_export(export_id)
        elif export.export_type == 'pdf':
            return self.generate_pdf_export(export_id)
        else:
            export.mark_as_failed(f'Type d\'export non supporté: {export.export_type}')
            return False
