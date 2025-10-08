"""
Service Import et Migration Intelligent WiseBook
Support migration Sage, SAP, Excel avec IA et validation automatique
"""
from typing import List, Dict, Optional, Any, Union
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import UploadedFile
from django.utils import timezone
from decimal import Decimal
import pandas as pd
import openpyxl
import csv
import xml.etree.ElementTree as ET
import re
import hashlib
import logging
from datetime import datetime, date

from ..models import (
    ImportSession, DataSourceConfiguration, ImportMapping, 
    AccountMappingRule, ImportValidationError, MigrationTemplate,
    ImportPreview, ImportResult
)
from apps.accounting.models import Company, ChartOfAccounts, JournalEntry, Journal
from apps.customers.models import Customer
from apps.suppliers.models import Supplier


logger = logging.getLogger(__name__)


class ImportService:
    """
    Service principal d'import avec intelligence artificielle
    Support multi-format et mapping automatique
    """
    
    @staticmethod
    def create_import_session(
        company: Company,
        session_data: Dict[str, Any],
        uploaded_file: UploadedFile,
        template_id: Optional[str] = None,
        user=None
    ) -> ImportSession:
        """
        Crée une nouvelle session d'import avec analyse automatique
        """
        with transaction.atomic():
            # Calcul empreinte fichier
            file_checksum = ImportService._calculate_file_checksum(uploaded_file)
            
            # Détection automatique du format si pas spécifié
            if not session_data.get('source_type'):
                session_data['source_type'] = ImportService._detect_source_type(uploaded_file)
            
            # Création session
            session = ImportSession.objects.create(
                company=company,
                session_name=session_data['session_name'],
                data_type=session_data['data_type'],
                source_file=uploaded_file,
                file_size=uploaded_file.size,
                file_checksum=file_checksum,
                created_by=user,
                status='UPLOADING'
            )
            
            # Application template si fourni
            if template_id:
                template = MigrationTemplate.objects.get(id=template_id)
                ImportService._apply_migration_template(session, template)
            
            # Analyse automatique du fichier
            ImportService._analyze_file_structure(session)
            
            session.status = 'PARSING'
            session.save()
            
            logger.info(f"Session import créée: {session.session_name} - {uploaded_file.name}")
            
            return session
    
    @staticmethod
    def _calculate_file_checksum(file: UploadedFile) -> str:
        """Calcule l'empreinte MD5 du fichier"""
        file.seek(0)
        hash_md5 = hashlib.md5()
        for chunk in file.chunks():
            hash_md5.update(chunk)
        file.seek(0)
        return hash_md5.hexdigest()
    
    @staticmethod
    def _detect_source_type(file: UploadedFile) -> str:
        """Détection automatique du type de source"""
        filename = file.name.lower()
        
        # Détection par extension
        if filename.endswith(('.xlsx', '.xls')):
            return 'EXCEL'
        elif filename.endswith('.csv'):
            return 'CSV'
        elif filename.endswith('.xml'):
            return 'XML'
        elif filename.endswith('.txt'):
            # Analyse contenu pour déterminer si FEC ou autre
            file.seek(0)
            sample = file.read(1024).decode('utf-8', errors='ignore')
            file.seek(0)
            
            if 'JournalCode' in sample and 'EcritureNum' in sample:
                return 'FEC'
            return 'TXT'
        
        return 'OTHER'
    
    @staticmethod
    def _analyze_file_structure(session: ImportSession):
        """
        Analyse intelligente de la structure du fichier
        """
        try:
            file_path = session.source_file.path
            detected_structure = {}
            
            if session.source_file.name.endswith(('.xlsx', '.xls')):
                # Analyse Excel
                detected_structure = ImportService._analyze_excel_structure(file_path)
            elif session.source_file.name.endswith('.csv'):
                # Analyse CSV
                detected_structure = ImportService._analyze_csv_structure(file_path)
            elif session.source_file.name.endswith('.xml'):
                # Analyse XML
                detected_structure = ImportService._analyze_xml_structure(file_path)
            
            # Détection automatique du type de données
            data_type = ImportService._detect_data_type(detected_structure)
            
            # Mise à jour session
            session.detected_structure = detected_structure
            session.total_rows_detected = detected_structure.get('total_rows', 0)
            if not session.data_type:
                session.data_type = data_type
            session.save()
            
        except Exception as e:
            logger.error(f"Erreur analyse structure fichier {session.id}: {e}")
            session.status = 'FAILED'
            session.save()
    
    @staticmethod
    def _analyze_excel_structure(file_path: str) -> Dict[str, Any]:
        """Analyse structure fichier Excel"""
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        structure = {
            'format': 'EXCEL',
            'worksheets': [],
            'total_rows': 0,
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Analyse en-têtes (première ligne)
            headers = []
            for col in range(1, min(max_col + 1, 50)):  # Limite à 50 colonnes
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Colonne_{col}")
            
            # Échantillon de données (lignes 2-11)
            sample_data = []
            for row in range(2, min(max_row + 1, 12)):
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                sample_data.append(row_data)
            
            structure['worksheets'].append({
                'name': sheet_name,
                'max_row': max_row,
                'max_col': max_col,
                'headers': headers,
                'sample_data': sample_data[:10],  # 10 lignes max
            })
            
            structure['total_rows'] += max_row - 1  # -1 pour en-tête
        
        workbook.close()
        return structure
    
    @staticmethod
    def _analyze_csv_structure(file_path: str) -> Dict[str, Any]:
        """Analyse structure fichier CSV"""
        structure = {
            'format': 'CSV',
            'headers': [],
            'sample_data': [],
            'total_rows': 0,
        }
        
        # Détection automatique du délimiteur
        with open(file_path, 'r', encoding='utf-8') as file:
            sample = file.read(1024)
            delimiter = ImportService._detect_csv_delimiter(sample)
        
        # Lecture avec pandas pour analyse rapide
        df = pd.read_csv(file_path, delimiter=delimiter, nrows=100)
        
        structure['headers'] = df.columns.tolist()
        structure['sample_data'] = df.head(10).values.tolist()
        structure['total_rows'] = len(df)
        structure['detected_delimiter'] = delimiter
        
        return structure
    
    @staticmethod
    def _detect_csv_delimiter(sample: str) -> str:
        """Détection automatique du délimiteur CSV"""
        delimiters = [';', ',', '\t', '|']
        delimiter_counts = {}
        
        for delimiter in delimiters:
            count = sample.count(delimiter)
            if count > 0:
                delimiter_counts[delimiter] = count
        
        if delimiter_counts:
            return max(delimiter_counts.items(), key=lambda x: x[1])[0]
        
        return ';'  # Par défaut français
    
    @staticmethod
    def _detect_data_type(structure: Dict) -> str:
        """
        Détection automatique du type de données selon les colonnes
        """
        headers = []
        
        if structure['format'] == 'EXCEL':
            # Collecte toutes les en-têtes des feuilles
            for worksheet in structure['worksheets']:
                headers.extend([h.lower() for h in worksheet['headers']])
        else:
            headers = [h.lower() for h in structure.get('headers', [])]
        
        headers_text = ' '.join(headers)
        
        # Détection par mots-clés
        if any(keyword in headers_text for keyword in ['compte', 'account', 'numero', 'libelle']):
            if any(keyword in headers_text for keyword in ['balance', 'solde', 'debit', 'credit']):
                return 'OPENING_BALANCE'
            return 'CHART_OF_ACCOUNTS'
        
        if any(keyword in headers_text for keyword in ['ecriture', 'journal', 'piece']):
            return 'JOURNAL_ENTRIES'
        
        if any(keyword in headers_text for keyword in ['client', 'customer', 'raison_sociale']):
            return 'CUSTOMERS'
        
        if any(keyword in headers_text for keyword in ['fournisseur', 'supplier', 'vendor']):
            return 'SUPPLIERS'
        
        return 'JOURNAL_ENTRIES'  # Par défaut
    
    @staticmethod
    def process_sage_migration(
        session: ImportSession,
        sage_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Migration spécialisée depuis Sage Comptabilité
        Avec mapping automatique plan comptable et écritures
        """
        migration_results = {
            'accounts_migrated': 0,
            'entries_migrated': 0,
            'customers_migrated': 0,
            'suppliers_migrated': 0,
            'errors': [],
            'warnings': [],
        }
        
        try:
            with transaction.atomic():
                session.status = 'IMPORTING'
                session.current_step = 'Migration plan comptable Sage'
                session.save()
                
                # 1. Migration du plan comptable avec mapping Sage → SYSCOHADA
                accounts_result = ImportService._migrate_sage_chart_of_accounts(session, sage_config)
                migration_results['accounts_migrated'] = accounts_result['created']
                
                session.current_step = 'Migration écritures comptables'
                session.progress_percent = 30
                session.save()
                
                # 2. Migration des écritures avec validation
                entries_result = ImportService._migrate_sage_journal_entries(session, sage_config)
                migration_results['entries_migrated'] = entries_result['created']
                
                session.current_step = 'Migration fichier tiers'
                session.progress_percent = 60
                session.save()
                
                # 3. Migration clients/fournisseurs
                if sage_config.get('include_third_parties', True):
                    customers_result = ImportService._migrate_sage_customers(session, sage_config)
                    suppliers_result = ImportService._migrate_sage_suppliers(session, sage_config)
                    
                    migration_results['customers_migrated'] = customers_result['created']
                    migration_results['suppliers_migrated'] = suppliers_result['created']
                
                session.current_step = 'Validation finale et contrôles'
                session.progress_percent = 90
                session.save()
                
                # 4. Validation globale cohérence
                validation_result = ImportService._validate_migration_coherence(session)
                migration_results['warnings'].extend(validation_result.get('warnings', []))
                
                # Finalisation
                session.status = 'COMPLETED'
                session.progress_percent = 100
                session.end_time = timezone.now()
                session.save()
                
                logger.info(f"Migration Sage terminée: {migration_results}")
                
                return migration_results
                
        except Exception as e:
            session.status = 'FAILED'
            session.error_summary = {'critical_error': str(e)}
            session.save()
            
            logger.error(f"Erreur migration Sage: {e}")
            raise
    
    @staticmethod
    def _migrate_sage_chart_of_accounts(session: ImportSession, config: Dict) -> Dict[str, int]:
        """
        Migration spécialisée plan comptable Sage vers SYSCOHADA
        """
        # Mapping Sage standard vers SYSCOHADA
        sage_to_syscohada_mapping = {
            # Classe 1 - Comptes de capitaux
            '101': '101',    # Capital social
            '106': '1061',   # Réserve légale
            '110': '110',    # Report à nouveau créditeur
            '120': '120',    # Résultat de l'exercice (bénéfice)
            
            # Classe 2 - Immobilisations
            '211': '211',    # Brevets, licences
            '213': '213',    # Fonds commercial
            '2154': '222',   # Terrains nus
            '2155': '223',   # Terrains bâtis
            '2313': '231',   # Bâtiments industriels
            '2154': '244',   # Matériel et mobilier
            '2182': '245',   # Matériel de transport
            '2183': '246',   # Matériel informatique
            
            # Classe 4 - Tiers (adaptation majeure Sage → SYSCOHADA)
            '401': '401',    # Fournisseurs
            '4011': '4011',  # Fournisseurs ordinaires
            '408': '408',    # Fournisseurs factures non parvenues
            '409': '409',    # Fournisseurs débiteurs
            '411': '411',    # Clients
            '4111': '4111',  # Clients ordinaires
            '413': '412',    # Clients effets à recevoir
            '416': '413',    # Clients douteux
            '419': '419',    # Clients créditeurs
            
            # Classe 5 - Trésorerie
            '512': '521',    # Banques (adaptation SYSCOHADA)
            '530': '531',    # Chèques postaux
            '571': '571',    # Caisse
            
            # Classe 6 - Charges (mapping enrichi)
            '601': '601',    # Achats marchandises
            '602': '602',    # Achats matières premières
            '606': '604',    # Achats prestations services
            '611': '611',    # Transports sur achats
            '613': '612',    # Transports sur ventes
            '621': '621',    # Sous-traitance
            '622': '622',    # Locations
            '625': '623',    # Réparations entretien
            '641': '641',    # Rémunérations personnel
            '645': '662',    # Charges sociales
            '661': '627',    # Services bancaires
            '671': '671',    # Intérêts emprunts
            
            # Classe 7 - Produits
            '701': '701',    # Ventes marchandises
            '706': '706',    # Prestations services
            '708': '708',    # Produits activités annexes
            '761': '761',    # Produits participation
            '764': '764',    # Revenus prêts
            '771': '771',    # Intérêts des prêts
        }
        
        created_count = 0
        
        # Lecture du plan comptable depuis le fichier
        chart_data = ImportService._parse_chart_of_accounts_data(session)
        
        for source_account in chart_data:
            sage_code = source_account.get('account_code', '')
            account_name = source_account.get('account_name', '')
            
            # Mapping automatique Sage → SYSCOHADA
            syscohada_code = ImportService._map_sage_to_syscohada(
                sage_code, account_name, sage_to_syscohada_mapping
            )
            
            if syscohada_code:
                # Création compte SYSCOHADA
                try:
                    account = ChartOfAccounts.objects.get_or_create(
                        company=session.company,
                        code=syscohada_code,
                        defaults={
                            'name': account_name,
                            'account_class': syscohada_code[0],
                            'account_type': 'DETAIL' if len(syscohada_code) >= 3 else 'TOTAL',
                            'normal_balance': ImportService._determine_normal_balance(syscohada_code),
                            'is_reconcilable': syscohada_code.startswith(('401', '411', '521')),
                        }
                    )[0]
                    
                    created_count += 1
                    
                    # Log de transformation
                    ImportService._log_transformation(
                        session, 'CREATE', 'Compte créé',
                        source_account, {'syscohada_code': syscohada_code}
                    )
                    
                except Exception as e:
                    ImportService._log_error(
                        session, 'MAPPING', f"Erreur création compte {syscohada_code}: {e}",
                        source_account
                    )
        
        return {'created': created_count}
    
    @staticmethod
    def _map_sage_to_syscohada(
        sage_code: str, 
        account_name: str, 
        mapping_table: Dict[str, str]
    ) -> Optional[str]:
        """
        Mapping intelligent Sage vers SYSCOHADA avec IA
        """
        # 1. Correspondance exacte
        if sage_code in mapping_table:
            return mapping_table[sage_code]
        
        # 2. Correspondance par préfixe (ex: 4011xx → 4011)
        for sage_prefix, syscohada_code in mapping_table.items():
            if sage_code.startswith(sage_prefix):
                return syscohada_code + sage_code[len(sage_prefix):]
        
        # 3. Analyse sémantique du libellé (IA basique)
        syscohada_suggestion = ImportService._suggest_syscohada_by_name(account_name)
        if syscohada_suggestion:
            return syscohada_suggestion
        
        # 4. Conservation structure si compatible SYSCOHADA
        if len(sage_code) >= 3 and sage_code[0] in '1234567':
            return sage_code  # Déjà compatible
        
        return None  # Mapping manuel requis
    
    @staticmethod
    def _suggest_syscohada_by_name(account_name: str) -> Optional[str]:
        """
        Suggestion compte SYSCOHADA par analyse du nom (IA sémantique)
        """
        name_lower = account_name.lower()
        
        # Dictionnaire de correspondances sémantiques
        semantic_mappings = {
            # Banques et trésorerie
            ('banque', 'compte bancaire', 'sgbc', 'boa', 'bicec'): '521',
            ('caisse', 'especes', 'liquidite'): '571',
            ('cheque', 'cheques postaux'): '531',
            
            # Clients
            ('client', 'creance client', 'debiteur'): '411',
            ('client douteux', 'provision client'): '413',
            
            # Fournisseurs  
            ('fournisseur', 'dette fournisseur', 'crediteur'): '401',
            ('facture non parvenue', 'fnp'): '408',
            
            # Charges courantes
            ('achat marchandise', 'marchandise'): '601',
            ('achat matiere premiere', 'matiere'): '602',
            ('transport', 'livraison'): '611',
            ('location', 'loyer'): '622',
            ('telephone', 'communication'): '636',
            ('electricite', 'eau', 'gaz'): '605',
            ('salaire', 'remuneration', 'personnel'): '661',
            ('charge sociale', 'cotisation'): '662',
            
            # Produits
            ('vente marchandise', 'chiffre affaire'): '701',
            ('prestation service', 'service'): '706',
            ('subvention', 'aide'): '711',
            
            # Immobilisations
            ('terrain', 'foncier'): '222',
            ('batiment', 'construction'): '231',
            ('materiel', 'equipement'): '244',
            ('vehicule', 'transport', 'automobile'): '245',
            ('informatique', 'ordinateur', 'logiciel'): '246',
        }
        
        for keywords, syscohada_code in semantic_mappings.items():
            if any(keyword in name_lower for keyword in keywords):
                return syscohada_code
        
        return None
    
    @staticmethod
    def _parse_chart_of_accounts_data(session: ImportSession) -> List[Dict]:
        """Parse les données du plan comptable selon le format"""
        if session.detected_structure['format'] == 'EXCEL':
            return ImportService._parse_excel_chart_data(session)
        elif session.detected_structure['format'] == 'CSV':
            return ImportService._parse_csv_chart_data(session)
        
        return []
    
    @staticmethod
    def _parse_excel_chart_data(session: ImportSession) -> List[Dict]:
        """Parse plan comptable depuis Excel"""
        file_path = session.source_file.path
        
        # Recherche de la feuille contenant le plan comptable
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        target_sheet = None
        
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['plan', 'compte', 'chart']):
                target_sheet = workbook[sheet_name]
                break
        
        if not target_sheet:
            target_sheet = workbook.active  # Première feuille par défaut
        
        accounts = []
        headers = [str(cell.value).strip() for cell in target_sheet[1]]  # En-têtes ligne 1
        
        # Détection colonnes essentielles
        code_col = ImportService._find_column_index(headers, ['code', 'numero', 'account'])
        name_col = ImportService._find_column_index(headers, ['libelle', 'nom', 'name', 'description'])
        
        if code_col is None or name_col is None:
            raise ValidationError("Colonnes 'Code' et 'Libellé' introuvables dans le fichier")
        
        # Lecture des comptes
        for row in target_sheet.iter_rows(min_row=2, values_only=True):
            if row[code_col] and row[name_col]:
                accounts.append({
                    'account_code': str(row[code_col]).strip(),
                    'account_name': str(row[name_col]).strip(),
                    'source_row': len(accounts) + 2,  # +2 car commence ligne 2
                })
        
        workbook.close()
        return accounts
    
    @staticmethod
    def _find_column_index(headers: List[str], possible_names: List[str]) -> Optional[int]:
        """Trouve l'index d'une colonne par ses noms possibles"""
        for i, header in enumerate(headers):
            header_lower = header.lower().strip()
            if any(name in header_lower for name in possible_names):
                return i
        return None
    
    @staticmethod
    def _migrate_sage_journal_entries(session: ImportSession, config: Dict) -> Dict[str, int]:
        """Migration des écritures Sage avec recodification automatique"""
        entries_data = ImportService._parse_journal_entries_data(session)
        created_count = 0
        
        # Regroupement par écriture (numéro pièce + journal + date)
        grouped_entries = {}
        
        for line_data in entries_data:
            key = (
                line_data.get('journal_code', ''),
                line_data.get('piece_number', ''),
                line_data.get('entry_date', '')
            )
            
            if key not in grouped_entries:
                grouped_entries[key] = []
            grouped_entries[key].append(line_data)
        
        # Création des écritures groupées
        for entry_key, lines in grouped_entries.items():
            try:
                # Validation équilibrage de l'écriture
                total_debit = sum(Decimal(str(line.get('debit_amount', 0))) for line in lines)
                total_credit = sum(Decimal(str(line.get('credit_amount', 0))) for line in lines)
                
                if abs(total_debit - total_credit) > Decimal('0.01'):
                    ImportService._log_error(
                        session, 'VALIDATION', 
                        f"Écriture déséquilibrée: {entry_key}",
                        {'debit': float(total_debit), 'credit': float(total_credit)}
                    )
                    continue
                
                # Création écriture WiseBook
                entry_created = ImportService._create_wisebook_entry(session, lines)
                if entry_created:
                    created_count += 1
                
            except Exception as e:
                ImportService._log_error(
                    session, 'CREATE', f"Erreur création écriture {entry_key}: {e}",
                    lines[0] if lines else {}
                )
        
        return {'created': created_count}
    
    @staticmethod
    def _create_wisebook_entry(session: ImportSession, lines_data: List[Dict]) -> bool:
        """Crée une écriture WiseBook depuis données Sage"""
        from apps.accounting.services.ecriture_service import EcritureService
        
        if not lines_data:
            return False
        
        first_line = lines_data[0]
        
        # Recherche journal WiseBook correspondant
        sage_journal_code = first_line.get('journal_code', '')
        wisebook_journal = ImportService._map_sage_journal_to_wisebook(
            session.company, sage_journal_code
        )
        
        if not wisebook_journal:
            raise ValidationError(f"Journal WiseBook introuvable pour code Sage: {sage_journal_code}")
        
        # Préparation données écriture
        entry_data = {
            'entry_date': datetime.strptime(first_line['entry_date'], '%Y-%m-%d').date(),
            'description': first_line.get('description', f"Migration Sage - {first_line.get('piece_number')}"),
            'reference': first_line.get('reference', ''),
            'source_document': f"Migration Sage - Session {session.id}",
        }
        
        # Préparation lignes avec mapping comptes
        mapped_lines = []
        for line in lines_data:
            sage_account_code = line.get('account_code', '')
            
            # Recherche compte WiseBook mappé
            wisebook_account = ImportService._find_mapped_account(
                session.company, sage_account_code
            )
            
            if not wisebook_account:
                # Tentative création automatique si mapping disponible
                wisebook_account = ImportService._create_account_from_mapping(
                    session.company, sage_account_code, line.get('account_name', '')
                )
            
            if wisebook_account:
                mapped_lines.append({
                    'account_code': wisebook_account.code,
                    'label': line.get('line_description', entry_data['description']),
                    'debit_amount': float(line.get('debit_amount', 0)),
                    'credit_amount': float(line.get('credit_amount', 0)),
                })
        
        if not mapped_lines:
            raise ValidationError("Aucune ligne mappée pour l'écriture")
        
        # Création via service WiseBook
        try:
            entry = EcritureService.create_journal_entry(
                company=session.company,
                journal=wisebook_journal,
                fiscal_year=session.fiscal_year or session.company.current_fiscal_year,
                entry_data=entry_data,
                lines_data=mapped_lines,
                auto_validate=False  # Validation manuelle après import
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Erreur création écriture WiseBook: {e}")
            return False
    
    @staticmethod
    def _map_sage_journal_to_wisebook(company: Company, sage_code: str) -> Optional[Journal]:
        """Mapping journal Sage vers WiseBook"""
        # Correspondances standards Sage → SYSCOHADA
        journal_mappings = {
            'VEN': 'VE',    # Ventes
            'ACH': 'AC',    # Achats
            'BAN': 'BQ',    # Banque
            'CAI': 'CA',    # Caisse
            'OD': 'OD',     # Opérations diverses
            'A-NOU': 'AN',  # À nouveaux
        }
        
        wisebook_code = journal_mappings.get(sage_code, sage_code)
        
        return Journal.objects.filter(
            company=company,
            code=wisebook_code
        ).first()
    
    @staticmethod
    def _find_mapped_account(company: Company, source_code: str) -> Optional[ChartOfAccounts]:
        """Recherche compte WiseBook mappé"""
        # Recherche directe
        account = ChartOfAccounts.objects.filter(
            company=company,
            code=source_code
        ).first()
        
        if account:
            return account
        
        # Recherche avec mapping standard Sage
        mapped_code = ImportService._map_sage_account_code(source_code)
        if mapped_code:
            return ChartOfAccounts.objects.filter(
                company=company,
                code=mapped_code
            ).first()
        
        return None
    
    @staticmethod
    def _map_sage_account_code(sage_code: str) -> Optional[str]:
        """Mapping automatique code Sage vers SYSCOHADA"""
        # Règles de transformation courantes
        if len(sage_code) >= 3:
            # La plupart des codes Sage sont déjà compatibles SYSCOHADA
            return sage_code
        
        return None
    
    @staticmethod
    def generate_migration_preview(session: ImportSession) -> ImportPreview:
        """
        Génère une prévisualisation complète de la migration
        """
        # Analyse des données à importer
        sample_data = ImportService._extract_sample_data(session, limit=100)
        
        # Estimation impact
        estimates = ImportService._estimate_import_impact(session, sample_data)
        
        # Création prévisualisation
        preview = ImportPreview.objects.create(
            import_session=session,
            sample_data=sample_data,
            accounts_to_create=estimates['accounts_count'],
            entries_to_create=estimates['entries_count'],
            third_parties_to_create=estimates['third_parties_count'],
            estimated_processing_time=estimates['processing_time'],
            estimated_disk_space=estimates['disk_space_mb']
        )
        
        return preview
    
    @staticmethod
    def _extract_sample_data(session: ImportSession, limit: int = 100) -> List[Dict]:
        """Extrait un échantillon des données transformées"""
        # Implémentation selon le format de fichier
        if session.detected_structure['format'] == 'EXCEL':
            return ImportService._extract_excel_sample(session, limit)
        elif session.detected_structure['format'] == 'CSV':
            return ImportService._extract_csv_sample(session, limit)
        
        return []
    
    @staticmethod
    def _estimate_import_impact(session: ImportSession, sample_data: List) -> Dict[str, Any]:
        """Estimation de l'impact de l'import"""
        return {
            'accounts_count': len(set(item.get('account_code') for item in sample_data if item.get('account_code'))),
            'entries_count': session.total_rows_detected,
            'third_parties_count': len(set(item.get('third_party_code') for item in sample_data if item.get('third_party_code'))),
            'processing_time': session.total_rows_detected * 0.1,  # 100ms par ligne
            'disk_space_mb': session.file_size / 1024 / 1024 * 3,  # 3x taille fichier
        }
    
    @staticmethod
    def _log_transformation(session: ImportSession, action: str, step: str, input_data: Dict, output_data: Dict):
        """Log une transformation"""
        from .models import DataTransformationLog
        
        DataTransformationLog.objects.create(
            import_session=session,
            action=action,
            step_name=step,
            source_row=input_data.get('source_row', 0),
            input_data=input_data,
            output_data=output_data,
            log_message=f"{action}: {step}"
        )
    
    @staticmethod
    def _log_error(session: ImportSession, error_type: str, message: str, data: Dict):
        """Log une erreur d'import"""
        ImportValidationError.objects.create(
            import_session=session,
            error_type=error_type,
            severity='ERROR',
            row_number=data.get('source_row', 0),
            error_code=f"{error_type}_ERROR",
            error_message=message,
            source_data=data
        )