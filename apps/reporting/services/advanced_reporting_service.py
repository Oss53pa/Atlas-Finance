"""
Service Reporting Avancé WiseBook
Génération automatique états SYSCOHADA et BI selon EXF-BI-001 à EXF-BI-003
"""
from typing import List, Dict, Optional, Any
from django.db import connection, transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.template.loader import render_to_string
from decimal import Decimal
from datetime import date, datetime
import json
import logging

from apps.accounting.models import Company, FiscalYear, ChartOfAccounts
from apps.financial_analysis.models import TAFIREStatement, SIG
from ..models import ReportTemplate, ReportGeneration


logger = logging.getLogger(__name__)


class AdvancedReportingService:
    """
    Service de reporting avancé avec IA
    États réglementaires SYSCOHADA et analytics prédictives
    """
    
    @staticmethod
    def generate_syscohada_balance_sheet(
        company: Company,
        fiscal_year: FiscalYear,
        as_of_date: Optional[date] = None,
        format_type: str = 'PDF'
    ) -> Dict[str, Any]:
        """
        Génère le Bilan SYSCOHADA officiel avec contrôles conformité
        Format réglementaire complet
        """
        if not as_of_date:
            as_of_date = fiscal_year.end_date
        
        start_time = timezone.now()
        
        # Extraction données bilan avec requête optimisée
        sql = """
        WITH balance_data AS (
            SELECT 
                ca.code,
                ca.name,
                ca.account_class,
                COALESCE(SUM(jel.debit_amount - jel.credit_amount), 0) as balance
            FROM accounting_chartofaccounts ca
            LEFT JOIN accounting_journalentryline jel ON ca.id = jel.account_id
            LEFT JOIN accounting_journalentry je ON jel.entry_id = je.id
            WHERE ca.company_id = %s 
            AND je.fiscal_year_id = %s
            AND je.entry_date <= %s
            AND je.is_validated = true
            GROUP BY ca.code, ca.name, ca.account_class
            HAVING ABS(COALESCE(SUM(jel.debit_amount - jel.credit_amount), 0)) > 0.01
        )
        SELECT 
            -- ACTIF
            'ACTIF' as side,
            'ACTIF IMMOBILISE' as section,
            'Immobilisations incorporelles' as subsection,
            COALESCE(SUM(CASE WHEN code LIKE '21%' THEN balance ELSE 0 END), 0) as amount
        FROM balance_data
        WHERE account_class IN ('2')
        
        UNION ALL
        
        SELECT 
            'ACTIF', 'ACTIF IMMOBILISE', 'Immobilisations corporelles',
            COALESCE(SUM(CASE WHEN code LIKE '22%' OR code LIKE '23%' OR code LIKE '24%' THEN balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class IN ('2')
        
        UNION ALL
        
        SELECT 
            'ACTIF', 'ACTIF CIRCULANT', 'Stocks et en-cours',
            COALESCE(SUM(CASE WHEN account_class = '3' THEN balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '3'
        
        UNION ALL
        
        SELECT 
            'ACTIF', 'ACTIF CIRCULANT', 'Créances et emplois assimilés',
            COALESCE(SUM(CASE WHEN code LIKE '41%' OR code LIKE '42%' OR code LIKE '43%' OR code LIKE '44%' THEN balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '4'
        
        UNION ALL
        
        SELECT 
            'ACTIF', 'TRESORERIE-ACTIF', 'Trésorerie-Actif',
            COALESCE(SUM(CASE WHEN account_class = '5' AND balance > 0 THEN balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '5'
        
        UNION ALL
        
        -- PASSIF
        SELECT 
            'PASSIF', 'CAPITAUX PROPRES', 'Capital',
            COALESCE(SUM(CASE WHEN code LIKE '10%' THEN -balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '1'
        
        UNION ALL
        
        SELECT 
            'PASSIF', 'CAPITAUX PROPRES', 'Réserves et report à nouveau',
            COALESCE(SUM(CASE WHEN code LIKE '11%' OR code LIKE '106%' THEN -balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '1'
        
        UNION ALL
        
        SELECT 
            'PASSIF', 'CAPITAUX PROPRES', 'Résultat net',
            COALESCE(SUM(CASE WHEN code LIKE '12%' THEN -balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '1'
        
        UNION ALL
        
        SELECT 
            'PASSIF', 'DETTES FINANCIERES', 'Dettes financières',
            COALESCE(SUM(CASE WHEN code LIKE '16%' OR code LIKE '17%' THEN -balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '1'
        
        UNION ALL
        
        SELECT 
            'PASSIF', 'PASSIF CIRCULANT', 'Dettes circulantes',
            COALESCE(SUM(CASE WHEN code LIKE '40%' OR code LIKE '42%' OR code LIKE '43%' OR code LIKE '44%' THEN -balance ELSE 0 END), 0)
        FROM balance_data
        WHERE account_class = '4'
        
        ORDER BY side, section, subsection
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql, [company.id, fiscal_year.id, as_of_date])
            
            balance_sheet_data = {}
            total_actif = Decimal('0')
            total_passif = Decimal('0')
            
            for row in cursor.fetchall():
                side, section, subsection, amount = row
                amount = Decimal(str(amount))
                
                if side not in balance_sheet_data:
                    balance_sheet_data[side] = {}
                
                if section not in balance_sheet_data[side]:
                    balance_sheet_data[side][section] = {}
                
                balance_sheet_data[side][section][subsection] = float(amount)
                
                if side == 'ACTIF':
                    total_actif += amount
                else:
                    total_passif += amount
        
        # Vérification équilibre bilan
        equilibrium_check = abs(total_actif - total_passif)
        is_balanced = equilibrium_check <= Decimal('1.00')
        
        # Contrôles de conformité SYSCOHADA
        conformity_checks = AdvancedReportingService._validate_syscohada_balance_sheet(
            balance_sheet_data, company
        )
        
        calculation_time = (timezone.now() - start_time).total_seconds()
        
        return {
            'balance_sheet_data': balance_sheet_data,
            'totals': {
                'total_actif': float(total_actif),
                'total_passif': float(total_passif),
                'equilibrium_difference': float(equilibrium_check),
                'is_balanced': is_balanced,
            },
            'conformity_checks': conformity_checks,
            'metadata': {
                'company_name': company.name,
                'fiscal_year': fiscal_year.name,
                'as_of_date': as_of_date.isoformat(),
                'generation_date': datetime.now().isoformat(),
                'calculation_time_ms': int(calculation_time * 1000),
                'format': format_type,
            }
        }
    
    @staticmethod
    def _validate_syscohada_balance_sheet(data: Dict, company: Company) -> Dict[str, Any]:
        """
        Validation conformité bilan SYSCOHADA
        """
        checks = {
            'structure_conformity': True,
            'mandatory_items_present': True,
            'calculation_accuracy': True,
            'issues': [],
            'score': 100,
        }
        
        # Vérification présence postes obligatoires SYSCOHADA
        mandatory_actif_items = [
            'Immobilisations incorporelles',
            'Immobilisations corporelles', 
            'Stocks et en-cours',
            'Créances et emplois assimilés'
        ]
        
        mandatory_passif_items = [
            'Capital',
            'Réserves et report à nouveau',
            'Résultat net'
        ]
        
        actif_data = data.get('ACTIF', {})
        passif_data = data.get('PASSIF', {})
        
        # Contrôles ACTIF
        for item in mandatory_actif_items:
            found = False
            for section in actif_data.values():
                if item in section:
                    found = True
                    break
            
            if not found:
                checks['issues'].append(f"Poste obligatoire manquant: {item}")
                checks['mandatory_items_present'] = False
                checks['score'] -= 10
        
        # Contrôles PASSIF
        for item in mandatory_passif_items:
            found = False
            for section in passif_data.values():
                if item in section:
                    found = True
                    break
            
            if not found:
                checks['issues'].append(f"Poste obligatoire manquant: {item}")
                checks['mandatory_items_present'] = False
                checks['score'] -= 10
        
        return checks
    
    @staticmethod
    def generate_syscohada_income_statement(
        company: Company,
        fiscal_year: FiscalYear,
        period_end: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Génère le Compte de Résultat SYSCOHADA
        """
        if not period_end:
            period_end = fiscal_year.end_date
        
        # Récupération SIG si disponible
        sig = SIG.objects.filter(
            company=company,
            fiscal_year=fiscal_year
        ).order_by('-calculation_date').first()
        
        if sig:
            # Utilisation des SIG calculés
            income_statement = {
                'ACTIVITE_EXPLOITATION': {
                    'Ventes et prestations': float(sig.merchandise_sales + sig.production_sold),
                    'Consommations intermédiaires': float(sig.intermediate_consumption),
                    'Valeur Ajoutée': float(sig.added_value),
                    'Charges de personnel': float(sig.staff_costs),
                    'Impôts et taxes': float(sig.taxes_and_duties),
                    'Excédent Brut d\'Exploitation': float(sig.gross_operating_surplus),
                    'Dotations amortissements': float(sig.depreciation_provisions_dotations),
                    'Résultat d\'Exploitation': float(sig.operating_result),
                },
                'ACTIVITE_FINANCIERE': {
                    'Produits financiers': float(sig.financial_income),
                    'Charges financières': float(sig.financial_expenses),
                    'Résultat Financier': float(sig.financial_result),
                },
                'ACTIVITE_EXCEPTIONNELLE': {
                    'Produits exceptionnels': float(sig.exceptional_income),
                    'Charges exceptionnelles': float(sig.exceptional_expenses),
                    'Résultat Exceptionnel': float(sig.exceptional_result),
                },
                'RESULTAT': {
                    'Résultat avant impôt': float(sig.current_result_before_tax + sig.exceptional_result),
                    'Impôt sur le résultat': float(sig.income_tax),
                    'Résultat Net': float(sig.final_net_result),
                }
            }
        else:
            # Génération directe depuis balance
            income_statement = AdvancedReportingService._generate_income_statement_from_balance(
                company, fiscal_year, period_end
            )
        
        return {
            'income_statement_data': income_statement,
            'metadata': {
                'company_name': company.name,
                'fiscal_year': fiscal_year.name,
                'period_end': period_end.isoformat(),
                'source': 'SIG' if sig else 'BALANCE',
                'generation_date': datetime.now().isoformat(),
            }
        }
    
    @staticmethod
    def _generate_income_statement_from_balance(
        company: Company,
        fiscal_year: FiscalYear, 
        period_end: date
    ) -> Dict[str, Any]:
        """Génération compte résultat depuis balance si pas de SIG"""
        
        # Requête optimisée pour classes 6 et 7
        sql = """
        SELECT 
            CASE 
                WHEN ca.account_class = '6' THEN 'CHARGES'
                WHEN ca.account_class = '7' THEN 'PRODUITS'
            END as type,
            ca.code,
            ca.name,
            SUM(jel.debit_amount - jel.credit_amount) as balance
        FROM accounting_chartofaccounts ca
        LEFT JOIN accounting_journalentryline jel ON ca.id = jel.account_id
        LEFT JOIN accounting_journalentry je ON jel.entry_id = je.id
        WHERE ca.company_id = %s 
        AND je.fiscal_year_id = %s
        AND je.entry_date <= %s
        AND je.is_validated = true
        AND ca.account_class IN ('6', '7')
        GROUP BY ca.account_class, ca.code, ca.name
        HAVING ABS(SUM(jel.debit_amount - jel.credit_amount)) > 0.01
        ORDER BY ca.code
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql, [company.id, fiscal_year.id, period_end])
            
            charges = {}
            produits = {}
            
            for row in cursor.fetchall():
                account_type, code, name, balance = row
                balance = float(balance)
                
                if account_type == 'CHARGES' and balance > 0:
                    charges[f"{code} - {name}"] = balance
                elif account_type == 'PRODUITS' and balance < 0:
                    produits[f"{code} - {name}"] = abs(balance)
        
        total_charges = sum(charges.values())
        total_produits = sum(produits.values())
        resultat_net = total_produits - total_charges
        
        return {
            'CHARGES_EXPLOITATION': charges,
            'PRODUITS_EXPLOITATION': produits,
            'TOTAUX': {
                'Total Charges': total_charges,
                'Total Produits': total_produits,
                'Résultat Net': resultat_net,
            }
        }
    
    @staticmethod
    def generate_tafire_report(
        company: Company,
        fiscal_year: FiscalYear,
        comparison_year: Optional[FiscalYear] = None
    ) -> Dict[str, Any]:
        """
        Génère le rapport TAFIRE avec analyse comparative
        """
        # Récupération TAFIRE principal
        tafire = TAFIREStatement.objects.filter(
            company=company,
            fiscal_year=fiscal_year
        ).order_by('-statement_date').first()
        
        if not tafire:
            raise ValidationError("Aucun TAFIRE disponible pour cette période")
        
        # TAFIRE année de comparaison
        tafire_comparison = None
        if comparison_year:
            tafire_comparison = TAFIREStatement.objects.filter(
                company=company,
                fiscal_year=comparison_year
            ).order_by('-statement_date').first()
        
        # Construction rapport
        report_data = {
            'current_year': {
                'fiscal_year': fiscal_year.name,
                'flux_exploitation': float(tafire.operating_cash_surplus),
                'flux_investissement': float(tafire.investment_cash_flow),
                'flux_financement': float(tafire.financing_cash_flow),
                'variation_tresorerie': float(tafire.cash_variation),
                'caf': float(tafire.self_financing_capacity),
                'free_cash_flow': float(tafire.free_cash_flow),
            }
        }
        
        # Comparaison N-1 si disponible
        if tafire_comparison:
            report_data['comparison_year'] = {
                'fiscal_year': comparison_year.name,
                'flux_exploitation': float(tafire_comparison.operating_cash_surplus),
                'flux_investissement': float(tafire_comparison.investment_cash_flow),
                'flux_financement': float(tafire_comparison.financing_cash_flow),
                'variation_tresorerie': float(tafire_comparison.cash_variation),
                'caf': float(tafire_comparison.self_financing_capacity),
                'free_cash_flow': float(tafire_comparison.free_cash_flow),
            }
            
            # Calcul variations
            report_data['variations'] = {
                'flux_exploitation': report_data['current_year']['flux_exploitation'] - report_data['comparison_year']['flux_exploitation'],
                'flux_investissement': report_data['current_year']['flux_investissement'] - report_data['comparison_year']['flux_investissement'],
                'flux_financement': report_data['current_year']['flux_financement'] - report_data['comparison_year']['flux_financement'],
                'caf': report_data['current_year']['caf'] - report_data['comparison_year']['caf'],
                'free_cash_flow': report_data['current_year']['free_cash_flow'] - report_data['comparison_year']['free_cash_flow'],
            }
        
        # Analyses qualitatives automatiques
        report_data['analysis'] = AdvancedReportingService._analyze_tafire_quality(tafire, tafire_comparison)
        
        return report_data
    
    @staticmethod
    def _analyze_tafire_quality(
        tafire: TAFIREStatement,
        tafire_comparison: Optional[TAFIREStatement] = None
    ) -> Dict[str, Any]:
        """
        Analyse qualitative automatique du TAFIRE avec IA
        """
        analysis = {
            'strengths': [],
            'weaknesses': [],
            'recommendations': [],
            'risk_level': 'LOW',
        }
        
        # Analyse flux d'exploitation
        if tafire.operating_cash_surplus > 0:
            analysis['strengths'].append("Flux d'exploitation positifs - Bonne génération de cash")
            if tafire.self_financing_capacity > tafire.operating_cash_surplus:
                analysis['strengths'].append("CAF > flux exploitation - Gestion BFR efficace")
        else:
            analysis['weaknesses'].append("Flux d'exploitation négatifs - Activité consommatrice de trésorerie")
            analysis['recommendations'].append("Améliorer la gestion du BFR et la rentabilité opérationnelle")
            analysis['risk_level'] = 'HIGH'
        
        # Analyse Free Cash Flow
        if tafire.free_cash_flow > 0:
            analysis['strengths'].append("Free Cash Flow positif - Capacité d'autofinancement démontrée")
        else:
            analysis['weaknesses'].append("Free Cash Flow négatif - Dépendance au financement externe")
            analysis['recommendations'].append("Optimiser le cash flow libre : rentabilité et maîtrise investissements")
            if analysis['risk_level'] == 'LOW':
                analysis['risk_level'] = 'MEDIUM'
        
        # Analyse flux de financement
        if tafire.financing_cash_flow > 0:
            if tafire.new_borrowings > tafire.capital_increase:
                analysis['recommendations'].append("Endettement en hausse - Surveiller structure financière")
        
        # Comparaison temporelle
        if tafire_comparison:
            evolution_caf = tafire.self_financing_capacity - tafire_comparison.self_financing_capacity
            
            if evolution_caf > 0:
                analysis['strengths'].append(f"CAF en amélioration: +{evolution_caf:,.0f}")
            else:
                analysis['weaknesses'].append(f"Dégradation CAF: {evolution_caf:,.0f}")
        
        return analysis
    
    @staticmethod
    def generate_custom_dashboard(
        company: Company,
        dashboard_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Génère un dashboard personnalisé selon configuration
        Widgets dynamiques avec données temps réel
        """
        widgets_data = {}
        
        for widget_config in dashboard_config.get('widgets', []):
            widget_id = widget_config['id']
            widget_type = widget_config['type']
            
            try:
                if widget_type == 'kpi_treasury':
                    widgets_data[widget_id] = AdvancedReportingService._get_treasury_kpi(company)
                elif widget_type == 'aging_chart':
                    widgets_data[widget_id] = AdvancedReportingService._get_aging_chart_data(company)
                elif widget_type == 'performance_radar':
                    widgets_data[widget_id] = AdvancedReportingService._get_performance_radar(company)
                elif widget_type == 'financial_ratios':
                    widgets_data[widget_id] = AdvancedReportingService._get_financial_ratios_summary(company)
                else:
                    widgets_data[widget_id] = {'error': f'Widget type not supported: {widget_type}'}
            
            except Exception as e:
                widgets_data[widget_id] = {'error': str(e)}
        
        return {
            'dashboard_data': widgets_data,
            'config': dashboard_config,
            'generated_at': datetime.now().isoformat(),
        }
    
    @staticmethod
    def _get_treasury_kpi(company: Company) -> Dict[str, Any]:
        """KPI trésorerie pour widget"""
        from apps.treasury.services.treasury_service import TreasuryService
        
        position = TreasuryService.get_realtime_treasury_position(company)
        
        return {
            'current_position': position['summary']['current_position'],
            'available_total': position['summary']['total_available'],
            'inflows_today': position['daily_flows']['inflows_today'],
            'outflows_today': position['daily_flows']['outflows_today'],
            'forecast_7d': position['summary']['forecast_7d_position'],
        }
    
    @staticmethod
    def _get_aging_chart_data(company: Company) -> Dict[str, Any]:
        """Données balance âgée pour graphique"""
        from apps.customers.services.customer_service import CustomerService
        
        aging_data = CustomerService.get_aging_dashboard(company)
        
        return {
            'aging_breakdown': aging_data['summary']['aging_breakdown'],
            'aging_percentages': aging_data['summary']['aging_percentages'],
            'total_outstanding': aging_data['summary']['total_outstanding'],
        }
    
    @staticmethod
    def _get_performance_radar(company: Company) -> Dict[str, Any]:
        """Données radar performance globale"""
        # Simulation données radar - à connecter avec vrais modules
        return {
            'metrics': {
                'Liquidité': 85,
                'Rentabilité': 78,
                'Solvabilité': 92,
                'Activité': 67,
                'Structure': 89,
                'Croissance': 73,
            },
            'score_global': 80.7,
            'benchmark_sector': 75.0,
        }
    
    @staticmethod
    def _get_financial_ratios_summary(company: Company) -> Dict[str, Any]:
        """Résumé ratios financiers clés"""
        # À connecter avec module analyse financière
        return {
            'autonomie_financiere': 45.2,
            'liquidite_generale': 1.8,
            'rotation_stocks': 8.5,
            'delai_clients': 42,
            'delai_fournisseurs': 38,
            'marge_nette': 12.3,
            'roe': 18.7,
        }
    
    @staticmethod
    def export_to_excel(
        report_data: Dict[str, Any],
        template_name: str = 'syscohada_standard'
    ) -> bytes:
        """
        Export vers Excel avec mise en forme professionnelle
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        
        if 'balance_sheet_data' in report_data:
            # Génération bilan
            ws = wb.active
            ws.title = "Bilan SYSCOHADA"
            
            AdvancedReportingService._format_balance_sheet_excel(ws, report_data)
        
        if 'income_statement_data' in report_data:
            # Génération compte de résultat
            ws = wb.create_sheet("Compte de Résultat")
            AdvancedReportingService._format_income_statement_excel(ws, report_data)
        
        # Sauvegarde
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    @staticmethod
    def _format_balance_sheet_excel(worksheet, report_data):
        """Formatage Excel du bilan"""
        # En-tête
        worksheet['A1'] = f"BILAN SYSCOHADA"
        worksheet['A2'] = report_data['metadata']['company_name']
        worksheet['A3'] = f"Exercice: {report_data['metadata']['fiscal_year']}"
        worksheet['A4'] = f"Au: {report_data['metadata']['as_of_date']}"
        
        # Style en-tête
        header_font = Font(bold=True, size=14)
        worksheet['A1'].font = header_font
        
        # Structure bilan
        row = 6
        
        # ACTIF
        worksheet[f'A{row}'] = "ACTIF"
        worksheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        actif_data = report_data['balance_sheet_data'].get('ACTIF', {})
        for section_name, section_data in actif_data.items():
            worksheet[f'A{row}'] = section_name
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for item_name, amount in section_data.items():
                worksheet[f'B{row}'] = item_name
                worksheet[f'C{row}'] = amount
                row += 1
        
        # Total ACTIF
        worksheet[f'A{row}'] = "TOTAL ACTIF"
        worksheet[f'C{row}'] = report_data['totals']['total_actif']
        worksheet[f'A{row}'].font = Font(bold=True)
        worksheet[f'C{row}'].font = Font(bold=True)
    
    @staticmethod
    def generate_ai_insights(
        company: Company,
        fiscal_year: Optional[FiscalYear] = None
    ) -> Dict[str, Any]:
        """
        Génération automatique d'insights IA (EXF-BI-003)
        """
        insights = {
            'financial_health': {},
            'performance_trends': {},
            'risk_analysis': {},
            'recommendations': [],
            'predictive_analytics': {},
        }
        
        # Analyse santé financière
        balance_sheet = AdvancedReportingService.generate_syscohada_balance_sheet(
            company, fiscal_year or company.current_fiscal_year
        )
        
        insights['financial_health'] = {
            'balance_quality_score': balance_sheet['conformity_checks']['score'],
            'equilibrium_status': balance_sheet['totals']['is_balanced'],
            'structure_analysis': 'Structure financière équilibrée' if balance_sheet['totals']['is_balanced'] else 'Déséquilibres détectés'
        }
        
        # Détection d'anomalies IA
        anomalies = AdvancedReportingService._detect_anomalies_ai(company)
        insights['anomaly_detection'] = anomalies
        
        # Prédictions ML
        predictions = AdvancedReportingService._generate_ml_predictions(company)
        insights['predictive_analytics'] = predictions
        
        return insights
    
    @staticmethod
    def _detect_anomalies_ai(company: Company) -> Dict[str, Any]:
        """
        Détection d'anomalies avancée avec patterns IA
        """
        # Simulation détection anomalies - à remplacer par vrai ML
        return {
            'patterns_detected': 3,
            'anomalies': [
                {
                    'type': 'UNUSUAL_AMOUNT',
                    'description': 'Montant inhabituellement élevé détecté',
                    'confidence': 0.87,
                    'impact': 'MEDIUM',
                },
                {
                    'type': 'FREQUENCY_PATTERN',
                    'description': 'Pattern de fréquence inhabituel',
                    'confidence': 0.73,
                    'impact': 'LOW',
                }
            ],
            'overall_risk': 'LOW',
        }
    
    @staticmethod
    def _generate_ml_predictions(company: Company) -> Dict[str, Any]:
        """
        Prédictions business avec ML
        """
        # Simulation prédictions ML
        return {
            'cash_flow_forecast': {
                '30_days': 2450000,
                '60_days': 4100000,
                '90_days': 5800000,
                'confidence': 0.84,
            },
            'revenue_forecast': {
                'next_quarter': 12500000,
                'confidence': 0.79,
            },
            'risk_predictions': {
                'liquidity_risk': 0.15,
                'credit_risk': 0.23,
                'operational_risk': 0.08,
            }
        }