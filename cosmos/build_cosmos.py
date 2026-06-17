# -*- coding: utf-8 -*-
"""Cosmos Report Builder v3.0 — 360° opérations centre commercial.
Zéro VBA. Bilingue FR/EN. Multi-granularité M/Q/Y. Rapport mono-onglet multipage.
Formules compatibles Excel 365 + moteur de recalcul LibreOffice (INDEX/MATCH, SUMIFS,
AVERAGEIFS, EDATE, SUMPRODUCT, IF imbriqués — aucune LAMBDA/SWITCH/XLOOKUP en cellule)."""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.pagebreak import Break
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.title import Title
from openpyxl.chart.text import Text
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.image import Image as XLImage
import os
random.seed(42)

# ---------------------------------------------------------------- palette (Cosmos design tokens)
PRIMARY="0A352E"; PRIMARY_L="14564A"; PRIMARY_XL="3C7A6B"; NEUTRAL="8C8A8A"
SUCCESS="2A7B5D"; WARNING="E6A742"; DANGER="B9513D"
AMBER="E6A742"                 # builder accent (readable on dark sheets)
DARK_BG="0A0A0A"; DARK_PANEL="14161A"; DARK_INPUT="2A2410"
TXT_LIGHT="F5F5F5"; TXT_SEC_D="9CA3AF"; WHITE="FFFFFF"; BAND="F1F6F4"
CARD="FFFFFF"; CARD_BORDER="D7E2DD"; TXT_PRIM="0A2A23"; TXT_SEC="5E6E69"; HDR_TBL="EAF1EE"
G_TXT,G_BG=SUCCESS,"DCEFE7"; A_TXT,A_BG="9A6B14","FBEED3"; R_TXT,R_BG="8F3B2C","F6E0DA"
LINKC="1F6F55"
ICON={"rh":"👥","lease":"🏬","rec":"💰","foot":"🚶","mkt":"📣","com":"🛍️","hsse":"🛡️","fac":"🛠️"}
UI="Segoe UI"; MONO="Consolas"; WORD="Grand Hotel"; PAGE_TINT="EEF3F1"
LOGO_PATH=os.path.join(os.path.dirname(os.path.abspath(__file__)),"_cosmos_logo.png")

def gen_logo():
    from PIL import Image, ImageDraw, ImageFont
    W,H=720,200; img=Image.new("RGBA",(W,H),(0,0,0,0)); d=ImageDraw.Draw(img)
    def F(sz,bold=True):
        for n in (("segoeuib.ttf" if bold else "segoeui.ttf"),"arialbd.ttf","arial.ttf","DejaVuSans-Bold.ttf"):
            try: return ImageFont.truetype(n,sz)
            except Exception: continue
        return ImageFont.load_default()
    disc=(7,37,32); gold=(230,167,66); white=(245,250,248); light=(207,224,217)
    cx,cy,r=95,100,72
    d.ellipse([cx-r,cy-r,cx+r,cy+r],fill=disc)
    d.ellipse([cx-r,cy-r,cx+r,cy+r],outline=gold,width=8)
    cf=F(70); tb=d.textbbox((0,0),"C",font=cf); d.text((cx-(tb[2]-tb[0])/2,cy-(tb[3]-tb[1])/2-tb[1]),"C",font=cf,fill=gold)
    d.text((190,52),"COSMOS",font=F(64),fill=white)
    d.text((192,128),"EMERGENCE PLAZA · YOPOUGON",font=F(24,False),fill=light)
    img.save(LOGO_PATH)
gen_logo()

ICON_DIR=os.path.join(os.path.dirname(os.path.abspath(__file__)),"_icons")
def gen_icons():
    from PIL import Image, ImageDraw
    import math
    os.makedirs(ICON_DIR,exist_ok=True); S=128; W=10
    def cv():
        im=Image.new("RGBA",(S,S),(0,0,0,0)); return im,ImageDraw.Draw(im)
    def people(d,c): d.ellipse([28,24,56,52],outline=c,width=W); d.ellipse([72,24,100,52],outline=c,width=W); d.arc([18,54,66,112],195,345,fill=c,width=W); d.arc([62,54,110,112],195,345,fill=c,width=W)
    def person(d,c): d.ellipse([46,22,82,58],outline=c,width=W); d.arc([30,58,98,122],195,345,fill=c,width=W)
    def shop(d,c): d.polygon([(30,30),(98,30),(106,52),(22,52)],outline=c,width=W); d.rectangle([30,52,98,104],outline=c,width=W); d.rectangle([54,74,74,104],outline=c,width=W)
    def coins(d,c):
        for y in (74,54,34): d.ellipse([34,y,94,y+22],outline=c,width=W)
    def steps(d,c): d.ellipse([30,44,58,96],outline=c,width=W); d.ellipse([72,30,100,82],outline=c,width=W)
    def mega(d,c): d.polygon([(26,58),(74,38),(74,92),(26,72)],outline=c,width=W); d.line([80,46,100,38],fill=c,width=W); d.line([80,64,102,64],fill=c,width=W); d.line([80,82,100,90],fill=c,width=W)
    def bag(d,c): d.rectangle([34,52,94,106],outline=c,width=W); d.arc([46,30,82,74],180,360,fill=c,width=W)
    def shield(d,c): d.polygon([(64,20),(102,36),(102,70),(64,108),(26,70),(26,36)],outline=c,width=W); d.line([46,64,58,78],fill=c,width=W); d.line([58,78,84,46],fill=c,width=W)
    def gear(d,c):
        d.ellipse([42,42,86,86],outline=c,width=W); d.ellipse([57,57,71,71],outline=c,width=W)
        for a in range(0,360,45):
            x=64+42*math.cos(math.radians(a)); y=64+42*math.sin(math.radians(a)); d.ellipse([x-7,y-7,x+7,y+7],fill=c)
    def chart(d,c):
        for i,h in enumerate((38,62,28,50)): d.rectangle([28+i*20,104-h,42+i*20,104],outline=c,width=W)
    def money(d,c): d.ellipse([34,34,94,94],outline=c,width=W); d.line([64,42,64,86],fill=c,width=W); d.line([52,54,76,54],fill=c,width=W); d.line([52,70,76,70],fill=c,width=W)
    def hashh(d,c): d.line([50,28,42,100],fill=c,width=W); d.line([82,28,74,100],fill=c,width=W); d.line([32,50,98,50],fill=c,width=W); d.line([30,78,96,78],fill=c,width=W)
    def clock(d,c): d.ellipse([32,32,96,96],outline=c,width=W); d.line([64,64,64,42],fill=c,width=W); d.line([64,64,82,72],fill=c,width=W)
    def star(d,c):
        pts=[]
        for k in range(10):
            a=-math.pi/2+k*math.pi/5; rr=46 if k%2==0 else 19; pts.append((64+rr*math.cos(a),64+rr*math.sin(a)))
        d.polygon(pts,outline=c,width=W)
    def ruler(d,c): d.polygon([(30,60),(60,30),(98,68),(68,98)],outline=c,width=W)
    def gauge(d,c): d.arc([30,40,98,108],180,360,fill=c,width=W); d.line([64,74,86,52],fill=c,width=W)
    deptmap={"rh":people,"lease":shop,"rec":coins,"foot":steps,"mkt":mega,"com":bag,"hsse":shield,"fac":gear,"synth":chart}
    catmap={"chart":chart,"money":money,"hash":hashh,"steps":steps,"person":person,"clock":clock,"star":star,"ruler":ruler,"gauge":gauge}
    white=(245,250,248,255); primary=(10,53,46,255)
    for n,fn in deptmap.items():
        im,d=cv(); fn(d,white); im.save(os.path.join(ICON_DIR,"dep_"+n+".png"))
    for n,fn in catmap.items():
        im,d=cv(); fn(d,primary); im.save(os.path.join(ICON_DIR,"cat_"+n+".png"))
gen_icons()
UCAT={"%":"chart","FCFA":"money","FCFA/m²":"ruler","nb":"hash","visites":"steps","ETP":"person","j":"clock","mois":"clock","score":"star"}
def place_icon(s,fname,cell,px=18):
    p=os.path.join(ICON_DIR,fname+".png")
    if os.path.exists(p):
        im=XLImage(p); im.width=px; im.height=px; s.add_image(im,cell)

F_FCFA='#,##0\\ "FCFA";[Red](#,##0\\ "FCFA");"-"'
F_FCFA_R='0.0,," M";[Red](0.0,," M");"–"'   # (unused) compact millions
FCM='#,##0.0" M";[Red](#,##0.0" M");"–"'    # value already divided by 1e6 in formula
F_INT='#,##0;[Red](#,##0);"-"'
F_PCT='0.0%;[Red]-0.0%;"-"'
F_DELTA='[Green]"▲ "0.0%;[Red]"▼ "0.0%;"–"'
F_NUM1='#,##0.0;[Red](#,##0.0);"-"'

def fill(h): return PatternFill("solid", fgColor=h)
def side(h,style="thin"): return Side(style=style,color=h)
def box(c): return Border(left=side(c),right=side(c),top=side(c),bottom=side(c))

# ---------------------------------------------------------------- depts & H2 sub-modules
DEPTS=[
 ("rh","Capital Humain",["Effectifs","Recrutement","Masse salariale","Climat social","Formation"]),
 ("lease","Gestion locative",["Occupation","Effort locataire","Loyers & réversion","Vacance & rétention","Incentives & WALT"]),
 ("rec","Recouvrement",["Encours & balance âgée","DSO & recouvrement","Suivi de recouvrement","Provisions","Litiges & contentieux"]),
 ("foot","Fréquentation",["Fréquentation vs budget","Densité & conversion","Dwell time","Répartition par zone"]),
 ("mkt","Marketing",["Événements","Marketing digital","Publicité & communication","Analyse données (CRM/fidélité)"]),
 ("com","Commercial",["Ventes & CA/m²","Location à bail","Pop-store / specialty","Le Market","Partenariats & activations","Remises"]),
 ("hsse","Sécurité / HSSE",["Incidents","Accidentologie","Sécurité incendie","Sûreté","Gestion des risques","Formation HSSE"]),
 ("fac","Facility",["Maintenance prév./corr.","Tickets OT","Énergie & utilities","CAPEX","Inspections","Hygiène & propreté","SLA prestataires"]),
]
DEPT_FR={d[0]:d[1] for d in DEPTS}

# ---------------------------------------------------------------- KPI master (key,fr,en,unit,agg,dept,h2,is_budget,base)
KPIS=[
 # RH
 ("kpi.rh.headcount","Effectif (ETP)","Headcount (FTE)","ETP","LAST","rh",1,False,120),
 ("kpi.rh.fte_gap","Écart ETP vs budget","FTE gap vs budget","ETP","LAST","rh",1,True,5),
 ("kpi.rh.open_pos","Postes ouverts","Open positions","nb","LAST","rh",2,False,6),
 ("kpi.rh.ttf","Time-to-fill","Time-to-fill","j","AVG","rh",2,False,38),
 ("kpi.rh.payroll","Masse salariale","Payroll","FCFA","SUM","rh",3,True,85000000),
 ("kpi.rh.payroll_var","Masse sal. vs budget","Payroll vs budget","%","RATIO","rh",3,False,1.02),
 ("kpi.rh.turnover","Turnover","Turnover","%","AVG","rh",4,False,0.03),
 ("kpi.rh.absence","Absentéisme","Absenteeism","%","AVG","rh",4,False,0.04),
 ("kpi.rh.train_hours","Heures form./agent","Training hrs/FTE","h","AVG","rh",5,False,2.5),
 ("kpi.rh.train_cover","Couverture formation","Training coverage","%","AVG","rh",5,False,0.60),
 # LEASE
 ("kpi.lease.occ_gla","Occupation (GLA)","Occupancy (GLA)","%","AVG","lease",1,True,0.90),
 ("kpi.lease.occ_unit","Occupation (unités)","Occupancy (units)","%","AVG","lease",1,False,0.88),
 ("kpi.lease.gla_vacant","GLA vacant","Vacant GLA","m²","LAST","lease",4,False,1800),
 ("kpi.lease.ocr","Taux d'effort (OCR)","Occupancy Cost Ratio","%","RATIO","lease",2,True,0.12),
 ("kpi.lease.passing","Loyer en place","Passing rent","FCFA","LAST","lease",3,True,130000000),
 ("kpi.lease.erv","Valeur locative (ERV)","Est. Rental Value","FCFA","LAST","lease",3,False,140000000),
 ("kpi.lease.reversion","Réversion","Reversion","%","RATIO","lease",3,False,0.07),
 ("kpi.lease.vac_age","Ancienneté vacance","Vacancy age","mois","AVG","lease",4,False,5),
 ("kpi.lease.retention","Rétention renouvel.","Renewal retention","%","AVG","lease",4,False,0.80),
 ("kpi.lease.incentives","Franchises/incentives","Incentives","FCFA","SUM","lease",5,True,6000000),
 ("kpi.lease.rent_sqm","Loyer moyen /m²","Avg rent/sqm","FCFA/m²","RATIO","lease",3,True,9000),
 ("kpi.lease.walt","Durée résiduelle (WALT)","WALT","ans","AVG","lease",5,False,3.2),
 # RECOUVREMENT
 ("kpi.rec.ar","Créances (AR)","Receivables","FCFA","LAST","rec",1,False,420000000),
 ("kpi.rec.aged_0_30","Âgée 0-30j","Aged 0-30d","FCFA","LAST","rec",1,False,180000000),
 ("kpi.rec.aged_30_60","Âgée 30-60j","Aged 30-60d","FCFA","LAST","rec",1,False,110000000),
 ("kpi.rec.aged_60_90","Âgée 60-90j","Aged 60-90d","FCFA","LAST","rec",1,False,70000000),
 ("kpi.rec.aged_90p","Âgée 90j+","Aged 90d+","FCFA","LAST","rec",1,False,60000000),
 ("kpi.rec.dso","DSO réel","DSO","j","RATIO","rec",2,False,52),
 ("kpi.rec.recovery","Taux recouvrement","Recovery rate","%","RATIO","rec",2,True,0.86),
 ("kpi.rec.collected","Encaissements","Collections","FCFA","SUM","rec",2,True,220000000),
 ("kpi.rec.billed","Facturé","Billed","FCFA","SUM","rec",2,False,255000000),
 ("kpi.rec.dunning","Relances émises","Dunning notices","nb","SUM","rec",3,False,320),
 ("kpi.rec.promises","Promesses de paiement","Payment promises","FCFA","SUM","rec",3,False,95000000),
 ("kpi.rec.promise_kept","Taux promesses tenues","Promise-kept rate","%","AVG","rec",3,False,0.72),
 ("kpi.rec.collected_dun","Encaissé sur relance","Collected via dunning","FCFA","SUM","rec",3,True,60000000),
 ("kpi.rec.provision","Provisions douteuses","Doubtful provisions","%","RATIO","rec",4,False,0.09),
 ("kpi.rec.litige_count","Dossiers en litige","Disputed cases","nb","LAST","rec",5,False,14),
 ("kpi.rec.litige_amount","Montant en litige","Disputed amount","FCFA","LAST","rec",5,False,85000000),
 ("kpi.rec.litige_resolved","Taux résolution litiges","Dispute resolution rate","%","AVG","rec",5,False,0.55),
 ("kpi.rec.litige_age","Ancienneté litige","Dispute age","mois","AVG","rec",5,False,7),
 # FOOTFALL
 ("kpi.foot.total","Fréquentation","Total footfall","visites","SUM","foot",1,True,480000),
 ("kpi.foot.vs_budget","Footfall vs budget","Footfall vs budget","%","RATIO","foot",1,False,0.97),
 ("kpi.foot.per_sqm","Footfall /m²","Footfall/sqm","v/m²","RATIO","foot",2,False,12),
 ("kpi.foot.conversion","Conversion→ventes","Footfall→sales conv.","%","RATIO","foot",2,False,0.22),
 ("kpi.foot.dwell","Temps de visite","Dwell time","min","AVG","foot",3,False,47),
 ("kpi.foot.peak_share","Part heure de pointe","Peak-hour share","%","AVG","foot",4,False,0.18),
 ("kpi.foot.weekend_share","Part week-end","Weekend share","%","AVG","foot",4,False,0.42),
 # MARKETING
 ("kpi.mkt.events","Événements","Events","nb","SUM","mkt",1,False,4),
 ("kpi.mkt.event_att","Affluence événements","Event attendance","visites","SUM","mkt",1,False,12000),
 ("kpi.mkt.event_roi","ROI événement","Event ROI","%","RATIO","mkt",1,False,1.8),
 ("kpi.mkt.event_sat","Satisfaction événement","Event satisfaction","score","AVG","mkt",1,False,8.2),
 ("kpi.mkt.engagement","Taux d'engagement","Engagement rate","%","AVG","mkt",2,False,0.045),
 ("kpi.mkt.followers","Abonnés","Followers","nb","LAST","mkt",2,False,78000),
 ("kpi.mkt.reach","Portée (reach)","Reach","nb","SUM","mkt",2,False,540000),
 ("kpi.mkt.cpa","Coût/acquisition","CPA","FCFA","RATIO","mkt",2,False,3500),
 ("kpi.mkt.digital_conv","Conversion digitale","Digital conversion","%","RATIO","mkt",2,False,0.03),
 ("kpi.mkt.spend","Dépense mkt","Marketing spend","FCFA","SUM","mkt",3,True,22000000),
 ("kpi.mkt.spend_pct","Spend % du CA","Spend % revenue","%","RATIO","mkt",3,False,0.03),
 ("kpi.mkt.impressions","Impressions pub","Ad impressions","nb","SUM","mkt",3,False,1500000),
 ("kpi.mkt.campaigns","Campagnes actives","Active campaigns","nb","LAST","mkt",3,False,3),
 ("kpi.mkt.nps","NPS","NPS","score","AVG","mkt",4,False,42),
 ("kpi.mkt.members","Membres fidélité","Loyalty members","nb","LAST","mkt",4,False,26000),
 ("kpi.mkt.member_act","Activation membres","Member activation","%","AVG","mkt",4,False,0.38),
 ("kpi.mkt.basket_member","Panier moyen membre","Avg member basket","FCFA","AVG","mkt",4,False,18500),
 # COMMERCIAL
 ("kpi.com.tenant_sales","Ventes locataires","Tenant sales","FCFA","SUM","com",1,True,1250000000),
 ("kpi.com.total_sales","CA total locataires","Total turnover","FCFA","SUM","com",1,True,1300000000),
 ("kpi.com.sales_sqm","CA /m²","Sales/sqm","FCFA/m²","RATIO","com",1,False,62000),
 ("kpi.com.lfl","Like-for-like","Like-for-like","%","RATIO","com",1,False,0.06),
 ("kpi.com.new_leases","Nouveaux baux","New leases","nb","SUM","com",2,True,3),
 ("kpi.com.pipeline","Pipeline pondéré","Weighted pipeline","FCFA","LAST","com",2,True,340000000),
 ("kpi.com.leases_renewed","Baux renouvelés","Leases renewed","nb","SUM","com",2,False,2),
 ("kpi.com.pipeline_conv","Conversion pipeline","Pipeline conversion","%","RATIO","com",2,False,0.35),
 ("kpi.com.popup_count","Locations pop-up (nb)","Pop-up rentals (nb)","nb","LAST","com",3,True,6),
 ("kpi.com.popup_rev","Revenus pop-up","Pop-up revenue","FCFA","SUM","com",3,True,18000000),
 ("kpi.com.kiosk_occ","Taux d'occupation pop-up","Pop-up occupancy","%","AVG","com",3,True,0.78),
 ("kpi.com.market_rev","Revenus Le Market","Le Market revenue","FCFA","SUM","com",4,True,42000000),
 ("kpi.com.market_occ","Taux d'occupation Le Market","Le Market occupancy","%","AVG","com",4,True,0.83),
 ("kpi.com.market_stalls","Kiosques loués (Le Market)","Le Market kiosks let","nb","LAST","com",4,True,45),
 ("kpi.com.media_sales","Régie publicitaire","Media sales","FCFA","SUM","com",5,True,28000000),
 ("kpi.com.partnership_rev","Revenus partenariats","Partnership revenue","FCFA","SUM","com",5,False,15000000),
 ("kpi.com.activations","Activations annonceurs","Advertiser activations","nb","SUM","com",5,False,5),
 ("kpi.com.activation_rev","Revenus activations","Activation revenue","FCFA","SUM","com",5,False,9000000),
 ("kpi.com.discount","Escompte % CA brut","Discount % gross","%","RATIO","com",6,False,0.04),
 # HSSE
 ("kpi.hsse.incidents","Incidents","Incidents","nb","SUM","hsse",1,False,2),
 ("kpi.hsse.incident_sev","Gravité incidents","Incident severity","score","AVG","hsse",1,False,2.1),
 ("kpi.hsse.tf","Taux de fréquence","Frequency rate","TF","RATIO","hsse",2,False,4.2),
 ("kpi.hsse.tg","Taux de gravité","Severity rate","TG","RATIO","hsse",2,False,0.15),
 ("kpi.hsse.nearmiss","Presqu'accidents","Near-misses","nb","SUM","hsse",2,False,5),
 ("kpi.hsse.days_no_lti","Jours sans accident","Days w/o LTI","j","LAST","hsse",2,False,120),
 ("kpi.hsse.evac_drills","Exercices d'évacuation","Evacuation drills","nb","SUM","hsse",3,False,1),
 ("kpi.hsse.fire_conform","Conformité incendie","Fire safety conformity","%","AVG","hsse",3,False,0.92),
 ("kpi.hsse.fire_alarms","Déclenchements alarme","Fire alarm triggers","nb","SUM","hsse",3,False,3),
 ("kpi.hsse.evac_time","Temps d'évacuation","Evacuation time","min","AVG","hsse",3,False,7),
 ("kpi.hsse.sec_interv","Interventions sûreté","Security interventions","nb","SUM","hsse",4,False,24),
 ("kpi.hsse.theft","Vols / malveillance","Theft / malice events","nb","SUM","hsse",4,False,3),
 ("kpi.hsse.cctv","Couverture CCTV","CCTV coverage","%","AVG","hsse",4,False,0.95),
 ("kpi.hsse.guards","Agents déployés","Guards deployed","nb","LAST","hsse",4,False,18),
 ("kpi.hsse.risks_id","Risques identifiés","Risks identified","nb","SUM","hsse",5,False,6),
 ("kpi.hsse.risks_treated","Risques traités","Risks treated","%","AVG","hsse",5,False,0.70),
 ("kpi.hsse.action_close","Clôture plan d'actions","Action closure","%","AVG","hsse",5,True,0.82),
 ("kpi.hsse.audits","Audits réalisés","Audits done","nb","SUM","hsse",5,False,1),
 ("kpi.hsse.train_cover","Couverture form. HSSE","HSSE training cov.","%","AVG","hsse",6,True,0.70),
 # FACILITY
 ("kpi.fac.prev_corr","Préventif/correctif","Preventive/corrective","%","AVG","fac",1,False,0.65),
 ("kpi.fac.wo_prev","OT préventifs","Preventive WOs","nb","SUM","fac",1,False,42),
 ("kpi.fac.wo_corr","OT correctifs","Corrective WOs","nb","SUM","fac",1,False,23),
 ("kpi.fac.ot_backlog","Backlog d'OT","WO backlog","nb","LAST","fac",2,False,12),
 ("kpi.fac.mttr","MTTR","Mean Time To Repair","h","AVG","fac",2,False,6.5),
 ("kpi.fac.ot_opened","OT ouverts","WOs opened","nb","SUM","fac",2,False,65),
 ("kpi.fac.ot_closed","OT clôturés","WOs closed","nb","SUM","fac",2,False,58),
 ("kpi.fac.energy_int","Intensité énergétique","Energy intensity","kWh/m²","RATIO","fac",3,True,14),
 ("kpi.fac.util_cost","Coût utilities vs budget","Utilities vs budget","%","RATIO","fac",3,False,1.03),
 ("kpi.fac.consumption","Conso eau/élec","Water/power use","unité","SUM","fac",3,True,32000),
 ("kpi.fac.capex","Suivi CAPEX","CAPEX tracking","FCFA","SUM","fac",4,True,45000000),
 ("kpi.fac.capex_var","CAPEX vs budget","CAPEX vs budget","%","RATIO","fac",4,False,0.96),
 ("kpi.fac.capex_proj","Projets CAPEX en cours","CAPEX projects","nb","LAST","fac",4,False,4),
 ("kpi.fac.inspections","Inspections réalisées","Inspections done","nb","SUM","fac",5,False,8),
 ("kpi.fac.insp_conform","Conformité inspections","Inspection conformity","%","AVG","fac",5,False,0.90),
 ("kpi.fac.nonconform","Non-conformités","Non-conformities","nb","SUM","fac",5,False,5),
 ("kpi.fac.cleanliness","Score propreté","Cleanliness score","score","AVG","fac",6,False,8.4),
 ("kpi.fac.hygiene_audits","Audits hygiène","Hygiene audits","nb","SUM","fac",6,False,4),
 ("kpi.fac.clean_complaints","Réclamations propreté","Cleaning complaints","nb","SUM","fac",6,False,3),
 ("kpi.fac.sla","Conformité SLA","Vendor SLA","%","AVG","fac",7,False,0.93),
]
KMETA={k[0]:k for k in KPIS}
def kdept(k): return k.split(".")[1]
def kmetric(k): return k.split(".")[2]
def is_pct(u): return u=="%"

# lower-is-better metrics (RAG sens = DOWN)
DOWN={"fte_gap","ocr","vac_age","gla_vacant","dso","provision","cpa","spend_pct","discount",
 "tf","tg","util_cost","energy_int","mttr","ot_backlog","turnover","absence","ttf",
 "litige_count","litige_amount","litige_age","incident_sev","incidents","evac_time","theft",
 "fire_alarms","nonconform","clean_complaints","wo_corr"}

DEPT_COLLECTE={"rh":"C1","lease":"C2","rec":"C3","foot":"C4","mkt":"C5","com":"C6","hsse":"C7","fac":"C8"}
DEPT_REPORT_KEY={d[0]:d[0] for d in DEPTS}
DEPT_KEYS={dep:[k for (k,*_r) in KPIS if kdept(k)==dep] for dep in DEPT_COLLECTE}
def dept_kpi_row(key): return 8 + DEPT_KEYS[kdept(key)].index(key)
def collecte_last(dep): return 7 + len(DEPT_KEYS[dep])
def subs_of(dep): return [d[2] for d in DEPTS if d[0]==dep][0]

# ---------------------------------------------------------------- i18n chrome
I18N=[
 ("nav.title","Reporting mensuel — Cosmos","Monthly Report — Cosmos"),
 ("nav.subtitle","Cosmos Yopougon · EPLSA – Emergence Plaza","Cosmos Yopougon · EPLSA – Emergence Plaza"),
 ("nav.period","Période","Period"),("nav.grain","Granularité","Granularity"),
 ("nav.lang","Langue","Language"),("nav.summary","Sommaire","Contents"),
 ("nav.synthesis","Synthèse exécutive","Executive summary"),("nav.methodo","Méthodologie","Methodology"),
 ("grain.month","Mois","Month"),("grain.quarter","Trimestre","Quarter"),("grain.year","Année","Year"),
 ("col.kpi","Indicateur","KPI"),("col.unit","Unité","Unit"),
 ("col.current","Courant","Current"),("col.prev","Période préc.","Prior period"),
 ("col.py","N-1","PY"),("col.delta_prev","Δ% préc.","Δ% prior"),("col.delta_py","Δ% N-1","Δ% PY"),
 ("col.budget","Budget","Budget"),("col.var","Écart%","Var%"),("col.rag","Statut","Status"),
 ("col.trend","Tendance","Trend"),("col.comment","Commentaire","Comment"),
 ("rag.green","Conforme","On track"),("rag.amber","Vigilance","Watch"),("rag.red","Alerte","Alert"),
 ("rag.global","RAG global","Global RAG"),
 ("dept.rh","Capital Humain","Human Capital"),("dept.lease","Gestion locative","Leasing"),
 ("dept.rec","Recouvrement","Collections"),("dept.foot","Fréquentation","Footfall"),
 ("dept.mkt","Marketing","Marketing"),("dept.com","Commercial","Commercial"),
 ("dept.hsse","Sécurité / HSSE","Safety / HSSE"),("dept.fac","Facility Management","Facility Management"),
 ("msg.collect","Collecte","Collection"),("msg.data","Données","Data"),
 ("msg.help","1 chiffre par KPI / mois — le rapport se remplit seul.","One figure per KPI / month — the report fills itself."),
 ("msg.updated","Mis à jour","Updated"),("msg.confidential","Confidentiel — diffusion restreinte","Confidential — restricted"),
 ("msg.version","Version 3.0","Version 3.0"),("msg.headline","Chiffre phare","Headline"),
 ("msg.heatmap","Heatmap RAG — départements × KPI clé","RAG heatmap — departments × key KPI"),
 ("msg.highlights","Faits marquants","Highlights"),("msg.tracking","Suivi de saisie","Input tracking"),
 ("msg.hero","Indicateurs groupe","Group indicators"),
 ("narr.up","{kpi} en hausse de {delta} vs période précédente.","{kpi} up {delta} vs prior period."),
 ("narr.down","{kpi} en baisse de {delta} vs période précédente.","{kpi} down {delta} vs prior period."),
]
for k,fr,en,u,agg,dep,h2,b,base in KPIS: I18N.append((k,fr,en))
for dep,title,subs in DEPTS:
    for i,s_ in enumerate(subs,1): I18N.append((f"h2.{dep}.{i}", s_, s_))
MONTHS=[("Janvier","January"),("Février","February"),("Mars","March"),("Avril","April"),
 ("Mai","May"),("Juin","June"),("Juillet","July"),("Août","August"),("Septembre","September"),
 ("Octobre","October"),("Novembre","November"),("Décembre","December")]

# ---------------------------------------------------------------- sheets
SH={"M0":"M0 · Pilotage","M1":"M1 · Paramètres","M2":"M2 · i18n","M3":"M3 · Base mensuelle","M4":"M4 · Dictionnaire",
 "C1":"C1 · Capital Humain","C2":"C2 · Lease","C3":"C3 · Recouvrement","C4":"C4 · Footfall",
 "C5":"C5 · Marketing","C6":"C6 · Commercial","C7":"C7 · HSSE","C8":"C8 · Facility",
 "RPT":"Rapport Cosmos","M5":"M5 · Données graphiques"}
def q(code): return "'%s'" % SH[code]
N_MONTHS=36; BASE_FIRST_COL=4
def m3_col(i): return get_column_letter(BASE_FIRST_COL+i)

wb=Workbook(); wb.remove(wb.active)
ws={code:wb.create_sheet(title=name) for code,name in SH.items()}
CALC=ws["M5"]; CALC.sheet_view.showGridLines=False  # feuille de calcul dédiée (masquée) pour les séries de graphiques

def T(key):
    return ('=IFERROR(IF(LANG="EN",INDEX(I18N_EN,MATCH("%s",I18N_KEY,0)),'
            'INDEX(I18N_FR,MATCH("%s",I18N_KEY,0))),"%s")')%(key,key,key)
def Tx(key): return T(key)[1:]   # without leading '='

# ================================================================ M2
def build_m2():
    s=ws["M2"]; s.sheet_properties.tabColor=AMBER; s.sheet_view.showGridLines=False
    for r in range(1,260):
        for c in range(1,13): s.cell(r,c).fill=fill(DARK_BG)
    s["B1"]="M2 · i18n (FR/EN) + MOIS"; s["B1"].font=Font(name=UI,bold=True,size=14,color=AMBER)
    for j,h in enumerate(["KEY","FR","EN"]):
        c=s.cell(3,2+j,h); c.font=Font(name=UI,bold=True,color=DARK_BG); c.fill=fill(AMBER); c.alignment=Alignment(horizontal="center")
    r=4
    for key,fr,en in I18N:
        s.cell(r,2,key).font=Font(name=MONO,size=9,color=TXT_SEC_D)
        s.cell(r,3,fr).font=Font(name=UI,color=TXT_LIGHT)
        s.cell(r,4,en).font=Font(name=UI,color=TXT_LIGHT); r+=1
    last=r-1
    s.cell(3,6,"Clés vides (attendu 0)").font=Font(name=UI,color=TXT_SEC_D)
    s.cell(3,7).value=f'=SUMPRODUCT(--(C4:C{last}=""))+SUMPRODUCT(--(D4:D{last}=""))'
    s.cell(3,7).font=Font(name=UI,bold=True,color=AMBER)
    for j,h in enumerate(["#","FR","EN"]):
        c=s.cell(3,9+j,h); c.font=Font(name=UI,bold=True,color=DARK_BG); c.fill=fill(AMBER)
    for i,(fr,en) in enumerate(MONTHS):
        s.cell(4+i,9,i+1).font=Font(name=MONO,color=TXT_LIGHT)
        s.cell(4+i,10,fr).font=Font(name=UI,color=TXT_LIGHT)
        s.cell(4+i,11,en).font=Font(name=UI,color=TXT_LIGHT)
    s.column_dimensions["B"].width=28; s.column_dimensions["C"].width=34; s.column_dimensions["D"].width=34
    wb.defined_names.add(DefinedName("I18N_KEY",attr_text=f"{q('M2')}!$B$4:$B${last}"))
    wb.defined_names.add(DefinedName("I18N_FR", attr_text=f"{q('M2')}!$C$4:$C${last}"))
    wb.defined_names.add(DefinedName("I18N_EN", attr_text=f"{q('M2')}!$D$4:$D${last}"))
    wb.defined_names.add(DefinedName("MOIS_FR", attr_text=f"{q('M2')}!$J$4:$J$15"))
    wb.defined_names.add(DefinedName("MOIS_EN", attr_text=f"{q('M2')}!$K$4:$K$15"))
    s.conditional_formatting.add(f"C4:D{last}", FormulaRule(formula=['C4=""'], fill=fill("7f1d1d")))

# ================================================================ M0
def build_m0():
    s=ws["M0"]; s.sheet_properties.tabColor=AMBER; s.sheet_view.showGridLines=False
    for r in range(1,60):
        for c in range(1,14): s.cell(r,c).fill=fill(DARK_BG)
    s["B1"]="M0 · PUPITRE DE PILOTAGE"; s["B1"].font=Font(name=UI,bold=True,size=16,color=AMBER)
    s["B2"]="Sélecteurs globaux — pilotent tout le classeur"; s["B2"].font=Font(name=UI,size=10,color=TXT_SEC_D)
    def sel(row,label,value,nf=None):
        s.cell(row,2,label).font=Font(name=UI,bold=True,color=TXT_LIGHT)
        c=s.cell(row,4,value); c.fill=fill(DARK_INPUT); c.font=Font(name=MONO,bold=True,size=12,color=AMBER)
        c.alignment=Alignment(horizontal="center"); c.border=box(AMBER)
        if nf: c.number_format=nf
        return c
    sel(4,"🌐 Langue (LANG)","FR"); sel(5,"Granularité (GRAIN)","M")
    sel(6,"Année (ANNEE)",2026); sel(7,"Mois (MOIS)",6)
    s.cell(8,2,"Période (PERIODE)").font=Font(name=UI,bold=True,color=TXT_LIGHT)
    pc=s.cell(8,4); pc.value="=DATE(ANNEE,MOIS,1)"; pc.number_format='yyyy-mm'
    pc.font=Font(name=MONO,bold=True,color=AMBER); pc.alignment=Alignment(horizontal="center")
    for cell,f1 in (("D4",'"FR,EN"'),("D5",'"M,Q,Y"')):
        dv=DataValidation(type="list",formula1=f1,allow_blank=False); s.add_data_validation(dv); dv.add(cell)
    dvy=DataValidation(type="whole",operator="between",formula1=2024,formula2=2026); s.add_data_validation(dvy); dvy.add("D6")
    dvm=DataValidation(type="whole",operator="between",formula1=1,formula2=12); s.add_data_validation(dvm); dvm.add("D7")
    s.cell(5,6,"M = Mois · Q = Trimestre · Y = Année").font=Font(name=UI,size=9,italic=True,color=TXT_SEC_D)
    s.cell(8,6,"Libellé :").font=Font(name=UI,size=9,color=TXT_SEC_D)
    s.cell(8,7).value='=IF(LANG="EN",INDEX(MOIS_EN,MONTH(PERIODE)),INDEX(MOIS_FR,MONTH(PERIODE)))&" "&YEAR(PERIODE)'
    s.cell(8,7).font=Font(name=UI,bold=True,color=AMBER)
    s.cell(11,2,"MOTEUR TEMPOREL — bornes").font=Font(name=UI,bold=True,color=AMBER)
    bounds=[(12,"P_START",'=IF(GRAIN="M",PERIODE,IF(GRAIN="Q",DATE(ANNEE,(ROUNDUP(MOIS/3,0)-1)*3+1,1),DATE(ANNEE,1,1)))'),
     (13,"P_END",'=IF(GRAIN="M",PERIODE,IF(GRAIN="Q",EDATE(P_START,2),DATE(ANNEE,12,1)))'),
     (14,"PP_START",'=IF(GRAIN="M",EDATE(P_START,-1),IF(GRAIN="Q",EDATE(P_START,-3),EDATE(P_START,-12)))'),
     (15,"PP_END",'=IF(GRAIN="M",EDATE(P_END,-1),IF(GRAIN="Q",EDATE(P_END,-3),EDATE(P_END,-12)))'),
     (16,"PY_START",'=EDATE(P_START,-12)'),(17,"PY_END",'=EDATE(P_END,-12)')]
    for row,name,f in bounds:
        s.cell(row,2,name).font=Font(name=MONO,color=TXT_SEC_D)
        c=s.cell(row,4); c.value=f; c.number_format='yyyy-mm-dd'; c.font=Font(name=MONO,color=TXT_LIGHT)
        wb.defined_names.add(DefinedName(name,attr_text=f"{q('M0')}!$D${row}"))
    s.cell(20,2,"SUIVI DE SAISIE (mois courant)").font=Font(name=UI,bold=True,color=AMBER)
    for j,h in enumerate(["Département","Complétude %","État"]):
        c=s.cell(21,2+j,h); c.font=Font(name=UI,bold=True,color=DARK_BG); c.fill=fill(AMBER)
    r=22
    for dep,title,subs in DEPTS:
        code=DEPT_COLLECTE[dep]; last=collecte_last(dep)
        s.cell(r,2).value=T("dept."+dep); s.cell(r,2).font=Font(name=UI,color=TXT_LIGHT)
        s.cell(r,3).value=(f"=IFERROR(COUNT(INDEX({q(code)}!$D$8:$AM${last},0,MATCH(PERIODE,{q(code)}!$D$6:$AM$6,0)))"
                           f"/ROWS({q(code)}!$C$8:$C${last}),0)")
        s.cell(r,3).number_format='0%'; s.cell(r,3).font=Font(name=MONO,color=TXT_LIGHT); s.cell(r,3).alignment=Alignment(horizontal="center")
        st=s.cell(r,4); st.value=f'=IF(C{r}>=1,"✓",IF(C{r}>0,"⌛","✗"))'
        st.font=Font(name=UI,bold=True,color=TXT_LIGHT); st.alignment=Alignment(horizontal="center"); r+=1
    s.column_dimensions["B"].width=22; s.column_dimensions["C"].width=16; s.column_dimensions["D"].width=10
    s.column_dimensions["F"].width=14; s.column_dimensions["G"].width=22
    for n,cellref in (("LANG","$D$4"),("GRAIN","$D$5"),("ANNEE","$D$6"),("MOIS","$D$7"),("PERIODE","$D$8")):
        wb.defined_names.add(DefinedName(n,attr_text=f"{q('M0')}!{cellref}"))

# ================================================================ M1
def build_m1():
    s=ws["M1"]; s.sheet_properties.tabColor=AMBER; s.sheet_view.showGridLines=False
    for r in range(1,len(KPIS)+40):
        for c in range(1,24): s.cell(r,c).fill=fill(DARK_BG)
    s["B1"]="M1 · PARAMÈTRES (RAG · budget · listes)"; s["B1"].font=Font(name=UI,bold=True,size=14,color=AMBER)
    s.cell(3,2,"tbl_param_rag — seuils paramétrables").font=Font(name=UI,bold=True,color=AMBER)
    for j,h in enumerate(["KEY","SENS","SEUIL_VERT","SEUIL_AMBRE"]):
        c=s.cell(4,2+j,h); c.font=Font(name=UI,bold=True,color=DARK_BG); c.fill=fill(AMBER)
    r=5; rag_first=r
    for key,fr,en,u,agg,dep,h2,b,base in KPIS:
        sens="DOWN" if kmetric(key) in DOWN else "UP"
        sv,sa=(1.0,0.95) if sens=="UP" else (1.0,1.05)
        s.cell(r,2,key).font=Font(name=MONO,size=9,color=TXT_SEC_D)
        s.cell(r,3,sens).font=Font(name=MONO,color=TXT_LIGHT)
        s.cell(r,4,sv).number_format='0.00'; s.cell(r,4).font=Font(name=MONO,color=TXT_LIGHT)
        s.cell(r,5,sa).number_format='0.00'; s.cell(r,5).font=Font(name=MONO,color=TXT_LIGHT); r+=1
    rag_last=r-1
    wb.defined_names.add(DefinedName("RAG_KEY",attr_text=f"{q('M1')}!$B${rag_first}:$B${rag_last}"))
    wb.defined_names.add(DefinedName("RAG_SENS",attr_text=f"{q('M1')}!$C${rag_first}:$C${rag_last}"))
    wb.defined_names.add(DefinedName("RAG_SV",attr_text=f"{q('M1')}!$D${rag_first}:$D${rag_last}"))
    wb.defined_names.add(DefinedName("RAG_SA",attr_text=f"{q('M1')}!$E${rag_first}:$E${rag_last}"))
    bc0=8
    s.cell(3,bc0,"tbl_param_budget — budget mensuel (KPI × 12 mois)").font=Font(name=UI,bold=True,color=AMBER)
    s.cell(4,bc0-1,"KEY").font=Font(name=UI,bold=True,color=DARK_BG); s.cell(4,bc0-1).fill=fill(AMBER)
    for m in range(12):
        c=s.cell(4,bc0+m,m+1); c.font=Font(name=UI,bold=True,color=DARK_BG); c.fill=fill(AMBER); c.alignment=Alignment(horizontal="center")
    r=5; bud_first=r; bud_rows={}
    for key,fr,en,u,agg,dep,h2,b,base in KPIS:
        s.cell(r,bc0-1,key).font=Font(name=MONO,size=9,color=TXT_SEC_D)
        if b:
            for m in range(12):
                tgt=base*(1.0+0.01*m if agg=="SUM" else 1.0)
                val=round(tgt,4) if is_pct(u) else int(round(tgt))
                cc=s.cell(r,bc0+m,val); cc.font=Font(name=MONO,size=9,color=TXT_LIGHT)
                cc.number_format=F_PCT if is_pct(u) else '#,##0'
            bud_rows[key]=r
        r+=1
    bud_last=r-1
    wb.defined_names.add(DefinedName("BUD_MONTHS",attr_text=f"{q('M1')}!${get_column_letter(bc0)}$4:${get_column_letter(bc0+11)}$4"))
    build_m1.bud=(bc0,bud_first,bud_last,bud_rows)
    s.cell(len(KPIS)+8,2,"tbl_param_locataires (extrait)").font=Font(name=UI,bold=True,color=AMBER)
    s.column_dimensions["B"].width=22

# ================================================================ M3
def build_m3():
    s=ws["M3"]; s.sheet_properties.tabColor=AMBER; s.sheet_view.showGridLines=False
    s["B1"]="M3 · BASE MENSUELLE — consolidation (réf. collectes : saisie unique)"
    s["B1"].font=Font(name=UI,bold=True,size=12,color=AMBER)
    s.cell(3,2,"Clé KPI").font=Font(name=UI,bold=True,color=DARK_BG); s.cell(3,2).fill=fill(AMBER)
    s.cell(3,3,"Agg").font=Font(name=UI,bold=True,color=DARK_BG); s.cell(3,3).fill=fill(AMBER)
    for i in range(N_MONTHS):
        y=2024+i//12; m=i%12+1
        c=s.cell(3,BASE_FIRST_COL+i); c.value=f"=DATE({y},{m},1)"; c.number_format='yyyy-mm'
        c.font=Font(name=MONO,size=8,color=DARK_BG); c.fill=fill(AMBER); c.alignment=Alignment(horizontal="center")
    rowmap={}; r=4
    for key,fr,en,u,agg,dep,h2,b,base in KPIS:
        s.cell(r,2,key).font=Font(name=MONO,size=8,color=TXT_SEC_D)
        s.cell(r,3,agg).font=Font(name=MONO,size=8,color=TXT_SEC_D); rowmap[key]=r
        for i in range(N_MONTHS):
            cell=s.cell(r,BASE_FIRST_COL+i)
            cell.number_format=F_PCT if is_pct(u) else '#,##0'; cell.font=Font(name=MONO,size=8,color=TXT_LIGHT)
            cell.value=f"={q(DEPT_COLLECTE[dep])}!{m3_col(i)}{dept_kpi_row(key)}"
        r+=1
    s.column_dimensions["B"].width=24
    a=get_column_letter(BASE_FIRST_COL); z=get_column_letter(BASE_FIRST_COL+N_MONTHS-1)
    wb.defined_names.add(DefinedName("HDR",attr_text=f"{q('M3')}!${a}$3:${z}$3"))
    build_m3.rowmap=rowmap

def m3_row_range(key):
    r=build_m3.rowmap[key]; a=get_column_letter(BASE_FIRST_COL); z=get_column_letter(BASE_FIRST_COL+N_MONTHS-1)
    return f"{q('M3')}!${a}${r}:${z}${r}"

# ================================================================ temporal formulas
def pval(rng,d1,d2,agg):
    if agg=="SUM": return f'SUMIFS({rng},HDR,">="&{d1},HDR,"<="&{d2})'
    if agg=="LAST": return f'INDEX({rng},MATCH({d2},HDR,0))'
    return f'AVERAGEIFS({rng},HDR,">="&{d1},HDR,"<="&{d2})'
def cur_f(key,agg,sc=""): return f'=IFERROR(({pval(m3_row_range(key),"P_START","P_END",agg)}){sc},"")'
def prev_f(key,agg,sc=""): return f'=IFERROR(({pval(m3_row_range(key),"PP_START","PP_END",agg)}){sc},"")'
def py_f(key,agg,sc=""): return f'=IFERROR(({pval(m3_row_range(key),"PY_START","PY_END",agg)}){sc},"")'
def scfor(u): return "/1000000" if u=="FCFA" else ""
def budget_f(key,agg,sc=""):
    bc0,bf,bl,brows=build_m1.bud
    if key not in brows: return None
    br=brows[key]; a=get_column_letter(bc0); z=get_column_letter(bc0+11)
    rng=f"{q('M1')}!${a}${br}:${z}${br}"; mask='(BUD_MONTHS>=MONTH(P_START))*(BUD_MONTHS<=MONTH(P_END))'
    if agg=="SUM": return f'=IFERROR((SUMPRODUCT({mask}*{rng})){sc},"")'
    return f'=IFERROR((SUMPRODUCT({mask}*{rng})/SUMPRODUCT({mask}*1)){sc},"")'

# ================================================================ collecte (generic, all depts)
def build_collecte(code,dept):
    s=ws[code]; s.sheet_properties.tabColor="334155"; s.sheet_view.showGridLines=False
    keys=DEPT_KEYS[dept]; ncols=BASE_FIRST_COL+N_MONTHS
    for r in range(1,8+2*len(keys)+6):
        for c in range(2,ncols+1): s.cell(r,c).fill=fill(DARK_BG)
    s.cell(1,3).value='="▌ "&'+Tx("msg.collect")+'&" · "&'+Tx("dept."+dept)
    s.cell(1,3).font=Font(name=UI,bold=True,size=14,color=AMBER)
    s.cell(2,3).value='='+Tx("msg.help"); s.cell(2,3).font=Font(name=UI,size=9,italic=True,color=TXT_SEC_D)
    s.cell(6,3,"KPI \\ Mois").font=Font(name=UI,bold=True,color=DARK_BG); s.cell(6,3).fill=fill(AMBER)
    for i in range(N_MONTHS):
        y=2024+i//12; m=i%12+1
        c=s.cell(6,BASE_FIRST_COL+i); c.value=f"=DATE({y},{m},1)"; c.number_format='mmm yy'
        c.font=Font(name=MONO,size=8,color=DARK_BG); c.fill=fill(AMBER); c.alignment=Alignment(horizontal="center")
    r=8
    for key in keys:
        _,fr,en,u,agg,dep,h2,b,base=KMETA[key]
        s.cell(r,3).value=T(key); s.cell(r,3).font=Font(name=UI,color=TXT_LIGHT,size=9)
        for i in range(N_MONTHS):
            val=base*(1.0+random.uniform(-0.05,0.15)*(i/(N_MONTHS-1))+random.uniform(-0.02,0.02))
            if is_pct(u): val=round(max(0,min(1.3,val)),4)
            elif u in("h","ans","TF","TG","min","kWh/m²","v/m²","score","mois") or agg=="RATIO": val=round(val,2)
            else: val=int(round(val))
            cell=s.cell(r,BASE_FIRST_COL+i,val); cell.fill=fill(DARK_INPUT)
            cell.font=Font(name=MONO,size=8,color=AMBER); cell.number_format=F_PCT if is_pct(u) else '#,##0'
            cell.protection=Protection(locked=False)
        r+=1
    last=r-1
    s.cell(last+2,3,"💬 Commentaires (KPI × mois)").font=Font(name=UI,bold=True,color=AMBER)
    cr=last+3
    for key in keys:
        s.cell(cr,3).value=T(key); s.cell(cr,3).font=Font(name=UI,size=8,color=TXT_SEC_D)
        for i in range(N_MONTHS):
            cc=s.cell(cr,BASE_FIRST_COL+i); cc.fill=fill(DARK_PANEL)
            cc.protection=Protection(locked=False); cc.font=Font(name=UI,size=8,color=TXT_LIGHT)
        cr+=1
    s.column_dimensions["C"].width=26
    for i in range(N_MONTHS): s.column_dimensions[get_column_letter(BASE_FIRST_COL+i)].width=9
    s.freeze_panes="D8"
    a=get_column_letter(BASE_FIRST_COL); z=get_column_letter(BASE_FIRST_COL+N_MONTHS-1)
    s.conditional_formatting.add(f"{a}6:{z}6", FormulaRule(formula=[f'{a}$6=PERIODE'], fill=fill(AMBER)))
    s.conditional_formatting.add(f"{a}8:{z}{last}", FormulaRule(formula=[f'AND({a}$6=PERIODE,ISBLANK({a}8))'], fill=fill("7f1d1d")))
    s.protection.sheet=True; s.protection.password="cosmos"; s.protection.formatCells=False

# ================================================================ SINGLE-SHEET REPORT
ANCHORS={}; TOC_LINKS=[]; PANELS=[]; HELP_COL=19  # helper chart data in cols S..AD (off print)
def panelize_white(s,r0,r1):
    tint="FF"+PAGE_TINT; bc=CARD_BORDER
    for r in range(r0,r1+1):
        for c in range(3,15):
            cl=s.cell(r,c)
            if cl.fill.fgColor.rgb in (tint,None,"00000000"): cl.fill=fill(WHITE)
    for c in range(3,15):
        t=s.cell(r0,c); t.border=Border(top=side(bc),left=(side(bc) if c==3 else t.border.left),right=(side(bc) if c==14 else t.border.right),bottom=t.border.bottom)
        b=s.cell(r1,c); b.border=Border(bottom=side(bc),left=(side(bc) if c==3 else None),right=(side(bc) if c==14 else None))
    for r in range(r0+1,r1):
        l=s.cell(r,3); l.border=Border(left=side(bc),top=l.border.top,right=l.border.right,bottom=l.border.bottom)
        rt=s.cell(r,14); rt.border=Border(right=side(bc),top=rt.border.top,left=rt.border.left,bottom=rt.border.bottom)
def chart_white(ch):
    try: ch.graphical_properties=GraphicalProperties(solidFill=WHITE)
    except Exception: pass

def write_kpi_row(s,rr,key):
    _,fr,en,u,agg,dep,h2,b,base=KMETA[key]
    s.cell(rr,3).value=T(key); s.cell(rr,3).font=Font(name=UI,size=9,color=TXT_PRIM)
    s.cell(rr,4).value=u; s.cell(rr,4).font=Font(name=UI,size=9,color=TXT_SEC); s.cell(rr,4).alignment=Alignment(horizontal="center")
    sc=scfor(u)
    nf=F_PCT if is_pct(u) else (FCM if u=="FCFA" else ('#,##0' if u=="FCFA/m²" else F_NUM1))
    cur=s.cell(rr,5); cur.value=cur_f(key,agg,sc); cur.number_format=nf
    prv=s.cell(rr,6); prv.value=prev_f(key,agg,sc); prv.number_format=nf
    dpv=s.cell(rr,7); dpv.value=f'=IFERROR((E{rr})/(F{rr})-1,"")'; dpv.number_format=F_DELTA
    py=s.cell(rr,8); py.value=py_f(key,agg,sc); py.number_format=nf
    dpy=s.cell(rr,9); dpy.value=f'=IFERROR((E{rr})/(H{rr})-1,"")'; dpy.number_format=F_DELTA
    bf=budget_f(key,agg,sc); bud=s.cell(rr,10)
    if bf: bud.value=bf; bud.number_format=nf
    var=s.cell(rr,11); var.value=f'=IFERROR((E{rr})/(J{rr})-1,"")'; var.number_format=F_DELTA
    ref=f'IF(N(J{rr})>0,J{rr},H{rr})'
    sens=f'INDEX(RAG_SENS,MATCH("{key}",RAG_KEY,0))'; sv=f'INDEX(RAG_SV,MATCH("{key}",RAG_KEY,0))'; sa=f'INDEX(RAG_SA,MATCH("{key}",RAG_KEY,0))'
    rx=f'(E{rr}/{ref})'
    tok=s.cell(rr,15)
    tok.value=(f'=IFERROR(IF({ref}=0,"",IF({sens}="UP",IF({rx}>={sv},"green",IF({rx}>={sa},"amber","red")),'
               f'IF({rx}<={sv},"green",IF({rx}<={sa},"amber","red")))),"")')
    tok.font=Font(color=WHITE,size=8)
    rag=s.cell(rr,12)
    rag.value=f'=IF(O{rr}="","",IF(O{rr}="green",{Tx("rag.green")},IF(O{rr}="amber",{Tx("rag.amber")},{Tx("rag.red")})))'
    rag.alignment=Alignment(horizontal="center"); rag.font=Font(name=UI,bold=True,size=9)
    for cc in (cur,prv,py,bud,var,dpv,dpy): cc.font=Font(name=MONO,size=9,color=TXT_PRIM)

def render_cards(s,row,keys):
    keys=keys[:4]; slots=[3,9]   # two hero cards per row, each spanning 6 cols (C:H, I:N)
    for j,key in enumerate(keys):
        _,fr,en,u,agg,dep,h2,b,base=KMETA[key]
        base_r=row+(j//2)*5; col=slots[j%2]; cend=col+5
        for c in range(col,cend+1): s.cell(base_r,c).fill=fill(PRIMARY)
        s.row_dimensions[base_r].height=5
        place_icon(s,"cat_"+UCAT.get(u,"gauge"),f"{get_column_letter(col)}{base_r+1}",16)
        lab=s.cell(base_r+1,col); lab.value='="       "&'+Tx(key); lab.font=Font(name=UI,size=10,color=TXT_SEC); lab.fill=fill(CARD)
        v=s.cell(base_r+2,col); v.value=cur_f(key,agg,scfor(u))
        v.number_format=F_PCT if is_pct(u) else (FCM if u=="FCFA" else ('#,##0' if u=="FCFA/m²" else F_NUM1))
        v.font=Font(name=MONO,bold=True,size=26,color=PRIMARY); v.fill=fill(CARD); v.alignment=Alignment(horizontal="left",vertical="center")
        d=s.cell(base_r+3,col); d.value='=IFERROR(('+cur_f(key,agg)[1:]+')/('+prev_f(key,agg)[1:]+')-1,"")'
        d.number_format=F_DELTA; d.font=Font(name=UI,bold=True,size=10,color=TXT_SEC); d.fill=fill(CARD); d.alignment=Alignment(horizontal="left")
        cl=get_column_letter(col)
        s.conditional_formatting.add(f"{cl}{base_r+3}", FormulaRule(formula=[f'{cl}{base_r+3}>0'], fill=fill(G_BG), font=Font(name=UI,bold=True,size=10,color=G_TXT)))
        s.conditional_formatting.add(f"{cl}{base_r+3}", FormulaRule(formula=[f'{cl}{base_r+3}<0'], fill=fill(R_BG), font=Font(name=UI,bold=True,size=10,color=R_TXT)))
        for rr in (base_r+1,base_r+2,base_r+3):
            s.merge_cells(start_row=rr,start_column=col,end_row=rr,end_column=cend)
        for rr in range(base_r,base_r+4):
            for c in range(col,cend+1):
                s.cell(rr,c).border=Border(left=side(CARD_BORDER) if c==col else None,
                                 right=side(CARD_BORDER) if c==cend else None,
                                 bottom=side(CARD_BORDER) if rr==base_r+3 else None)
        s.row_dimensions[base_r+1].height=18; s.row_dimensions[base_r+2].height=42; s.row_dimensions[base_r+3].height=20
    return row+((len(keys)+1)//2)*5+1

def add_chart(s,key,anchor_row,helper_row):
    u=KMETA[key][3]; rng=m3_row_range(key); HS=HELP_COL
    lbl=helper_row; v1=helper_row+1
    for k in range(12):
        col=HS+k; idx=f'(MATCH(P_END,HDR,0)-11+{k})'
        CALC.cell(lbl,col).value=f'=IFERROR(IF({idx}<1,"",TEXT(INDEX(HDR,{idx}),"mmm yy")),"")'
        c1=CALC.cell(v1,col); c1.value=f'=IFERROR(IF({idx}<1,NA(),INDEX({rng},{idx})),NA())'
        c1.number_format=F_PCT if u=="%" else '#,##0'
    cats=Reference(CALC,min_col=HS,max_col=HS+11,min_row=lbl,max_row=lbl)
    ch=LineChart(); ch.height=5.4; ch.width=24; ch.legend=None; ch.style=2
    ch.add_data(Reference(CALC,min_col=HS,max_col=HS+11,min_row=v1,max_row=v1),from_rows=True); ch.set_categories(cats)
    try:
        ser=ch.series[0]; ser.graphicalProperties.line.solidFill=SUCCESS
        ser.graphicalProperties.line.width=26000; ser.smooth=True
    except Exception: pass
    chart_white(ch); s.add_chart(ch,f"C{anchor_row}")

def add_detail_chart(s,cols,hr,bot,anchor_row):
    """Donut for share tables, column chart for monthly/amount tables. Returns next free row."""
    if bot<=hr: return anchor_row
    li=pi=vi=None
    for j,(h,k) in enumerate(cols):
        if k=="t" and li is None: li=j
        if k=="p" and pi is None: pi=j
        if k in("f","n") and vi is None: vi=j
    if li is None: return anchor_row
    cats=Reference(s,min_col=3+li,max_col=3+li,min_row=hr+1,max_row=bot)
    n=bot-hr
    if pi is not None and n<=9:
        ch=DoughnutChart(); ch.height=6.2; ch.width=10; ch.holeSize=58
        ch.add_data(Reference(s,min_col=3+pi,max_col=3+pi,min_row=hr+1,max_row=bot)); ch.set_categories(cats)
        ch.dataLabels=DataLabelList(); ch.dataLabels.showPercent=False; ch.dataLabels.showVal=False
        s.add_chart(ch,f"C{anchor_row}"); return anchor_row+max(7,n+1)
    if vi is not None and n>=3:
        cb=BarChart(); cb.type="col"; cb.height=6.2; cb.width=18; cb.legend=None
        cb.add_data(Reference(s,min_col=3+vi,max_col=3+vi,min_row=hr+1,max_row=bot)); cb.set_categories(cats)
        try: cb.series[0].graphicalProperties.solidFill=PRIMARY
        except Exception: pass
        s.add_chart(cb,f"C{anchor_row}"); return anchor_row+max(7,n+1)
    return anchor_row

def _banner(s,row,num,fr,en,icon="▸"):
    for c in range(3,16): s.cell(row,c).fill=fill(PRIMARY)
    c=s.cell(row,3); c.value='="%s  %s  "&IF(LANG="EN","%s","%s")'%(icon,num,en,fr)
    c.font=Font(name=UI,bold=True,size=11,color=WHITE); c.alignment=Alignment(vertical="center")
    s.row_dimensions[row].height=19
    return row+1

def _comment_box(s,row,body_formula,editable=False):
    cz=s.cell(row,3); cz.value='="💬 "&'+Tx("col.comment"); cz.font=Font(name=UI,bold=True,size=9,color=DARK_BG)
    for cc in range(3,16): s.cell(row,cc).fill=fill(AMBER)
    row+=1
    cb=s.cell(row,3); cb.value=body_formula; cb.font=Font(name=UI,size=10,color=TXT_PRIM,italic=editable)
    cb.alignment=Alignment(wrap_text=True,vertical="top")
    s.merge_cells(f"C{row}:O{row+2}")
    for cc in range(3,16):
        s.cell(row,cc).border=Border(left=side(CARD_BORDER),right=side(CARD_BORDER),top=side(CARD_BORDER))
        s.cell(row+2,cc).border=Border(left=side(CARD_BORDER),right=side(CARD_BORDER),bottom=side(CARD_BORDER))
    return row+4

_GH=[1]   # gauge helper-row cursor (col GAUGE_COL, off print)
GAUGE_COL=33
def add_gauge(s,key,anchor_col,anchor_row):
    agg=KMETA[key][4]; hr=_GH[0]; _GH[0]+=3; HG=GAUGE_COL
    gvref=f"{q('M5')}!{get_column_letter(HG)}{hr}"
    CALC.cell(hr,HG).value=cur_f(key,agg)
    CALC.cell(hr+1,HG).value=f'=MAX(0,1-{gvref})'
    ch=DoughnutChart(); ch.holeSize=72; ch.height=4.0; ch.width=4.4; ch.legend=None
    ch.add_data(Reference(CALC,min_col=HG,max_col=HG,min_row=hr,max_row=hr+1),titles_from_data=False)
    ch.series[0].data_points=[DataPoint(idx=0,spPr=GraphicalProperties(solidFill=SUCCESS)),
                              DataPoint(idx=1,spPr=GraphicalProperties(solidFill="E0E9E5"))]
    lab=s.cell(anchor_row,anchor_col); lab.value=T(key); lab.font=Font(name=UI,size=9,bold=True,color=TXT_SEC)
    chart_white(ch); s.add_chart(ch,f"{get_column_letter(anchor_col)}{anchor_row+1}")
    cap=s.cell(anchor_row+8,anchor_col); cap.value=cur_f(key,KMETA[key][4]); cap.number_format=F_PCT
    cap.font=Font(name=MONO,bold=True,size=14,color=PRIMARY); cap.alignment=Alignment(horizontal="center")
    s.merge_cells(start_row=anchor_row+8,start_column=anchor_col,end_row=anchor_row+8,end_column=anchor_col+2)

_CH=[300]   # combo helper-row cursor (HELP_COL region, off print)
def add_combo(s,key,anchor_row):
    agg=KMETA[key][4]; mrow=m3_row_range(key)
    bc0,bf,bl,brows=build_m1.bud; br=brows.get(key)
    a=get_column_letter(bc0); z=get_column_letter(bc0+11)
    budrng=f"{q('M1')}!${a}${br}:${z}${br}" if br else None
    HS=HELP_COL; _CH[0]+=4; lbl=_CH[0]; act=lbl+1; bud=lbl+2
    for k in range(12):
        col=HS+k; idx=f'(MATCH(P_END,HDR,0)-11+{k})'
        CALC.cell(lbl,col).value=f'=IFERROR(IF({idx}<1,"",TEXT(INDEX(HDR,{idx}),"mmm yy")),"")'
        CALC.cell(act,col).value=f'=IFERROR(IF({idx}<1,NA(),INDEX({mrow},{idx})),NA())'
        CALC.cell(bud,col).value=(f'=IFERROR(IF({idx}<1,NA(),INDEX({budrng},MONTH(INDEX(HDR,{idx})))),NA())' if budrng else "=NA()")
    cats=Reference(CALC,min_col=HS,max_col=HS+11,min_row=lbl,max_row=lbl)
    bar=BarChart(); bar.type="col"; bar.height=5.6; bar.width=24; bar.legend=None
    bar.add_data(Reference(CALC,min_col=HS,max_col=HS+11,min_row=act,max_row=act),from_rows=True); bar.set_categories(cats)
    try: bar.series[0].graphicalProperties.solidFill=PRIMARY_XL
    except Exception: pass
    ln=LineChart(); ln.add_data(Reference(CALC,min_col=HS,max_col=HS+11,min_row=bud,max_row=bud),from_rows=True)
    try:
        ln.series[0].graphicalProperties.line.solidFill=WARNING; ln.series[0].graphicalProperties.line.width=28000; ln.series[0].smooth=True
    except Exception: pass
    bar += ln
    chart_white(bar); s.add_chart(bar,f"C{anchor_row}")

def render_budget(s,row,num,fr,en,dept):
    row=_banner(s,row,num,fr,en,icon="🎯")
    bkeys=[k for k in DEPT_KEYS[dept] if KMETA[k][7]]
    lead=bkeys[0] if bkeys else DEPT_KEYS[dept][0]
    s.cell(row,3).value='="▌ "&IF(LANG="EN","Actual vs Budget — 12 months","Réalisé vs Budget — 12 mois")'
    s.cell(row,3).font=Font(name=UI,bold=True,size=10,color=PRIMARY); row+=1
    add_combo(s,lead,row); row+=12
    if not bkeys: return _comment_box(s,row,'=IF(LANG="EN","No budget data.","Pas de budget.")',editable=True)
    hr=row
    for j,h in enumerate(["col.kpi","col.unit","col.current","col.budget","col.var","col.rag"]):
        c=s.cell(hr,3+j); c.value=T(h); c.font=Font(name=UI,bold=True,size=9,color=WHITE); c.fill=fill(PRIMARY); c.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=hr+1
    for key in bkeys:
        _,fr2,en2,u,agg,dep,h2,b,base=KMETA[key]; sc=scfor(u)
        nf=F_PCT if is_pct(u) else (FCM if u=="FCFA" else ('#,##0' if u=="FCFA/m²" else F_NUM1))
        s.cell(rr,3).value=T(key); s.cell(rr,3).font=Font(name=UI,size=9,color=TXT_PRIM)
        s.cell(rr,4).value=u; s.cell(rr,4).font=Font(name=UI,size=9,color=TXT_SEC); s.cell(rr,4).alignment=Alignment(horizontal="center")
        cur=s.cell(rr,5); cur.value=cur_f(key,agg,sc); cur.number_format=nf; cur.font=Font(name=MONO,size=9,color=TXT_PRIM)
        bud=s.cell(rr,6); bud.value=budget_f(key,agg,sc); bud.number_format=nf; bud.font=Font(name=MONO,size=9,color=TXT_PRIM)
        var=s.cell(rr,7); var.value=f'=IFERROR((E{rr})/(F{rr})-1,"")'; var.number_format=F_DELTA; var.font=Font(name=MONO,size=9)
        sens=f'INDEX(RAG_SENS,MATCH("{key}",RAG_KEY,0))'; sv=f'INDEX(RAG_SV,MATCH("{key}",RAG_KEY,0))'; sa=f'INDEX(RAG_SA,MATCH("{key}",RAG_KEY,0))'
        rx=f'(E{rr}/F{rr})'
        tok=s.cell(rr,16); tok.value=(f'=IFERROR(IF(F{rr}=0,"",IF({sens}="UP",IF({rx}>={sv},"green",IF({rx}>={sa},"amber","red")),'
                   f'IF({rx}<={sv},"green",IF({rx}<={sa},"amber","red")))),"")'); tok.font=Font(color=WHITE,size=8)
        rag=s.cell(rr,8); rag.value=f'=IF(P{rr}="","",IF(P{rr}="green",{Tx("rag.green")},IF(P{rr}="amber",{Tx("rag.amber")},{Tx("rag.red")})))'
        rag.alignment=Alignment(horizontal="center"); rag.font=Font(name=UI,bold=True,size=9); rr+=1
    bot=rr-1
    s.conditional_formatting.add(f"C{hr+1}:H{bot}", FormulaRule(formula=['MOD(ROW(),2)=0'], fill=fill(BAND)))
    for st,bg,tx in (("green",G_BG,G_TXT),("amber",A_BG,A_TXT),("red",R_BG,R_TXT)):
        s.conditional_formatting.add(f"H{hr+1}:H{bot}", FormulaRule(formula=[f'$P{hr+1}="{st}"'], fill=fill(bg), font=Font(color=tx,bold=True)))
    return _comment_box(s,bot+1,'=IF(LANG="EN","Budget variance commentary.","Commentaire écart budgétaire.")',editable=True)

def render_kpi_section(s,row,num,fr,en,keys):
    row=_banner(s,row,num,fr,en); hr=row
    for j,h in enumerate(["col.kpi","col.unit","col.current","col.prev","col.delta_prev","col.py","col.delta_py","col.budget","col.var","col.rag"]):
        c=s.cell(hr,3+j); c.value=T(h); c.font=Font(name=UI,bold=True,size=9,color=WHITE)
        c.fill=fill(PRIMARY); c.alignment=Alignment(horizontal="center",wrap_text=True)
    rr=hr+1
    for key in keys: write_kpi_row(s,rr,key); rr+=1
    bot=rr-1
    s.conditional_formatting.add(f"C{hr+1}:L{bot}", FormulaRule(formula=['MOD(ROW(),2)=0'], fill=fill(BAND)))
    for st,bg,tx in (("green",G_BG,G_TXT),("amber",A_BG,A_TXT),("red",R_BG,R_TXT)):
        s.conditional_formatting.add(f"L{hr+1}:L{bot}", FormulaRule(formula=[f'$O{hr+1}="{st}"'], fill=fill(bg), font=Font(color=tx,bold=True)))
    row=bot+1; a=hr+1
    fm=s.cell(row,3); fm.value='="◆ "&'+Tx("msg.highlights"); fm.font=Font(name=UI,bold=True,size=9,color=AMBER); row+=1
    s.cell(row,3).value=f'=IFERROR("▲ "&INDEX(C{a}:C{bot},MATCH(MAX(G{a}:G{bot}),G{a}:G{bot},0))&"   "&TEXT(MAX(G{a}:G{bot}),"+0.0%;-0.0%"),"")'
    s.cell(row,3).font=Font(name=UI,size=9,color=G_TXT)
    s.cell(row,9).value=f'=IFERROR("▼ "&INDEX(C{a}:C{bot},MATCH(MIN(G{a}:G{bot}),G{a}:G{bot},0))&"   "&TEXT(MIN(G{a}:G{bot}),"+0.0%;-0.0%"),"")'
    s.cell(row,9).font=Font(name=UI,size=9,color=R_TXT); row+=1
    fk=keys[0]; dep0=kdept(fk); idx=DEPT_KEYS[dep0].index(fk)
    comref=q(DEPT_COLLECTE[dep0])+"!$D$"+str(collecte_last(dep0)+3+idx)
    narr='SUBSTITUTE(SUBSTITUTE(IF(G'+str(a)+'>=0,'+Tx("narr.up")+','+Tx("narr.down")+'),"{kpi}",'+Tx(fk)+'),"{delta}",TEXT(ABS(G'+str(a)+'),"0.0%"))'
    return _comment_box(s,row,'=IF('+comref+'<>"",'+comref+',IFERROR('+narr+',""))')

def render_detail(s,row,num,fr,en,cols,rows):
    row=_banner(s,row,num,fr,en); hr=row; ncol=len(cols)
    for j,(htxt,kind) in enumerate(cols):
        c=s.cell(hr,3+j,htxt); c.font=Font(name=UI,bold=True,size=9,color=WHITE); c.fill=fill(PRIMARY)
        c.alignment=Alignment(horizontal=("left" if kind=="t" else "center"),wrap_text=True)
    rr=hr+1
    for rowvals in rows:
        for j,(val,(htxt,kind)) in enumerate(zip(rowvals,cols)):
            c=s.cell(rr,3+j,val)
            if kind=="f": c.number_format='#,##0'; c.font=Font(name=MONO,size=9); c.alignment=Alignment(horizontal="right")
            elif kind=="p": c.number_format=F_PCT; c.font=Font(name=MONO,size=9); c.alignment=Alignment(horizontal="right")
            elif kind=="n": c.number_format='#,##0'; c.font=Font(name=MONO,size=9); c.alignment=Alignment(horizontal="right")
            else: c.font=Font(name=UI,size=9,color=TXT_PRIM); c.alignment=Alignment(horizontal="left")
        rr+=1
    bot=rr-1
    last_col=get_column_letter(2+ncol)
    s.conditional_formatting.add(f"C{hr+1}:{last_col}{bot}", FormulaRule(formula=['MOD(ROW(),2)=0'], fill=fill(BAND)))
    # data bars on percent / fcfa magnitude columns (premium gauge feel)
    for j,(htxt,kind) in enumerate(cols):
        col=get_column_letter(3+j)
        if kind=="f": s.column_dimensions[col].width=18
        elif kind=="t": s.column_dimensions[col].width=24
        if kind in ("p","f","n") and bot>hr:
            try:
                from openpyxl.formatting.rule import DataBarRule
                s.conditional_formatting.add(f"{col}{hr+1}:{col}{bot}",
                    DataBarRule(start_type="min",end_type="max",color=(SUCCESS if kind=="p" else PRIMARY_XL),showValue=True))
            except Exception: pass
    row=add_detail_chart(s,cols,hr,bot,bot+2)
    return _comment_box(s,row,'=IF(LANG="EN","Commentary / data source: management systems (demo).","Commentaire / source : systèmes de gestion (démo).")',editable=True)

def render_org(s,row,num,fr,en,nodes):
    row=_banner(s,row,num,fr,en)
    for lvl,fr2,en2 in nodes:
        c=s.cell(row,3+lvl); c.value='=IF(LANG="EN","%s","%s")'%(en2,fr2)
        c.font=Font(name=UI,bold=(lvl==0),size=(11 if lvl==0 else 10),color=(AMBER if lvl==0 else TXT_PRIM))
        c.fill=fill(HDR_TBL if lvl<=1 else WHITE); c.border=box(CARD_BORDER); row+=1
    return _comment_box(s,row+1,'=IF(LANG="EN","Org chart — demo.","Organigramme — démo.")',editable=True)

def render_note(s,row,num,fr,en):
    row=_banner(s,row,num,fr,en)
    return _comment_box(s,row,'=IF(LANG="EN","Commentary.","Commentaire.")',editable=True)

def render_dept(s,dept,dnum,row,helper):
    keys=DEPT_KEYS[dept]; ANCHORS[dept]=row
    for c in range(3,16): s.cell(row,c).fill=fill(PRIMARY)
    s.row_dimensions[row].height=28
    place_icon(s,"dep_"+dept,f"C{row}",26)
    h=s.cell(row,3); h.value=('="         %d.   "&'%dnum)+Tx("dept."+dept)+'&"   ·   "&IF(LANG="EN",INDEX(MOIS_EN,MONTH(PERIODE)),INDEX(MOIS_FR,MONTH(PERIODE)))&" "&YEAR(PERIODE)'
    h.font=Font(name=UI,bold=True,size=15,color=WHITE); h.alignment=Alignment(vertical="center")
    bl=s.cell(row,15); bl.value='="↑ "&'+Tx("nav.summary"); bl.font=Font(name=UI,size=8,color="CFE0D9",underline="single")
    bl.fill=fill(PRIMARY); bl.hyperlink="#"+q("RPT")+"!C1"; bl.alignment=Alignment(horizontal="right",vertical="center"); row+=1
    sub=s.cell(row,3); sub.value='=""&'+Tx("nav.grain")+'&": "&IF(GRAIN="M",'+Tx("grain.month")+',IF(GRAIN="Q",'+Tx("grain.quarter")+','+Tx("grain.year")+'))'
    sub.font=Font(name=UI,size=9,italic=True,color=TXT_SEC); row+=2
    row=render_cards(s,row,keys[:4])
    lead=keys[0]
    s.cell(row,3).value='="▌ "&'+Tx(lead)+'&" — 12 mois / 12M"'
    s.cell(row,3).font=Font(name=UI,bold=True,size=10,color=PRIMARY)
    add_chart(s,lead,row+1,helper); helper+=3; row+=12
    gk=[k for k in keys if KMETA[k][3]=="%"][:3]
    if gk:
        s.cell(row,3).value='="▌ "&IF(LANG="EN","Key gauges","Jauges clés")'
        s.cell(row,3).font=Font(name=UI,bold=True,size=10,color=PRIMARY); row+=1
        for gi,gkey in enumerate(gk): add_gauge(s,gkey,3+gi*4,row)
        row+=12
    for num,fr,en,typ,payload in SECTIONS.get(dept,[]):
        sstart=row
        if typ=="kpi": row=render_kpi_section(s,row,num,fr,en,payload)
        elif typ=="detail": row=render_detail(s,row,num,fr,en,payload[0],payload[1])
        elif typ=="budget": row=render_budget(s,row,num,fr,en,payload)
        elif typ=="org": row=render_org(s,row,num,fr,en,payload)
        else: row=render_note(s,row,num,fr,en)
        PANELS.append((sstart,row-1)); row+=1
    for cc in range(3,16): s.cell(row,cc).fill=fill(BAND)
    s.row_breaks.append(Break(id=row)); row+=2
    return row,helper

def render_front_cover(s,row):
    ANCHORS['cover']=row
    for rr in range(row,row+45):
        for c in range(3,15): s.cell(rr,c).fill=fill(PRIMARY)
    try:
        img=XLImage(LOGO_PATH); img.width=430; img.height=119; s.add_image(img,"D"+str(row+4))
    except Exception: pass
    t=s.cell(row+13,3); t.value='='+Tx("nav.title"); t.font=Font(name=WORD,bold=True,size=40,color=WHITE)
    s.merge_cells(start_row=row+13,start_column=3,end_row=row+13,end_column=14); s.row_dimensions[row+13].height=46
    st=s.cell(row+16,3); st.value='='+Tx("nav.subtitle"); st.font=Font(name=UI,size=14,color="CFE0D9")
    s.merge_cells(start_row=row+16,start_column=3,end_row=row+16,end_column=14)
    for c in range(3,8): s.cell(row+18,c).fill=fill(WARNING)
    pr=s.cell(row+20,3); pr.value='=""&'+Tx("nav.period")+'&" : "&IF(LANG="EN",INDEX(MOIS_EN,MONTH(PERIODE)),INDEX(MOIS_FR,MONTH(PERIODE)))&" "&YEAR(PERIODE)'
    pr.font=Font(name=UI,bold=True,size=18,color=WARNING); s.row_dimensions[row+20].height=26
    rg=s.cell(row+22,3); rg.value='='+Tx("rag.global"); rg.font=Font(name=UI,size=11,color="CFE0D9")
    foot=s.cell(row+42,3); foot.value='=""&'+Tx("msg.confidential")+'&"   ·   "&'+Tx("msg.version")+'&"   ·   "&'+Tx("msg.updated")+'&": "&TEXT(TODAY(),"yyyy-mm-dd")'
    foot.font=Font(name=UI,size=10,color="CFE0D9")
    s.row_breaks.append(Break(id=row+45)); return row+46

def render_toc(s,row):
    ANCHORS['toc']=row
    for c in range(3,16): s.cell(row,c).fill=fill(PRIMARY)
    s.row_dimensions[row].height=26
    try:
        img=XLImage(LOGO_PATH); img.width=165; img.height=46; s.add_image(img,"C"+str(row))
    except Exception: pass
    h=s.cell(row,3); h.value='="            "&'+Tx("nav.summary"); h.font=Font(name=UI,bold=True,size=15,color=WHITE); h.alignment=Alignment(vertical="center"); row+=2
    sy=s.cell(row,3); sy.value='="◈    "&'+Tx("nav.synthesis"); sy.font=Font(name=UI,size=12,bold=True,color=PRIMARY,underline="single")
    TOC_LINKS.append((sy,'synth')); s.row_dimensions[row].height=22; row+=1
    headline={"rh":"kpi.rh.headcount","lease":"kpi.lease.occ_gla","rec":"kpi.rec.ar","foot":"kpi.foot.total",
              "mkt":"kpi.mkt.nps","com":"kpi.com.total_sales","hsse":"kpi.hsse.days_no_lti","fac":"kpi.fac.sla"}
    for di,(dep,title,subs) in enumerate(DEPTS):
        key=headline[dep]; u=KMETA[key][3]; agg=KMETA[key][4]
        lk=s.cell(row,3); lk.value=('="%d.    "&'%(di+2))+Tx("dept."+dep); lk.font=Font(name=UI,size=12,color=PRIMARY,underline="single")
        TOC_LINKS.append((lk,dep))
        v=s.cell(row,9); v.value=cur_f(key,agg,scfor(u)); v.number_format=F_PCT if u=="%" else (FCM if u=="FCFA" else F_INT)
        v.font=Font(name=MONO,size=11,color=TXT_SEC); s.row_dimensions[row].height=20; row+=1
    me=s.cell(row,3); me.value='="ⓘ    "&'+Tx("nav.methodo"); me.font=Font(name=UI,size=12,color=PRIMARY,underline="single"); me.hyperlink="#"+q("M4")+"!A1"; row+=2
    s.row_breaks.append(Break(id=row)); return row+1

def render_back_cover(s,row):
    for rr in range(row,row+45):
        for c in range(3,15): s.cell(rr,c).fill=fill(PRIMARY)
    try:
        img=XLImage(LOGO_PATH); img.width=430; img.height=119; s.add_image(img,"D"+str(row+6))
    except Exception: pass
    t=s.cell(row+17,3); t.value='=IF(LANG="EN","Thank you","Merci")'; t.font=Font(name=WORD,bold=True,size=38,color=WHITE)
    s.merge_cells(start_row=row+17,start_column=3,end_row=row+17,end_column=14); s.row_dimensions[row+17].height=44
    for c in range(3,8): s.cell(row+20,c).fill=fill(WARNING)
    c1=s.cell(row+23,3); c1.value="EPLSA – Emergence Plaza · Cosmos Yopougon, Abidjan"; c1.font=Font(name=UI,size=12,color="CFE0D9")
    c2=s.cell(row+25,3); c2.value='=IF(LANG="EN","Prepared by the Centre Management — confidential.","Établi par la Direction du centre — confidentiel.")'; c2.font=Font(name=UI,size=11,color="CFE0D9")
    foot=s.cell(row+42,3); foot.value='=""&'+Tx("msg.confidential")+'&"   ·   "&'+Tx("msg.version")
    foot.font=Font(name=UI,size=10,color="CFE0D9")
    return row+45

def render_synthese(s,row):
    ANCHORS['synth']=row
    for c in range(3,16): s.cell(row,c).fill=fill(PRIMARY)
    s.row_dimensions[row].height=26
    place_icon(s,"dep_synth",f"C{row}",24)
    h=s.cell(row,3); h.value='="          1.   "&'+Tx("nav.synthesis"); h.font=Font(name=UI,bold=True,size=15,color=WHITE); h.alignment=Alignment(vertical="center")
    bl=s.cell(row,15); bl.value='="↑ "&'+Tx("nav.summary"); bl.font=Font(name=UI,size=8,color="CFE0D9",underline="single"); bl.fill=fill(PRIMARY)
    bl.hyperlink="#"+q("RPT")+"!C1"; bl.alignment=Alignment(horizontal="right",vertical="center"); row+=2
    s.cell(row,3).value='="▌ "&'+Tx("msg.hero"); s.cell(row,3).font=Font(name=UI,bold=True,size=12,color=PRIMARY); row+=1
    hero=[("kpi.com.total_sales","FCFA","SUM"),("kpi.lease.occ_gla","%","AVG"),
          ("kpi.foot.total","visites","SUM"),("kpi.rec.recovery","%","RATIO"),("kpi.lease.passing","FCFA","LAST")]
    col=3
    for key,u,agg in hero:
        for c in range(col,col+2): s.cell(row,c).fill=fill(PRIMARY)
        s.row_dimensions[row].height=5
        lab=s.cell(row+1,col); lab.value=T(key); lab.font=Font(name=UI,size=8,color=TXT_SEC); lab.fill=fill(CARD)
        v=s.cell(row+2,col); v.value=cur_f(key,agg,scfor(u)); v.number_format=F_PCT if u=="%" else (FCM if u=="FCFA" else F_INT)
        v.font=Font(name=MONO,bold=True,size=12,color=PRIMARY); v.fill=fill(CARD)
        for rr in (row+1,row+2): s.merge_cells(start_row=rr,start_column=col,end_row=rr,end_column=col+1)
        for rr in range(row,row+3):
            for c in range(col,col+2):
                s.cell(rr,c).border=Border(left=side(CARD_BORDER) if c==col else None,right=side(CARD_BORDER) if c==col+1 else None,bottom=side(CARD_BORDER) if rr==row+2 else None)
        col+=2
    row+=4
    s.cell(row,3).value='="▌ "&'+Tx("msg.heatmap"); s.cell(row,3).font=Font(name=UI,bold=True,size=12,color=PRIMARY); row+=1
    hr=row
    for j,h in enumerate(["Dépt / Dept","KPI","Achievement"]):
        c=s.cell(hr,3+j); c.value=h; c.font=Font(name=UI,bold=True,size=9); c.fill=fill(HDR_TBL); c.border=Border(bottom=side(AMBER,"medium"))
    rr=hr+1
    rep={"rh":"kpi.rh.payroll_var","lease":"kpi.lease.occ_gla","rec":"kpi.rec.recovery","foot":"kpi.foot.vs_budget",
         "mkt":"kpi.mkt.event_roi","com":"kpi.com.lfl","hsse":"kpi.hsse.action_close","fac":"kpi.fac.sla"}
    for dep,title,subs in DEPTS:
        key=rep[dep]; agg=KMETA[key][4]
        s.cell(rr,3).value=T("dept."+dep); s.cell(rr,3).font=Font(name=UI,size=9)
        s.cell(rr,4).value=T(key); s.cell(rr,4).font=Font(name=UI,size=9,color=TXT_SEC)
        v=s.cell(rr,5); v.value=cur_f(key,agg); v.number_format='0.00'; v.font=Font(name=MONO,size=9); rr+=1
    s.conditional_formatting.add(f"E{hr+1}:E{rr-1}",
        ColorScaleRule(start_type="num",start_value=0.7,start_color=R_BG,mid_type="num",mid_value=0.95,mid_color=A_BG,end_type="num",end_value=1.1,end_color=G_BG))
    PANELS.append((hr,rr-1))
    row=rr+1
    s.row_breaks.append(Break(id=row)); return row+1

def build_report_single():
    s=ws["RPT"]; s.sheet_view.showGridLines=False; s.sheet_properties.tabColor=PRIMARY
    s.column_dimensions["A"].width=2.2; s.column_dimensions["B"].width=2.2
    s.column_dimensions["C"].width=26; s.column_dimensions["D"].width=7
    for col in "EFGHIJKLMN": s.column_dimensions[col].width=8.6
    s.column_dimensions["O"].hidden=True; s.column_dimensions["P"].hidden=True
    row=1; helper=1
    row=render_front_cover(s,row)
    row=render_toc(s,row)
    row=render_synthese(s,row)
    for i,(dep,title,subs) in enumerate(DEPTS):
        row,helper=render_dept(s,dep,i+2,row,helper)
    row=render_back_cover(s,row)
    for cell,key in TOC_LINKS:
        cell.hyperlink="#"+q("RPT")+"!C"+str(ANCHORS.get(key,1))
    maxrow=row
    # tinted page background on untouched content cells (cards float in white)
    for r in range(1,maxrow+1):
        for c in range(3,15):
            cl=s.cell(r,c)
            if cl.fill.fgColor.rgb in (None,"00000000"): cl.fill=fill(PAGE_TINT)
    for p in PANELS: panelize_white(s,p[0],p[1])
    s.print_area=f"C1:N{maxrow}"
    s.page_setup.orientation="portrait"; s.page_setup.paperSize=9
    s.page_setup.fitToWidth=1; s.page_setup.fitToHeight=0
    try: s.sheet_properties.pageSetUpPr.fitToPage=True
    except Exception: pass
    s.page_margins.left=0.35; s.page_margins.right=0.35; s.page_margins.top=0.45; s.page_margins.bottom=0.55
    s.oddFooter.left.text='&"Segoe UI"&8 COSMOS · Emergence Plaza'
    s.oddFooter.center.text='&"Segoe UI"&8 Page &P / &N'
    s.oddFooter.right.text='&"Segoe UI"&8 Confidentiel'
    s.evenFooter.left.text=s.oddFooter.left.text; s.evenFooter.center.text=s.oddFooter.center.text; s.evenFooter.right.text=s.oddFooter.right.text

# ================================================================ M4 dictionary + methodo
def build_m4():
    s=ws["M4"]; s.sheet_properties.tabColor="94A3B8"; s.sheet_view.showGridLines=False
    bl=s.cell(1,3); bl.value='="↩ "&'+Tx("nav.title"); bl.font=Font(name=UI,size=9,color=LINKC,underline="single")
    bl.hyperlink="#"+q("RPT")+"!C1"
    s.cell(2,3,"M4 · Dictionnaire KPI & Méthodologie").font=Font(name=UI,bold=True,size=16,color=TXT_PRIM)
    for c in range(3,15): s.cell(3,c).fill=fill(AMBER)
    for j,h in enumerate(["Clé","FR","EN","Unité","Agg","Dépt","Budget","Sous-module H2"]):
        c=s.cell(5,3+j,h); c.font=Font(name=UI,bold=True,size=9,color=TXT_PRIM); c.fill=fill(HDR_TBL); c.border=Border(bottom=side(AMBER,"medium"))
    r=6
    for key,fr,en,u,agg,dep,h2,b,base in KPIS:
        s.cell(r,3,key).font=Font(name=MONO,size=8,color=TXT_SEC)
        s.cell(r,4,fr).font=Font(name=UI,size=9); s.cell(r,5,en).font=Font(name=UI,size=9)
        s.cell(r,6,u).font=Font(name=UI,size=9); s.cell(r,7,agg).font=Font(name=MONO,size=9)
        s.cell(r,8,dep).font=Font(name=UI,size=9); s.cell(r,9,"✓" if b else "").font=Font(name=UI,size=9)
        s.cell(r,10,subs_of(dep)[h2-1]).font=Font(name=UI,size=8,color=TXT_SEC); r+=1
    notes=[
     "MÉTHODOLOGIE & NOTES",
     "• Rapport mono-onglet « Rapport Cosmos » : couverture → synthèse → 8 départements, chacun en sous-modules (H2),",
     "  chaque sous-module = tableau KPI + faits marquants + zone commentaire ; sauts de page entre départements.",
     "• i18n : cellules =IF(LANG=\"EN\",INDEX(I18N_EN,…),INDEX(I18N_FR,…)) ; bascule LANG (FR/EN) sur M0.",
     "• Moteur temporel : bornes P_/PP_/PY_ sur M0 (IF imbriqués) ; agrégation par KPI (SUM/AVG/LAST).",
     "• RATIO : agrégé en moyenne mensuelle (v1) ; V2 = SUM(num)/SUM(den).",
     "• RAG paramétrable : tbl_param_rag (M1) — sens UP/DOWN + seuils ; ref = budget sinon N-1.",
     "• Saisie unique : chaque KPI de M3 référence sa collecte C1…C8 (cellules ambre, feuille protégée — mdp : cosmos).",
     "• Graphiques : fenêtre 12 mois glissants non volatile (INDEX/MATCH sur P_END).",
     "",
     "RACCOURCIS EXCEL 365 (Name Manager — équivalents LAMBDA, CDC §4.3 / Annexe B) :",
     "T    = LAMBDA(k, XLOOKUP(k, I18N_KEY, IF(LANG=\"EN\",I18N_EN,I18N_FR), k))",
     "PVAL = LAMBDA(row,hdr,d1,d2,agg, SWITCH(agg,\"SUM\",SUMIFS(row,hdr,\">=\"&d1,hdr,\"<=\"&d2),",
     "        \"AVG\",AVERAGEIFS(row,hdr,\">=\"&d1,hdr,\"<=\"&d2),\"LAST\",XLOOKUP(d2,hdr,row),0))",
    ]
    nr=r+2
    for line in notes:
        c=s.cell(nr,3,line)
        mono=line.startswith(("T","PVAL","        "))
        c.font=Font(name=(MONO if mono else UI), size=9, bold=(not mono and not line.startswith("•") and line!=""),
                    color=TXT_PRIM if (not mono and not line.startswith("•")) else TXT_SEC); nr+=1
    s.column_dimensions["C"].width=26; s.column_dimensions["D"].width=26; s.column_dimensions["E"].width=26; s.column_dimensions["J"].width=22
    s.freeze_panes="C6"

# ================================================================ SECTION SPEC (per TOC)
def K(dep,*ms): return ["kpi.%s.%s"%(dep,m) for m in ms]
# column kinds: t=text  n=number  f=FCFA  p=percent
SECTIONS={
 "rh":[
  ("2.1","Organigramme","Organizational Chart","org",[
     (0,"Direction du centre","Centre Management"),
     (1,"Property / Lease Manager","Property / Lease Manager"),
     (2,"Gestion locative","Leasing"),(2,"Finance & Recouvrement","Finance & Recovery"),
     (2,"Marketing","Marketing"),(2,"Sécurité / HSSE","Security / HSSE"),(2,"Facility","Facility"),
     (1,"Support & Administration","Support & Administration")]),
  ("2.2","Tableau de bord Capital Humain","Human Capital Dashboard","kpi",
     K("rh","headcount","fte_gap","open_pos","ttf","payroll","payroll_var","turnover","absence","train_hours","train_cover")),
 ],
 "lease":[
  ("3.1","Synthèse de l'immeuble","Building Summary","detail",([
     ("Élément / Item","t"),("Valeur / Value","n"),("Unité / Unit","t")],[
     ("GLA totale / Total GLA",18000,"m²"),("Unités / Units",96,"nb"),("Unités louées / Let units",86,"nb"),
     ("GLA vacante / Vacant GLA",1800,"m²"),("Niveaux / Floors",3,"nb"),("Places de parking / Parking bays",450,"nb"),
     ("Locataires actifs / Active tenants",92,"nb")])),
  ("3.2","Activité locative","Rental Trading","kpi",K("lease","occ_gla","occ_unit","passing","erv","rent_sqm")),
  ("3.3","Rent roll cible (net de charges)","Targeted Rent Roll (net of service charges)","detail",([
     ("Zone","t"),("GLA (m²)","n"),("Loyer cible / Target rent","f"),("Réalisé / Achieved","f"),("Atteinte / Achievement","p")],[
     ("RDC / Ground floor",6500,52000000,49000000,0.94),("Niveau 1 / Level 1",5800,41000000,38500000,0.94),
     ("Niveau 2 / Level 2",4200,28000000,24000000,0.857),("Food court",1000,9000000,8800000,0.978),
     ("Kiosques / Kiosks",500,4500000,3900000,0.867)])),
  ("3.4","Top 10 locataires","Top 10 Tenant Breakdown","detail",([
     ("Locataire / Tenant","t"),("Secteur / Sector","t"),("GLA (m²)","n"),("Loyer mensuel / Monthly rent","f"),("% Rent roll","p")],[
     ("Carrefour","Alimentaire / Grocery",4200,38000000,0.28),("Cinéma Pathé","Loisirs / Leisure",2100,16000000,0.12),
     ("Zara","Mode / Fashion",780,11200000,0.083),("Nike","Mode / Fashion",460,8700000,0.064),
     ("KFC","Restauration / F&B",320,7800000,0.058),("Game","Électronique / Electronics",640,7300000,0.054),
     ("Orange","Télécom",180,6200000,0.046),("Adidas","Mode / Fashion",520,9500000,0.07),
     ("Pharmacie Plaza","Santé / Health",160,3100000,0.023),("Jumia Pay","Services",90,2400000,0.018)])),
  ("3.5","Locataires prospects","Prospective Tenants","detail",([
     ("Prospect","t"),("Secteur / Sector","t"),("GLA (m²)","n"),("Valeur locative / Rental value","f"),("Étape / Stage","t"),("Proba","p")],[
     ("H&M","Mode / Fashion",900,12500000,"Négociation / Negotiation",0.6),("Burger King","F&B",300,7200000,"STP / Subject-to",0.75),
     ("Decathlon","Sport",1500,15000000,"LOI",0.4),("Yves Rocher","Beauté / Beauty",120,2600000,"STP / Subject-to",0.8),
     ("Apple Premium","Électronique",250,9000000,"Prospection / Prospecting",0.3)])),
  ("3.6","Durée résiduelle moyenne (WALE)","Weighted Average Lease Expiry","detail",([
     ("Année / Year","t"),("Baux / Leases","n"),("GLA (m²)","n"),("Loyer / Rent","f"),("% du total / of total","p")],[
     ("2026",18,3200,42000000,0.21),("2027",22,4100,55000000,0.27),("2028",15,2800,33000000,0.16),
     ("2029",20,3900,41000000,0.20),("2030+",21,3000,32000000,0.16)])),
  ("3.7","Renouvellements / sorties","Lease Renewal / Disposal Cases","detail",([
     ("Réf / Ref","t"),("Locataire / Tenant","t"),("Type","t"),("GLA (m²)","n"),("Loyer / Rent","f"),("Statut / Status","t")],[
     ("RN-001","KFC","Renewal",320,7800000,"Signé / Signed"),("RN-002","Orange","Renewal",180,6200000,"En cours / Ongoing"),
     ("DS-001","Mango","Disposal",420,0,"Sortie / Exit"),("RN-003","Game","Renewal",640,7300000,"Signé / Signed"),
     ("DS-002","Celio","Disposal",260,0,"Préavis / Notice")])),
  ("3.8","Rent roll contractuel","Contracted Rent Roll","detail",([
     ("Mois / Month","t"),("Loyer contractuel / Contracted","f"),("Charges / Service charges","f"),("Net","f")],[
     ("2026-01",132000000,18000000,114000000),("2026-02",133000000,18100000,114900000),("2026-03",134500000,18200000,116300000),
     ("2026-04",135000000,18300000,116700000),("2026-05",137000000,18400000,118600000),("2026-06",138500000,18500000,120000000)])),
  ("3.9","Rent roll projeté","Projected Rent Roll","detail",([
     ("Période / Period","t"),("Rent roll projeté / Projected","f"),("Hypothèse / Assumption","t")],[
     ("T1 2026 / Q1",132000000,"Base"),("T2 2026 / Q2",136000000,"+ New In H&M"),
     ("T3 2026 / Q3",140000000,"+ Burger King"),("T4 2026 / Q4",143000000,"Renouvellements / Renewals")])),
  ("3.10","Analyse du rent roll","Rent Roll Analysis","kpi",K("lease","ocr","reversion","vac_age","gla_vacant","retention","incentives","walt")),
 ],
 "rec":[
  ("4.1","Factures à recouvrer — DSO −30 j","Rental Invoices to Recover — DSO −30d","detail",([
     ("Facture / Invoice","t"),("Locataire / Tenant","t"),("Montant / Amount","f"),("Échéance / Due","t"),("Âge (j) / Age","n")],[
     ("INV-2601","Zara",11200000,"2026-06-15",12),("INV-2602","Nike",8700000,"2026-06-18",9),
     ("INV-2603","KFC",7800000,"2026-06-20",7),("INV-2604","Game",7300000,"2026-06-22",5),
     ("INV-2605","Orange",6200000,"2026-06-25",2)])),
  ("4.2","Factures à recouvrer — DSO +30 j","Invoices to Recover — DSO +30d","detail",([
     ("Facture / Invoice","t"),("Locataire / Tenant","t"),("Montant / Amount","f"),("Échéance / Due","t"),("Âge (j) / Age","n")],[
     ("INV-2588","Mango",9400000,"2026-04-30",47),("INV-2576","Celio",6100000,"2026-04-12",65),
     ("INV-2560","Jennyfer",4800000,"2026-03-28",80),("INV-2541","Bata",3900000,"2026-03-05",103)])),
  ("4.3","Autres revenus","Other Revenue","detail",([
     ("Source","t"),("Montant / Amount","f"),("Statut / Status","t")],[
     ("Parking",18500000,"Encaissé / Collected"),("Régie publicitaire / Media",12000000,"Encaissé / Collected"),
     ("Kiosques / Kiosks",6400000,"Partiel / Partial"),("Pop-up stores",4200000,"Encaissé / Collected"),
     ("Locations événementielles / Event hire",3100000,"En attente / Pending")])),
  ("4.4","Détail des encaissements","Details of Cash Receipts","kpi",
     K("rec","collected","billed","recovery","dso","collected_dun","promises","promise_kept","dunning","provision")),
  ("4.4.1","Pop-up — locations · occupation · revenus","Pop-up — rentals · occupancy · revenue","kpi",
     ["kpi.com.popup_count","kpi.com.kiosk_occ","kpi.com.popup_rev"]),
  ("4.4.1b","Pop-up — détail des emplacements","Pop-up — location details","detail",([
     ("Enseigne / Brand","t"),("Emplacement / Location","t"),("Mois loués / Months","n"),("Revenu / Revenue","f"),("Statut / Status","t")],[
     ("Samsung Experience","Mail central / Central",2,6500000,"Loué / Let"),("Nestlé","RDC / Ground",1,3200000,"Loué / Let"),
     ("MTN Promo","Entrée / Entrance",1,2800000,"Loué / Let"),("Glovo","Food court",2,1500000,"Loué / Let"),
     ("Emplacement libre / Vacant","RDC / Ground",0,0,"Vacant")])),
  ("4.4.2","Le Market — locations · occupation · revenus","Le Market — rentals · occupancy · revenue","kpi",
     ["kpi.com.market_stalls","kpi.com.market_occ","kpi.com.market_rev"]),
  ("4.4.2b","Le Market — kiosques d'artisans","Le Market — artisan kiosks","detail",([
     ("Kiosque / Kiosk","t"),("Artisan","t"),("Métier / Craft","t"),("Loyer / Rent","f"),("Statut / Status","t")],[
     ("K-01","Awa Couture","Couture / Tailoring",450000,"Loué / Let"),("K-02","Kossi Bois","Artisanat bois / Woodcraft",380000,"Loué / Let"),
     ("K-03","Bijoux Yas","Bijouterie / Jewelry",520000,"Loué / Let"),("K-04","Cuir & Co","Maroquinerie / Leather",410000,"Loué / Let"),
     ("K-05","Déco Akwaba","Décoration / Decor",0,"Vacant"),("K-06","Épices Mama","Épices / Spices",300000,"Loué / Let")])),
  ("4.4.3","Contentieux 2024-2025","Legal Case 2024-2025","detail",([
     ("Réf / Ref","t"),("Locataire / Tenant","t"),("Montant / Amount","f"),("Statut / Status","t")],[
     ("LC-2401","Ex-locataire A / Former tenant A",24000000,"En cours / Ongoing"),
     ("LC-2402","Ex-locataire B / Former tenant B",13500000,"Jugement / Judgment"),
     ("LC-2503","Prestataire X / Vendor X",8200000,"Médiation / Mediation")])),
  ("4.4.4","Litiges à provisionner","Litigation to be Provisioned","detail",([
     ("Réf / Ref","t"),("Locataire / Tenant","t"),("Montant / Amount","f"),("Provision proposée / Proposed","f")],[
     ("LC-2604","Locataire C / Tenant C",18000000,9000000),("LC-2605","Locataire D / Tenant D",11000000,5500000),
     ("LC-2606","Locataire E / Tenant E",7500000,3750000)])),
  ("4.4.5","Litiges provisionnés","Provisioned Litigation","detail",([
     ("Réf / Ref","t"),("Locataire / Tenant","t"),("Montant / Amount","f"),("Provision / Provision","f"),("Taux / Rate","p")],[
     ("LC-2401","Ex-locataire A",24000000,12000000,0.5),("LC-2402","Ex-locataire B",13500000,13500000,1.0),
     ("LC-2503","Prestataire X",8200000,4100000,0.5)])),
  ("4.4.6","Escomptes accordés","Discounts Granted","detail",([
     ("Locataire / Tenant","t"),("Motif / Reason","t"),("Escompte / Discount","f"),("% CA / of rent","p")],[
     ("Carrefour","Ancrage / Anchor",3800000,0.10),("Cinéma Pathé","Fréquentation / Footfall driver",1600000,0.10),
     ("Zara","Négociation / Negotiation",1100000,0.10),("Food court (global)","Lancement / Launch",900000,0.05)])),
  ("4.5","Commentaire","Commentary","note",None),
 ],
 "foot":[
  ("5","Analyse de fréquentation","Footfall Analysis","kpi",K("foot","total","vs_budget","per_sqm","conversion","dwell","peak_share","weekend_share")),
  ("5.1","Répartition par zone / entrée","Footfall by Zone / Entrance","detail",([
     ("Zone / Entrance","t"),("Visites / Visits","n"),("Part / Share","p"),("Δ vs N-1 / vs PY","p")],[
     ("Entrée principale / Main",172000,0.36,0.04),("Parking",96000,0.20,0.02),("Food court",84000,0.175,0.06),
     ("Galerie Nord / North mall",72000,0.15,-0.01),("Galerie Sud / South mall",56000,0.115,0.03)])),
 ],
 "mkt":[
  ("6.1","Réseaux sociaux","Social Media","kpi",K("mkt","followers","engagement","reach","impressions")),
  ("6.1.1","Recrutement d'abonnés par réseau","Follower Acquisition by Network","detail",([
     ("Réseau / Network","t"),("Abonnés / Followers","n"),("Nouveaux / New","n"),("Croissance / Growth","p")],[
     ("Facebook",42000,1800,0.045),("Instagram",26000,1500,0.061),("TikTok",8500,820,0.106),("LinkedIn",1500,80,0.056)])),
  ("6.2","Marketing direct","Direct Marketing","kpi",K("mkt","cpa","digital_conv","campaigns")),
  ("6.3","Analyse base clients","Customers Database Analysis","note",None),
  ("6.3.1","Profil & localisation clients","Customer Profile and Location","detail",([
     ("Segment","t"),("Part / Share","p"),("Localisation / Location","t")],[
     ("Familles / Families",0.38,"Yopougon"),("Jeunes actifs / Young pros",0.27,"Plateau / Cocody"),
     ("Étudiants / Students",0.18,"Yopougon"),("Touristes / Tourists",0.07,"Hors zone / Out of area"),
     ("Seniors",0.10,"Yopougon")])),
  ("6.3.2","Comportement & satisfaction clients","Customer Behaviour & Satisfaction","detail",([
     ("Indicateur / Metric","t"),("Valeur / Value","t")],[
     ("Panier moyen / Avg basket","18 500 FCFA"),("Fréquence visite / Visit frequency","2,3 / mois"),
     ("CSAT","84%"),("NPS","42"),("Taux de recommandation / Recommendation","79%"),
     ("Réclamations / Complaints","18 (mois / month)")])),
  ("6.4","Événements marketing & commerciaux","Marketing and Commercial Events","detail",([
     ("Événement / Event","t"),("Date","t"),("Affluence / Attendance","n"),("Coût / Cost","f"),("ROI","p")],[
     ("Fête des mères / Mothers Day","2026-05",4200,3500000,2.1),("Soldes d'été / Summer sales","2026-06",6800,5200000,2.6),
     ("Kids Festival","2026-06",3100,2800000,1.4),("Concert live","2026-06",2400,4100000,0.9)])),
  ("6.5","Budget marketing","Marketing Budget","kpi",K("mkt","spend","spend_pct","event_roi","event_sat")),
  ("6.6","Évaluation prestataires marketing","Marketing Vendor Assessment","detail",([
     ("Prestataire / Vendor","t"),("Service","t"),("Score /10","n"),("Statut / Status","t")],[
     ("Agence Pulse","Community management",8,"Conforme / OK"),("PrintCI","Impression / Print",7,"Conforme / OK"),
     ("EventPro","Événementiel / Events",6,"Vigilance / Watch"),("DataReach","Média digital / Digital",9,"Conforme / OK")])),
  ("6.7","Analyse CRM / fidélité","CRM / Loyalty Analysis","kpi",K("mkt","nps","members","member_act","basket_member")),
 ],
 "com":[
  ("7.1","Tableau de bord des ventes","Sales Dashboard","kpi",K("com","tenant_sales","total_sales","sales_sqm","lfl")),
  ("7.2","CA cumulé (YTD)","Sales Revenues Year to Date","detail",([
     ("Mois / Month","t"),("CA / Sales","f"),("Cumul / YTD","f"),("vs Budget","p")],[
     ("2026-01",1180000000,1180000000,-0.03),("2026-02",1210000000,2390000000,-0.02),("2026-03",1265000000,3655000000,0.0),
     ("2026-04",1240000000,4895000000,-0.01),("2026-05",1295000000,6190000000,0.01),("2026-06",1300000000,7490000000,0.0)])),
  ("7.3","CA par catégorie","Revenue by Products","detail",([
     ("Catégorie / Category","t"),("CA / Sales","f"),("Part / Share","p")],[
     ("Alimentaire / Grocery",420000000,0.32),("Mode / Fashion",338000000,0.26),("Restauration / F&B",234000000,0.18),
     ("Électronique / Electronics",156000000,0.12),("Loisirs / Leisure",104000000,0.08),("Services",48000000,0.04)])),
  ("7.4","Gestion commerciale","Sales Management","note",None),
  ("7.4.1","Indicateurs de vente","Sales Metrics","kpi",K("com","new_leases","leases_renewed","pipeline","pipeline_conv","discount")),
  ("7.4.2","Opportunités d'affaires","Business Opportunities","detail",([
     ("Opportunité / Opportunity","t"),("Valeur / Value","f"),("Étape / Stage","t"),("Proba","p")],[
     ("Régie pub 2027 / Media 2027",60000000,"Proposition / Proposal",0.5),("Naming food court",35000000,"Négociation / Negotiation",0.6),
     ("Sponsor concert / Concert sponsor",18000000,"Qualifié / Qualified",0.4),("Partenariat banque / Bank partner",24000000,"STP",0.7)])),
  ("7.4.3","Pop-up & tendances marché","Sales Pop-up and Market Trends","kpi",
     K("com","popup_count","popup_rev","kiosk_occ","market_rev","market_occ","market_stalls","media_sales","partnership_rev","activations","activation_rev")),
 ],
 "hsse":[
  ("8.1","Journal des événements / dangers","Events Hazard Logbook","detail",([
     ("Réf / Ref","t"),("Date","t"),("Type","t"),("Lieu / Location","t"),("Gravité / Severity","t")],[
     ("EV-2601","2026-06-03","Chute / Slip","Parking","Mineur / Minor"),("EV-2602","2026-06-09","Incendie évité / Near-fire","Local technique / Plant room","Majeur / Major"),
     ("EV-2603","2026-06-14","Malaise / Faintness","Food court","Mineur / Minor"),("EV-2604","2026-06-21","Vol / Theft","Galerie / Mall","Modéré / Moderate")])),
  ("8.2","Indicateurs événements / dangers","Events / Hazard Metrics","kpi",K("hsse","incidents","incident_sev","nearmiss","days_no_lti")),
  ("8.2.1","Analyse par poste de travail","Analysis by Workstation","detail",([
     ("Poste / Workstation","t"),("Événements / Events","n"),("Jours perdus / Lost days","n")],[
     ("Sécurité / Security",4,2),("Maintenance",3,5),("Nettoyage / Cleaning",2,1),("Accueil / Reception",1,0)])),
  ("8.2.2","Analyse par type d'incident","Analysis by Incident Type","detail",([
     ("Type d'incident / Incident type","t"),("Nombre / Count","n"),("Part / Share","p")],[
     ("Chute / Slip",4,0.4),("Vol / Theft",3,0.3),("Malaise / Medical",2,0.2),("Incendie / Fire",1,0.1)])),
  ("8.2.3","Analyse par jours perdus","Analysis by Lost Working Days","detail",([
     ("Mois / Month","t"),("Jours perdus / Lost days","n")],[
     ("2026-01",1),("2026-02",0),("2026-03",3),("2026-04",2),("2026-05",0),("2026-06",2)])),
  ("8.2.4","Analyse par parties prenantes","Analysis by Stakeholders","detail",([
     ("Partie prenante / Stakeholder","t"),("Événements / Events","n")],[
     ("Visiteurs / Visitors",5),("Personnel / Staff",3),("Prestataires / Vendors",2),("Locataires / Tenants",1)])),
  ("8.2.5","Analyse du taux de fréquence","Frequency Rate Analysis","kpi",K("hsse","tf","tg")),
  ("8.3","Réunions toolbox","Toolbox Meeting","detail",([
     ("Date","t"),("Thème / Topic","t"),("Participants","n")],[
     ("2026-06-02","Évacuation / Evacuation",22),("2026-06-12","EPI / PPE",18),("2026-06-23","Premiers secours / First aid",20)])),
  ("8.4","Suivi des actions sûreté","Security Action Tracking","note",None),
  ("8.4.1","Actions en cours","Current Action Items","detail",([
     ("Réf / Ref","t"),("Action","t"),("Responsable / Owner","t"),("Échéance / Due","t"),("Statut / Status","t")],[
     ("AC-01","Renforcer CCTV parking / Add CCTV","Sûreté / Security","2026-07-15","En cours / Ongoing"),
     ("AC-02","Audit extincteurs / Extinguisher audit","HSSE","2026-07-01","En cours / Ongoing"),
     ("AC-03","Formation évacuation / Evac training","HSSE","2026-07-20","Planifié / Planned")])),
  ("8.4.2","Actions à venir","Upcoming Action Items","detail",([
     ("Réf / Ref","t"),("Action","t"),("Responsable / Owner","t"),("Échéance / Due","t"),("Statut / Status","t")],[
     ("AC-04","Exercice incendie T3 / Q3 fire drill","HSSE","2026-08-10","À venir / Upcoming"),
     ("AC-05","Revue plan de crise / Crisis plan review","Direction / Mgmt","2026-09-01","À venir / Upcoming")])),
  ("8.5","Suivi budgétaire & sûreté","Budget & Security Tracking","kpi",
     K("hsse","guards","sec_interv","theft","cctv","fire_conform","evac_time","evac_drills","fire_alarms","risks_id","risks_treated","action_close","audits","train_cover")),
  ("8.6","Évaluation prestataires","Vendors Assessment","detail",([
     ("Prestataire / Vendor","t"),("Service","t"),("SLA","p"),("Score /10","n"),("Statut / Status","t")],[
     ("SecuGuard","Gardiennage / Guarding",0.94,8,"Conforme / OK"),("FireSafe","Incendie / Fire",0.97,9,"Conforme / OK"),
     ("SafeMed","Secours / Medical",0.90,7,"Vigilance / Watch")])),
 ],
 "fac":[
  ("9.1","Relevés de rondes (walkthrough)","Walkthrough Findings Records","detail",([
     ("Réf / Ref","t"),("Zone","t"),("Constat / Finding","t"),("Priorité / Priority","t"),("Statut / Status","t")],[
     ("WT-01","Toiture / Roof","Infiltration / Leak","Haute / High","En cours / Ongoing"),
     ("WT-02","CVC / HVAC","Filtre à changer / Filter due","Moyenne / Medium","Planifié / Planned"),
     ("WT-03","Sanitaires / Restrooms","Robinetterie / Faucet","Basse / Low","Clôturé / Closed"),
     ("WT-04","Parking","Éclairage / Lighting","Moyenne / Medium","En cours / Ongoing")])),
  ("9.2","Gestion des utilities","Utilities Management","note",None),
  ("9.2.1","Analyse conso électricité","Electricity Consumption Analysis","detail",([
     ("Mois / Month","t"),("Conso (kWh)","n"),("Coût / Cost","f"),("vs Budget","p")],[
     ("2026-01",178000,23800000,-0.02),("2026-02",172000,22900000,-0.04),("2026-03",185000,24700000,0.01),
     ("2026-04",181000,24100000,-0.01),("2026-05",190000,25300000,0.03),("2026-06",186000,24800000,0.02)])),
  ("9.2.2","Analyse conso eau de ville","City Water Consumption Analysis","detail",([
     ("Mois / Month","t"),("Conso (m³)","n"),("Coût / Cost","f"),("vs Budget","p")],[
     ("2026-01",3100,3700000,-0.03),("2026-02",2950,3520000,-0.06),("2026-03",3250,3880000,0.02),
     ("2026-04",3180,3800000,0.0),("2026-05",3320,3960000,0.04),("2026-06",3210,3830000,0.01)])),
  ("9.2.3","Analyse conso carburant (groupe)","Generator Fuel Consumption","detail",([
     ("Mois / Month","t"),("Conso (L)","n"),("Coût / Cost","f"),("Heures groupe / Genset hrs","n")],[
     ("2026-01",4200,5040000,38),("2026-02",3900,4680000,34),("2026-03",4800,5760000,46),
     ("2026-04",4500,5400000,41),("2026-05",5200,6240000,52),("2026-06",4600,5520000,44)])),
  ("9.2.4","Plan de maintenance préventive","Preventive Maintenance Schedule","detail",([
     ("Équipement / Equipment","t"),("Fréquence / Frequency","t"),("Prochaine / Next","t"),("Statut / Status","t")],[
     ("CVC / HVAC","Mensuelle / Monthly","2026-07-05","À jour / Up to date"),("Ascenseurs / Lifts","Mensuelle / Monthly","2026-07-08","À jour / Up to date"),
     ("Groupe électrogène / Genset","Trimestrielle / Quarterly","2026-09-01","Planifié / Planned"),("Incendie / Fire system","Trimestrielle / Quarterly","2026-08-15","Planifié / Planned")])),
  ("9.2.5","Historique de maintenance","Maintenance History","detail",([
     ("Date","t"),("Équipement / Equipment","t"),("Intervention","t"),("Coût / Cost","f")],[
     ("2026-06-04","CVC / HVAC","Remplacement filtre / Filter",850000),("2026-06-11","Ascenseur 2 / Lift 2","Réparation porte / Door",1200000),
     ("2026-06-19","Plomberie / Plumbing","Fuite RDC / Leak",420000),("2026-06-27","Électricité / Electrical","Tableau / Switchboard",680000)])),
  ("9.2.6","Indicateurs maintenance & énergie","Maintenance & Energy KPI","kpi",
     K("fac","prev_corr","wo_prev","wo_corr","ot_backlog","mttr","ot_opened","ot_closed","energy_int","util_cost","consumption")),
  ("9.3","Gestion des déchets","Waste Management","detail",([
     ("Mois / Month","t"),("Déchets (t) / Waste","n"),("Recyclé / Recycled","p"),("Coût / Cost","f")],[
     ("2026-01",40,0.31,2500000),("2026-02",38,0.33,2400000),("2026-03",44,0.35,2700000),
     ("2026-04",41,0.34,2550000),("2026-05",46,0.37,2800000),("2026-06",42,0.36,2600000)])),
  ("9.4","Actions facility","Facility Action Items","detail",([
     ("Réf / Ref","t"),("Action","t"),("Responsable / Owner","t"),("Échéance / Due","t"),("Statut / Status","t")],[
     ("FA-01","Étanchéité toiture / Roof sealing","Facility","2026-07-30","En cours / Ongoing"),
     ("FA-02","Relamping LED parking","Facility","2026-08-05","Planifié / Planned"),
     ("FA-03","Contrat nettoyage / Cleaning contract","Achats / Procurement","2026-07-12","En cours / Ongoing")])),
  ("9.5","Évaluation prestataires","Vendors Assessment","detail",([
     ("Prestataire / Vendor","t"),("Service","t"),("SLA","p"),("Score /10","n"),("Statut / Status","t")],[
     ("CleanPro","Nettoyage / Cleaning",0.93,8,"Conforme / OK"),("CoolTech","CVC / HVAC",0.95,9,"Conforme / OK"),
     ("LiftCare","Ascenseurs / Lifts",0.91,7,"Vigilance / Watch"),("EcoWaste","Déchets / Waste",0.88,6,"Vigilance / Watch")])),
  ("9.6","Suivi budgétaire","Budget Tracking","kpi",K("fac","capex","capex_var","capex_proj","inspections","insp_conform","nonconform","cleanliness","hygiene_audits","clean_complaints","sla")),
 ],
}

# Budget vs Réalisé — une sous-section par département (combo réel/budget + table)
_BNUM={"rh":"2.3","lease":"3.11","rec":"4.7","foot":"5.2","mkt":"6.8","com":"7.5","hsse":"8.7","fac":"9.7"}
for _dep in SECTIONS:
    SECTIONS[_dep].append((_BNUM[_dep],"Budget vs Réalisé","Budget vs Actual","budget",_dep))

# ================================================================ build
build_m2(); build_m0(); build_m1(); build_m3()
for dep,code in DEPT_COLLECTE.items(): build_collecte(code,dep)
build_report_single()
build_m4()

ws["M1"].sheet_state="hidden"; ws["M2"].sheet_state="hidden"; ws["M3"].sheet_state="hidden"; ws["M5"].sheet_state="hidden"
order=["RPT","M0","C1","C2","C3","C4","C5","C6","C7","C8","M4","M1","M2","M3","M5"]
wb._sheets.sort(key=lambda sh:order.index([k for k,v in SH.items() if v==sh.title][0]))
wb.active=0
wb.save("Cosmos-Report-Builder-v3.0.xlsx")
print("SAVED. sheets:",len(wb.sheetnames),"| KPIs:",len(KPIS))
