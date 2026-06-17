import json, os, subprocess, sys
from pathlib import Path
from openpyxl import load_workbook

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
MACRO_DIR = os.path.expandvars(r"%APPDATA%\LibreOffice\4\user\basic\Standard")
MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave()
  ThisComponent.calculateAll()
  ThisComponent.store()
  ThisComponent.close(True)
End Sub
</script:module>"""

def setup():
    if not os.path.exists(MACRO_DIR):
        subprocess.run([SOFFICE,"--headless","--terminate_after_init"],capture_output=True,timeout=60)
        os.makedirs(MACRO_DIR, exist_ok=True)
    Path(os.path.join(MACRO_DIR,"Module1.xba")).write_text(MACRO, encoding="utf-8")

def recalc(fn, timeout=120):
    abs_path=str(Path(fn).absolute())
    setup()
    cmd=[SOFFICE,"--headless","--norestore",
         "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
         abs_path]
    r=subprocess.run(cmd,capture_output=True,text=True,timeout=timeout)
    errs=["#VALUE!","#DIV/0!","#REF!","#NAME?","#NULL!","#NUM!","#N/A"]
    wb=load_workbook(fn,data_only=True)
    details={e:[] for e in errs}; total=0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                if isinstance(c.value,str):
                    for e in errs:
                        if e in c.value:
                            details[e].append(f"{sn}!{c.coordinate}"); total+=1; break
    wb.close()
    out={"status":"success" if total==0 else "errors_found","total_errors":total,"error_summary":{}}
    for e,l in details.items():
        if l: out["error_summary"][e]={"count":len(l),"locations":l[:25]}
    return out

if __name__=="__main__":
    fn=sys.argv[1]; print(json.dumps(recalc(fn, int(sys.argv[2]) if len(sys.argv)>2 else 120),indent=2))
